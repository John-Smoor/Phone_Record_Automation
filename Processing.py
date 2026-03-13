# import os
# import pandas as pd
# import re
# from datetime import datetime, date
# import openpyxl
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from openpyxl.formatting.rule import ColorScaleRule
# from openpyxl.utils import get_column_letter

# # ── Branch → City / Format mapping ──────────────────────────────────────────
# mapping_dict = {
#     "1 MG": {"City": "Bangalore", "Format": "Café"},
#     "1 Mg Cafe": {"City": "Bangalore", "Format": "Café"},
#     "1mg cafe": {"City": "Bangalore", "Format": "Café"},
#     "Adarsh Palm Retreat": {"City": "Bangalore", "Format": "Café"},
#     "Arumbakkam Cloud Kitchen": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "arumbakkam ck": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Avadi CH": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Bandra": {"City": "Mumbai", "Format": "Café"},
#     "Banni Square": {"City": "Delhi NCR", "Format": "Kiosk"},
#     "Basaveshwara Nagar CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "BEL Road": {"City": "Bangalore", "Format": "Café"},
#     "Bellandur": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Bhartiya City Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "BLR Airport T2 International Bakery": {"City": "Bangalore", "Format": "Airport"},
#     "BLR Airport Terminal 1": {"City": "Bangalore", "Format": "Airport"},
#     "Central Arcade DLF": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "dlf": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Chennai ECR": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Cummulus": {"City": "Bangalore", "Format": "Café"},
#     "Dwarka Cloud Kitchen": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "dwarka ck": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Electronic City": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Express Avenue SMOOR": {"City": "Chennai", "Format": "Kiosk"},
#     "express avenue": {"City": "Chennai", "Format": "Kiosk"},
#     "Forum Shantiniketan": {"City": "Bangalore", "Format": "Kiosk"},
#     "Forum South Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "Garuda Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "Godrej One": {"City": "Mumbai", "Format": "Kiosk"},
#     "Grant Road CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Head Office": {"City": "Bangalore", "Format": "NA"},
#     "Hesaraghatta": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Hinjewadi": {"City": "Pune", "Format": "Cloud Kitchen"},
#     "HSR": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "IC Colony CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Indiranagar": {"City": "Bangalore", "Format": "Signature"},
#     "Infinity Mall Malad": {"City": "Mumbai", "Format": "Kiosk"},
#     "J.P Nagar": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "j.p nagar": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Jayanagar": {"City": "Bangalore", "Format": "Café"},
#     "Kadugodi CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Kalkaji CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Kamanahalli": {"City": "Bangalore", "Format": "Café"},
#     "kammanahalli": {"City": "Bangalore", "Format": "Café"},
#     "Kharghar CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Koramangala": {"City": "Bangalore", "Format": "Café"},
#     "Koreagaon Park": {"City": "Pune", "Format": "Signature"},
#     "Lavelle Road": {"City": "Bangalore", "Format": "Signature"},
#     "Laxmi Nagar": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Magarpatta": {"City": "Pune", "Format": "Cloud Kitchen"},
#     "Mahadevpura": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Malad": {"City": "Mumbai", "Format": "Kiosk"},
#     "MAP Smoor": {"City": "Bangalore", "Format": "Café"},
#     "Marathahalli CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Medavakkam TN": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Mira Road CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Mukherjee Nagar": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Netaji Shubhash Palace": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "nsp": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Nexus Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "Noida Sec 46": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Noida Sec 83": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Orion Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "Perungudi CK": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Phoenix Mall Pune": {"City": "Pune", "Format": "Kiosk"},
#     "Powai": {"City": "Mumbai", "Format": "Signature"},
#     "Rajouri Garden": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "rajouri": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Ramapuram TN": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "ramapuram": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "RT Nagar CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Sadashivanagar": {"City": "Bangalore", "Format": "Café"},
#     "Sakinaka CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Sanpada": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Sarjapur CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Seawoods Mall": {"City": "Mumbai", "Format": "Café"},
#     "Smoor - Defence Colony CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "defence colony ck": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Test.Rebel": {"City": "NA", "Format": "NA"},
#     "Udyog Vihar CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Undri": {"City": "Pune", "Format": "Cloud Kitchen"},
#     "Uttarahalli CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Vandalur": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Vashi Smoor": {"City": "Mumbai", "Format": "Kiosk"},
#     "Vasundra CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Vega City": {"City": "Bangalore", "Format": "Kiosk"},
#     "Versova": {"City": "Mumbai", "Format": "Café"},
#     "Viviana Lounge Mumbai": {"City": "Mumbai", "Format": "Kiosk"},
#     "viviana": {"City": "Mumbai", "Format": "Kiosk"},
#     "West Tambaram": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Whitefield": {"City": "Bangalore", "Format": "Signature"},
#     "Whitefield Lounge": {"City": "Bangalore", "Format": "Signature"},
#     "Worli": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Yelahanka Cloud Kitchen": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "yelahanka": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Kandivali CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Airoli CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Ambience Mall": {"City": "Delhi NCR", "Format": "Kiosk"},
#     "Bhandup CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Green Park CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Gurugram sector 31 Cloud Kitchen": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Malviya Nagar CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "RR Nagar CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Rohini CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Henjewadi": {"City": "Pune", "Format": "Cloud Kitchen"},
#     "Hesarghatta": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "hessaraghatta": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Embassy Lake": {"City": "Bangalore", "Format": "Café"},
#     "Chembur East CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Hesaraghatta Outlet": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Mahakali CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Noida Sector 46": {"City": "NCR", "Format": "Cloud Kitchen"},
#     "Noida Sector 83": {"City": "NCR", "Format": "Cloud Kitchen"},
#     "Velachery": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Smoor Deloitte": {"City": "Bangalore", "Format": "Kiosk"},
#     "Zirakpur": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "RBI Officers Canteen": {"City": "Bangalore", "Format": "Kiosk"},
# }

# # ── Phone cleaning ────────────────────────────────────────────────────────────
# def clean_phone(raw):
#     if pd.isna(raw):
#         return None
#     phone = str(raw).strip()
#     phone = re.sub(r'[\s\-\(\)\+\.]', '', phone)
#     phone = re.sub(r'\D', '', phone)
#     if not phone:
#         return None
#     if len(phone) == 12 and phone.startswith('91'):
#         phone = phone[2:]
#     if len(phone) == 11 and phone.startswith('0'):
#         phone = phone[1:]
#     if len(phone) != 10:
#         return None
#     for digit in set(phone):
#         if phone.count(digit) > 5:
#             return None
#     if phone[0] in ('0', '1'):
#         return None
#     if phone[0] not in ('6', '7', '8', '9'):
#         return None
#     return phone


# # ── Main analysis ─────────────────────────────────────────────────────────────
# def build_report(input_csv: str, output_xlsx: str):

#     # 1. Load & filter
#     df = pd.read_csv(input_csv, dtype=str, low_memory=False)
#     df = df[df['channel'].str.strip().str.lower().isin(['dine in', 'take away', 'dine-in', 'takeaway'])]

#     # 2. Clean phones
#     df['clean_phone'] = df['customer_phoneNumber'].apply(clean_phone)

#     # 3. Deduplicate at invoice level
#     inv = df.drop_duplicates(subset='invoiceNumber')[
#         ['invoiceNumber', 'branchName', 'channel', 'date', 'clean_phone']
#     ].copy()
#     inv['date'] = pd.to_datetime(inv['date'], errors='coerce').dt.date

#     # 4. Add City / Format via mapping
#     mapping_df = pd.DataFrame.from_dict(mapping_dict, orient='index').reset_index()
#     mapping_df.columns = ['branchName', 'City', 'Format']
#     inv = inv.merge(mapping_df, on='branchName', how='left')
#     inv['City']   = inv['City'].fillna('Unknown')
#     inv['Format'] = inv['Format'].fillna('Unknown')
#     inv = inv[~inv['City'].isin(['Unknown', 'NA'])]

#     # ── EXCLUDE Cloud Kitchen outlets ────────────────────────────────────────
#     inv = inv[inv['Format'] != 'Cloud Kitchen']

#     # 5. Week / day buckets
#     inv['date_dt']  = pd.to_datetime(inv['date'], errors='coerce')
#     inv['iso_week'] = inv['date_dt'].dt.isocalendar().week.astype(int)
#     inv['iso_year'] = inv['date_dt'].dt.isocalendar().year.astype(int)
#     inv['yw']       = inv['iso_year'] * 100 + inv['iso_week']

#     last3_yw          = sorted(inv['yw'].dropna().unique())[-3:]
#     target_week_pairs = [(int(yw // 100), int(yw % 100)) for yw in last3_yw]

#     all_days    = sorted(inv['date'].dropna().unique())
#     target_days = all_days[-3:] if len(all_days) >= 3 else all_days

#     # 6. Per-branch metrics
#     order_counts = inv.groupby('branchName')['invoiceNumber'].nunique().reset_index()
#     order_counts.columns = ['branchName', 'total_orders']

#     branches = (
#         inv[['branchName', 'City', 'Format']]
#         .drop_duplicates('branchName')
#         .merge(order_counts, on='branchName', how='left')
#         .sort_values('total_orders', ascending=False)
#         .reset_index(drop=True)
#     )

#     def capture(sub):
#         total     = sub['invoiceNumber'].nunique()
#         has_phone = sub[sub['clean_phone'].notna()]['invoiceNumber'].nunique()
#         return has_phone / total if total > 0 else None

#     day_fmt   = lambda d: d.strftime('%b %#d') if os.name == 'nt' else d.strftime('%b %-d')
#     latest_yw = last3_yw[-1] if last3_yw else None

#     # Week cols
#     week_cols = []
#     for (yr, wk) in target_week_pairs:
#         yw_key = yr * 100 + wk
#         col    = f'W{wk} WTD' if yw_key == latest_yw else f'W{wk}'
#         week_cols.append(col)
#         rates = []
#         for _, row in branches.iterrows():
#             sub = inv[(inv['branchName'] == row['branchName']) &
#                       (inv['iso_week'] == wk) & (inv['iso_year'] == yr)]
#             rates.append(capture(sub))
#         branches[col] = rates

#     # Day cols
#     day_cols = []
#     for d in target_days:
#         col = day_fmt(d)
#         day_cols.append(col)
#         rates = []
#         for _, row in branches.iterrows():
#             sub = inv[(inv['branchName'] == row['branchName']) & (inv['date'] == d)]
#             rates.append(capture(sub))
#         branches[col] = rates

#     metric_cols = week_cols + day_cols

#     # 7. Build Excel ────────────────────────────────────────────────────────────
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Phone No Record Tracker"

#     ROYAL_BLUE = "1A3A6B"
#     HDR_WEEK   = "2C5282"
#     HDR_DAY    = "1A3A6B"
#     WHITE      = "FFFFFF"
#     LIGHT_ROW  = "F7F9FC"
#     ALT_ROW    = "EEF2F9"
#     BORDER_C   = "C5CFE3"

#     thin = Side(style='thin', color=BORDER_C)
#     bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

#     total_cols      = 3 + len(metric_cols)
#     last_col_letter = get_column_letter(total_cols)

#     # Title row
#     ws.merge_cells(f'A1:{last_col_letter}1')
#     c = ws['A1']
#     c.value     = "Phone No Record Tracker  |  Dine In & Take Away"
#     c.font      = Font(name='Arial', bold=True, size=14, color=WHITE)
#     c.fill      = PatternFill('solid', start_color=ROYAL_BLUE)
#     c.alignment = Alignment(horizontal='center', vertical='center')
#     ws.row_dimensions[1].height = 30

#     ws.merge_cells(f'A2:{last_col_letter}2')
#     ws['A2'].fill = PatternFill('solid', start_color=ROYAL_BLUE)
#     ws.row_dimensions[2].height = 6

#     # Group header row (row 3)
#     for c_idx in [1, 2, 3]:
#         ws.cell(row=3, column=c_idx).fill = PatternFill('solid', start_color=ROYAL_BLUE)

#     ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=6)
#     wk_cell = ws.cell(row=3, column=4, value="← Last 3 Weeks →")
#     wk_cell.font      = Font(name='Arial', bold=True, size=9, color=WHITE)
#     wk_cell.fill      = PatternFill('solid', start_color=HDR_WEEK)
#     wk_cell.alignment = Alignment(horizontal='center')

#     ws.merge_cells(start_row=3, start_column=7, end_row=3, end_column=9)
#     dy_cell = ws.cell(row=3, column=7, value="← Last 3 Days →")
#     dy_cell.font      = Font(name='Arial', bold=True, size=9, color=WHITE)
#     dy_cell.fill      = PatternFill('solid', start_color=HDR_DAY)
#     dy_cell.alignment = Alignment(horizontal='center')

#     # Column label row (row 4)
#     headers  = ['City', 'Branch', 'Format'] + metric_cols
#     col_meta = {}
#     for h in ['City', 'Branch', 'Format']:
#         col_meta[h] = (ROYAL_BLUE, WHITE)
#     for c in week_cols:
#         col_meta[c] = (HDR_WEEK, WHITE)
#     for c in day_cols:
#         col_meta[c] = (HDR_DAY, WHITE)

#     for ci, h in enumerate(headers, start=1):
#         cell = ws.cell(row=4, column=ci, value=h)
#         bg, fg = col_meta.get(h, (ROYAL_BLUE, WHITE))
#         cell.fill      = PatternFill('solid', start_color=bg)
#         cell.font      = Font(name='Arial', bold=True, size=10, color=fg)
#         cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
#         cell.border    = bdr
#     ws.row_dimensions[4].height = 28

#     # Data rows
#     for ri, row in branches.iterrows():
#         excel_row = ri + 5
#         fill_bg   = LIGHT_ROW if ri % 2 == 0 else ALT_ROW
#         vals = [row['City'], row['branchName'], row['Format']] + [row[c] for c in metric_cols]
#         for ci, val in enumerate(vals, start=1):
#             cell        = ws.cell(row=excel_row, column=ci)
#             cell.border = bdr
#             if ci <= 3:
#                 cell.value     = val
#                 cell.font      = Font(name='Arial', size=9)
#                 cell.fill      = PatternFill('solid', start_color=fill_bg)
#                 cell.alignment = Alignment(horizontal='left', vertical='center')
#             else:
#                 if val is None:
#                     cell.value     = "—"
#                     cell.font      = Font(name='Arial', size=9, color='BBBBBB')
#                     cell.fill      = PatternFill('solid', start_color=fill_bg)
#                     cell.alignment = Alignment(horizontal='center')
#                 else:
#                     cell.value         = val
#                     cell.number_format = '0.0%'
#                     cell.font          = Font(name='Arial', size=9)
#                     cell.fill          = PatternFill('solid', start_color=fill_bg)
#                     cell.alignment     = Alignment(horizontal='center')

#     total_data_rows = len(branches)
#     last_row        = 4 + total_data_rows

#     # Conditional formatting: ONLY on day cols
#     day_col_start = 3 + len(week_cols) + 1
#     for ci_offset in range(len(day_cols)):
#         ci         = day_col_start + ci_offset
#         col_letter = get_column_letter(ci)
#         ws.conditional_formatting.add(
#             f"{col_letter}5:{col_letter}{last_row}",
#             ColorScaleRule(
#                 start_type='min', start_color='FFFFFF',
#                 end_type='max',   end_color='C0392B'
#             )
#         )

#     col_widths = [14, 28, 14, 9, 9, 9, 9, 9, 9]
#     for ci, w in enumerate(col_widths, start=1):
#         ws.column_dimensions[get_column_letter(ci)].width = w

#     ws.freeze_panes = 'D5'

#     # 8. Raw Dump sheet ─────────────────────────────────────────────────────────
#     ws2 = wb.create_sheet("Raw Dump")

#     extra_cols_wanted = ['invoiceDate', 'customer_name', 'customer_email',
#                          'customer_id', 'orderNumber', 'netAmount', 'grossAmount']
#     available_extra = [c for c in extra_cols_wanted if c in df.columns]

#     df['_qty'] = pd.to_numeric(df.get('item_quantity', 0), errors='coerce').fillna(0)
#     sale_qty   = df.groupby('invoiceNumber')['_qty'].sum().reset_index()
#     sale_qty.columns = ['invoiceNumber', 'sale_qty']

#     raw = df.drop_duplicates(subset='invoiceNumber')[
#         ['invoiceNumber', 'branchName', 'channel', 'date', 'clean_phone'] + available_extra
#     ].copy()
#     raw = raw.merge(mapping_df[['branchName', 'City', 'Format']], on='branchName', how='left')
#     raw = raw.merge(sale_qty, on='invoiceNumber', how='left')
#     raw['City']   = raw['City'].fillna('Unknown')
#     raw['Format'] = raw['Format'].fillna('Unknown')
#     raw['date']   = pd.to_datetime(raw['date'], errors='coerce').dt.date

#     # ── EXCLUDE Cloud Kitchen from Raw Dump too ───────────────────────────────
#     raw = raw[raw['Format'] != 'Cloud Kitchen']

#     for nc in ['netAmount', 'grossAmount']:
#         if nc in raw.columns:
#             raw[nc] = pd.to_numeric(raw[nc], errors='coerce')

#     raw_dump = (
#         raw[raw['clean_phone'].notna()]
#         .merge(order_counts, on='branchName', how='left')
#         .sort_values(['total_orders', 'date', 'branchName'], ascending=[False, True, True])
#         .drop(columns=['total_orders'])
#         .reset_index(drop=True)
#     )

#     ordered_cols = ['invoiceNumber', 'date']
#     for c in ['invoiceDate', 'branchName', 'City', 'Format', 'channel',
#               'customer_name', 'clean_phone', 'customer_email', 'customer_id',
#               'orderNumber', 'sale_qty', 'netAmount', 'grossAmount']:
#         if c in raw_dump.columns:
#             ordered_cols.append(c)
#     raw_dump = raw_dump[ordered_cols]

#     header_labels = {
#         'invoiceNumber': 'Invoice No', 'date': 'Date', 'invoiceDate': 'Invoice Date',
#         'branchName': 'Branch', 'City': 'City', 'Format': 'Format', 'channel': 'Channel',
#         'customer_name': 'Customer Name', 'clean_phone': 'Phone Number',
#         'customer_email': 'Email', 'customer_id': 'Customer ID', 'orderNumber': 'Order No',
#         'sale_qty': 'Sale Qty', 'netAmount': 'Net Amount', 'grossAmount': 'Gross Amount',
#     }

#     numeric_cols_raw = {'netAmount', 'grossAmount', 'sale_qty'}
#     left_cols        = {'branchName', 'customer_name', 'customer_email'}
#     hdr_fill  = PatternFill('solid', start_color="1A3A6B")
#     hdr_font  = Font(name='Arial', bold=True, size=10, color="FFFFFF")
#     hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

#     for ci, col in enumerate(ordered_cols, start=1):
#         cell = ws2.cell(row=1, column=ci, value=header_labels.get(col, col))
#         cell.fill = hdr_fill; cell.font = hdr_font
#         cell.alignment = hdr_align; cell.border = bdr
#     ws2.row_dimensions[1].height = 28

#     for ri, row_data in raw_dump.iterrows():
#         excel_row = ri + 2
#         fill_bg   = "F9FAFB" if ri % 2 == 0 else "FFFFFF"
#         for ci, col in enumerate(ordered_cols, start=1):
#             val  = row_data[col]
#             cell = ws2.cell(row=excel_row, column=ci)
#             cell.fill   = PatternFill('solid', start_color=fill_bg)
#             cell.font   = Font(name='Arial', size=9)
#             cell.border = bdr
#             if col in numeric_cols_raw:
#                 try:
#                     cell.value = float(val) if val is not None and str(val) != 'nan' else None
#                 except (ValueError, TypeError):
#                     cell.value = None
#                 cell.number_format = '#,##0.00' if col in {'netAmount', 'grossAmount'} else '#,##0'
#                 cell.alignment = Alignment(horizontal='right', vertical='center')
#             elif col == 'clean_phone':
#                 cell.value         = str(val) if val is not None else None
#                 cell.number_format = '@'
#                 cell.alignment     = Alignment(horizontal='center', vertical='center')
#             else:
#                 cell.value     = val
#                 cell.alignment = Alignment(
#                     horizontal='left' if col in left_cols else 'center',
#                     vertical='center'
#                 )

#     col_widths2 = {
#         'invoiceNumber': 16, 'date': 12, 'invoiceDate': 14, 'branchName': 26,
#         'City': 12, 'Format': 14, 'channel': 12, 'customer_name': 20,
#         'clean_phone': 15, 'customer_email': 26, 'customer_id': 14,
#         'orderNumber': 14, 'sale_qty': 10, 'netAmount': 14, 'grossAmount': 14,
#     }
#     for ci, col in enumerate(ordered_cols, start=1):
#         ws2.column_dimensions[get_column_letter(ci)].width = col_widths2.get(col, 14)

#     ws2.freeze_panes = 'A2'

#     wb.save(output_xlsx)
#     print(f"✅ Report saved → {output_xlsx}")
#     print(f"   Branches analysed             : {len(branches)}")
#     print(f"   Total invoices (Dine In / TA) : {len(inv)}")
#     print(f"   Raw Dump rows (with phone)    : {len(raw_dump)}")
#     phone_fill_pct = (inv['clean_phone'].notna().sum() / len(inv) * 100) if len(inv) else 0
#     print(f"   Overall phone capture         : {phone_fill_pct:.1f}%")


# # ── Entry point ───────────────────────────────────────────────────────────────
# if __name__ == "__main__":
#     import sys

#     ROLLING_CSV = r"C:/Users/jeryy/OneDrive - BLISS CHOCOLATES INDIA PRIVATE LIMITED/Desktop/Documents/Phone No/Data/sales_rolling.csv"
#     OUTPUT_DIR  = r"C:/Users/jeryy/OneDrive - BLISS CHOCOLATES INDIA PRIVATE LIMITED/Desktop/Documents/Phone No/output"

#     input_file = sys.argv[1] if len(sys.argv) > 1 else ROLLING_CSV
#     if len(sys.argv) > 2:
#         output_file = sys.argv[2]
#     else:
#         os.makedirs(OUTPUT_DIR, exist_ok=True)
#         date_str    = datetime.today().strftime("%Y-%m-%d")
#         output_file = os.path.join(OUTPUT_DIR, f"Phone No Record - {date_str}.xlsx")

#     build_report(input_file, output_file)




# import os
# import pandas as pd
# import re
# from datetime import datetime, date
# import openpyxl
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from openpyxl.formatting.rule import ColorScaleRule
# from openpyxl.utils import get_column_letter

# # ── Branch → City / Format mapping ──────────────────────────────────────────
# mapping_dict = {
#     "1 MG": {"City": "Bangalore", "Format": "Café"},
#     "1 Mg Cafe": {"City": "Bangalore", "Format": "Café"},
#     "1mg cafe": {"City": "Bangalore", "Format": "Café"},
#     "Adarsh Palm Retreat": {"City": "Bangalore", "Format": "Café"},
#     "Arumbakkam Cloud Kitchen": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "arumbakkam ck": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Avadi CH": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Bandra": {"City": "Mumbai", "Format": "Café"},
#     "Banni Square": {"City": "Delhi NCR", "Format": "Kiosk"},
#     "Basaveshwara Nagar CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "BEL Road": {"City": "Bangalore", "Format": "Café"},
#     "Bellandur": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Bhartiya City Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "BLR Airport T2 International Bakery": {"City": "Bangalore", "Format": "Airport"},
#     "BLR Airport Terminal 1": {"City": "Bangalore", "Format": "Airport"},
#     "Central Arcade DLF": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "dlf": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Chennai ECR": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Cummulus": {"City": "Bangalore", "Format": "Café"},
#     "Dwarka Cloud Kitchen": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "dwarka ck": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Electronic City": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Express Avenue SMOOR": {"City": "Chennai", "Format": "Kiosk"},
#     "express avenue": {"City": "Chennai", "Format": "Kiosk"},
#     "Forum Shantiniketan": {"City": "Bangalore", "Format": "Kiosk"},
#     "Forum South Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "Garuda Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "Godrej One": {"City": "Mumbai", "Format": "Kiosk"},
#     "Grant Road CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Head Office": {"City": "Bangalore", "Format": "NA"},
#     "Hesaraghatta": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Hinjewadi": {"City": "Pune", "Format": "Cloud Kitchen"},
#     "HSR": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "IC Colony CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Indiranagar": {"City": "Bangalore", "Format": "Signature"},
#     "Infinity Mall Malad": {"City": "Mumbai", "Format": "Kiosk"},
#     "J.P Nagar": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "j.p nagar": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Jayanagar": {"City": "Bangalore", "Format": "Café"},
#     "Kadugodi CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Kalkaji CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Kamanahalli": {"City": "Bangalore", "Format": "Café"},
#     "kammanahalli": {"City": "Bangalore", "Format": "Café"},
#     "Kharghar CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Koramangala": {"City": "Bangalore", "Format": "Café"},
#     "Koreagaon Park": {"City": "Pune", "Format": "Signature"},
#     "Lavelle Road": {"City": "Bangalore", "Format": "Signature"},
#     "Laxmi Nagar": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Magarpatta": {"City": "Pune", "Format": "Cloud Kitchen"},
#     "Mahadevpura": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Malad": {"City": "Mumbai", "Format": "Kiosk"},
#     "MAP Smoor": {"City": "Bangalore", "Format": "Café"},
#     "Marathahalli CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Medavakkam TN": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Mira Road CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Mukherjee Nagar": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Netaji Shubhash Palace": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "nsp": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Nexus Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "Noida Sec 46": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Noida Sec 83": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Orion Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "Perungudi CK": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Phoenix Mall Pune": {"City": "Pune", "Format": "Kiosk"},
#     "Powai": {"City": "Mumbai", "Format": "Signature"},
#     "Rajouri Garden": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "rajouri": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Ramapuram TN": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "ramapuram": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "RT Nagar CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Sadashivanagar": {"City": "Bangalore", "Format": "Café"},
#     "Sakinaka CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Sanpada": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Sarjapur CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Seawoods Mall": {"City": "Mumbai", "Format": "Café"},
#     "Smoor - Defence Colony CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "defence colony ck": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Test.Rebel": {"City": "NA", "Format": "NA"},
#     "Udyog Vihar CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Undri": {"City": "Pune", "Format": "Cloud Kitchen"},
#     "Uttarahalli CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Vandalur": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Vashi Smoor": {"City": "Mumbai", "Format": "Kiosk"},
#     "Vasundra CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Vega City": {"City": "Bangalore", "Format": "Kiosk"},
#     "Versova": {"City": "Mumbai", "Format": "Café"},
#     "Viviana Lounge Mumbai": {"City": "Mumbai", "Format": "Kiosk"},
#     "viviana": {"City": "Mumbai", "Format": "Kiosk"},
#     "West Tambaram": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Whitefield": {"City": "Bangalore", "Format": "Signature"},
#     "Whitefield Lounge": {"City": "Bangalore", "Format": "Signature"},
#     "Worli": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Yelahanka Cloud Kitchen": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "yelahanka": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Kandivali CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Airoli CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Ambience Mall": {"City": "Delhi NCR", "Format": "Kiosk"},
#     "Bhandup CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Green Park CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Gurugram sector 31 Cloud Kitchen": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Malviya Nagar CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "RR Nagar CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Rohini CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Henjewadi": {"City": "Pune", "Format": "Cloud Kitchen"},
#     "Hesarghatta": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "hessaraghatta": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Embassy Lake": {"City": "Bangalore", "Format": "Café"},
#     "Chembur East CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Hesaraghatta Outlet": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Mahakali CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Noida Sector 46": {"City": "NCR", "Format": "Cloud Kitchen"},
#     "Noida Sector 83": {"City": "NCR", "Format": "Cloud Kitchen"},
#     "Velachery": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Smoor Deloitte": {"City": "Bangalore", "Format": "Kiosk"},
#     "Zirakpur": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "RBI Officers Canteen": {"City": "Bangalore", "Format": "Kiosk"},
# }

# # ── Phone cleaning ────────────────────────────────────────────────────────────
# def clean_phone(raw):
#     if pd.isna(raw):
#         return None
#     phone = str(raw).strip()
#     phone = re.sub(r'[\s\-\(\)\+\.]', '', phone)
#     phone = re.sub(r'\D', '', phone)
#     if not phone:
#         return None
#     if len(phone) == 12 and phone.startswith('91'):
#         phone = phone[2:]
#     if len(phone) == 11 and phone.startswith('0'):
#         phone = phone[1:]
#     if len(phone) != 10:
#         return None
#     for digit in set(phone):
#         if phone.count(digit) > 5:
#             return None
#     if phone[0] in ('0', '1'):
#         return None
#     if phone[0] not in ('6', '7', '8', '9'):
#         return None
#     return phone


# # ── Main analysis ─────────────────────────────────────────────────────────────
# def build_report(input_csv: str, output_xlsx: str):

#     # 1. Load & filter
#     df = pd.read_csv(input_csv, dtype=str, low_memory=False)
#     df = df[df['channel'].str.strip().str.lower().isin(['dine in', 'take away', 'dine-in', 'takeaway'])]

#     # 2. Clean phones
#     df['clean_phone'] = df['customer_phoneNumber'].apply(clean_phone)

#     # 3. Deduplicate at invoice level
#     inv = df.drop_duplicates(subset='invoiceNumber')[
#         ['invoiceNumber', 'branchName', 'channel', 'date', 'clean_phone']
#     ].copy()
#     inv['date'] = pd.to_datetime(inv['date'], errors='coerce').dt.date

#     # 4. Add City / Format via mapping
#     mapping_df = pd.DataFrame.from_dict(mapping_dict, orient='index').reset_index()
#     mapping_df.columns = ['branchName', 'City', 'Format']
#     inv = inv.merge(mapping_df, on='branchName', how='left')
#     inv['City']   = inv['City'].fillna('Unknown')
#     inv['Format'] = inv['Format'].fillna('Unknown')
#     inv = inv[~inv['City'].isin(['Unknown', 'NA'])]

#     # ── EXCLUDE Cloud Kitchen outlets ────────────────────────────────────────
#     inv = inv[inv['Format'] != 'Cloud Kitchen']

#     # 5. Week / day buckets
#     inv['date_dt']  = pd.to_datetime(inv['date'], errors='coerce')
#     inv['iso_week'] = inv['date_dt'].dt.isocalendar().week.astype(int)
#     inv['iso_year'] = inv['date_dt'].dt.isocalendar().year.astype(int)
#     inv['yw']       = inv['iso_year'] * 100 + inv['iso_week']

#     last3_yw          = sorted(inv['yw'].dropna().unique())[-3:]
#     target_week_pairs = [(int(yw // 100), int(yw % 100)) for yw in last3_yw]

#     all_days    = sorted(inv['date'].dropna().unique())
#     target_days = all_days[-3:] if len(all_days) >= 3 else all_days

#     # 6. Per-branch metrics
#     order_counts = inv.groupby('branchName')['invoiceNumber'].nunique().reset_index()
#     order_counts.columns = ['branchName', 'total_orders']

#     branches = (
#         inv[['branchName', 'City', 'Format']]
#         .drop_duplicates('branchName')
#         .merge(order_counts, on='branchName', how='left')
#         .sort_values('total_orders', ascending=False)
#         .reset_index(drop=True)
#     )

#     # ── FIXED: Unique valid phone numbers ÷ Unique invoices ──────────────────
#     def capture(sub):
#         total     = sub['invoiceNumber'].nunique()
#         has_phone = sub['clean_phone'].nunique()  # unique valid phone numbers
#         return has_phone / total if total > 0 else None

#     day_fmt   = lambda d: d.strftime('%b %#d') if os.name == 'nt' else d.strftime('%b %-d')
#     latest_yw = last3_yw[-1] if last3_yw else None

#     # Week cols
#     week_cols = []
#     for (yr, wk) in target_week_pairs:
#         yw_key = yr * 100 + wk
#         col    = f'W{wk} WTD' if yw_key == latest_yw else f'W{wk}'
#         week_cols.append(col)
#         rates = []
#         for _, row in branches.iterrows():
#             sub = inv[(inv['branchName'] == row['branchName']) &
#                       (inv['iso_week'] == wk) & (inv['iso_year'] == yr)]
#             rates.append(capture(sub))
#         branches[col] = rates

#     # Day cols
#     day_cols = []
#     for d in target_days:
#         col = day_fmt(d)
#         day_cols.append(col)
#         rates = []
#         for _, row in branches.iterrows():
#             sub = inv[(inv['branchName'] == row['branchName']) & (inv['date'] == d)]
#             rates.append(capture(sub))
#         branches[col] = rates

#     metric_cols = week_cols + day_cols

#     # 7. Build Excel ────────────────────────────────────────────────────────────
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Phone No Record Tracker"

#     ROYAL_BLUE = "1A3A6B"
#     HDR_WEEK   = "2C5282"
#     HDR_DAY    = "1A3A6B"
#     WHITE      = "FFFFFF"
#     LIGHT_ROW  = "F7F9FC"
#     ALT_ROW    = "EEF2F9"
#     BORDER_C   = "C5CFE3"

#     thin = Side(style='thin', color=BORDER_C)
#     bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

#     total_cols      = 3 + len(metric_cols)
#     last_col_letter = get_column_letter(total_cols)

#     # Title row
#     ws.merge_cells(f'A1:{last_col_letter}1')
#     c = ws['A1']
#     c.value     = "Phone No Record Tracker  |  Dine In & Take Away"
#     c.font      = Font(name='Arial', bold=True, size=14, color=WHITE)
#     c.fill      = PatternFill('solid', start_color=ROYAL_BLUE)
#     c.alignment = Alignment(horizontal='center', vertical='center')
#     ws.row_dimensions[1].height = 30

#     ws.merge_cells(f'A2:{last_col_letter}2')
#     ws['A2'].fill = PatternFill('solid', start_color=ROYAL_BLUE)
#     ws.row_dimensions[2].height = 6

#     # Group header row (row 3)
#     for c_idx in [1, 2, 3]:
#         ws.cell(row=3, column=c_idx).fill = PatternFill('solid', start_color=ROYAL_BLUE)

#     ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=6)
#     wk_cell = ws.cell(row=3, column=4, value="← Last 3 Weeks →")
#     wk_cell.font      = Font(name='Arial', bold=True, size=9, color=WHITE)
#     wk_cell.fill      = PatternFill('solid', start_color=HDR_WEEK)
#     wk_cell.alignment = Alignment(horizontal='center')

#     ws.merge_cells(start_row=3, start_column=7, end_row=3, end_column=9)
#     dy_cell = ws.cell(row=3, column=7, value="← Last 3 Days →")
#     dy_cell.font      = Font(name='Arial', bold=True, size=9, color=WHITE)
#     dy_cell.fill      = PatternFill('solid', start_color=HDR_DAY)
#     dy_cell.alignment = Alignment(horizontal='center')

#     # Column label row (row 4)
#     headers  = ['City', 'Branch', 'Format'] + metric_cols
#     col_meta = {}
#     for h in ['City', 'Branch', 'Format']:
#         col_meta[h] = (ROYAL_BLUE, WHITE)
#     for c in week_cols:
#         col_meta[c] = (HDR_WEEK, WHITE)
#     for c in day_cols:
#         col_meta[c] = (HDR_DAY, WHITE)

#     for ci, h in enumerate(headers, start=1):
#         cell = ws.cell(row=4, column=ci, value=h)
#         bg, fg = col_meta.get(h, (ROYAL_BLUE, WHITE))
#         cell.fill      = PatternFill('solid', start_color=bg)
#         cell.font      = Font(name='Arial', bold=True, size=10, color=fg)
#         cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
#         cell.border    = bdr
#     ws.row_dimensions[4].height = 28

#     # Data rows
#     for ri, row in branches.iterrows():
#         excel_row = ri + 5
#         fill_bg   = LIGHT_ROW if ri % 2 == 0 else ALT_ROW
#         vals = [row['City'], row['branchName'], row['Format']] + [row[c] for c in metric_cols]
#         for ci, val in enumerate(vals, start=1):
#             cell        = ws.cell(row=excel_row, column=ci)
#             cell.border = bdr
#             if ci <= 3:
#                 cell.value     = val
#                 cell.font      = Font(name='Arial', size=9)
#                 cell.fill      = PatternFill('solid', start_color=fill_bg)
#                 cell.alignment = Alignment(horizontal='left', vertical='center')
#             else:
#                 if val is None:
#                     cell.value     = "—"
#                     cell.font      = Font(name='Arial', size=9, color='BBBBBB')
#                     cell.fill      = PatternFill('solid', start_color=fill_bg)
#                     cell.alignment = Alignment(horizontal='center')
#                 else:
#                     cell.value         = val
#                     cell.number_format = '0.0%'
#                     cell.font          = Font(name='Arial', size=9)
#                     cell.fill          = PatternFill('solid', start_color=fill_bg)
#                     cell.alignment     = Alignment(horizontal='center')

#     total_data_rows = len(branches)
#     last_row        = 4 + total_data_rows

#     # Conditional formatting: ONLY on day cols
#     day_col_start = 3 + len(week_cols) + 1
#     for ci_offset in range(len(day_cols)):
#         ci         = day_col_start + ci_offset
#         col_letter = get_column_letter(ci)
#         ws.conditional_formatting.add(
#             f"{col_letter}5:{col_letter}{last_row}",
#             ColorScaleRule(
#                 start_type='min', start_color='FFFFFF',
#                 end_type='max',   end_color='C0392B'
#             )
#         )

#     col_widths = [14, 28, 14, 9, 9, 9, 9, 9, 9]
#     for ci, w in enumerate(col_widths, start=1):
#         ws.column_dimensions[get_column_letter(ci)].width = w

#     ws.freeze_panes = 'D5'

#     # 8. Raw Dump sheet ─────────────────────────────────────────────────────────
#     ws2 = wb.create_sheet("Raw Dump")

#     extra_cols_wanted = ['invoiceDate', 'customer_name', 'customer_email',
#                          'customer_id', 'orderNumber', 'netAmount', 'grossAmount']
#     available_extra = [c for c in extra_cols_wanted if c in df.columns]

#     df['_qty'] = pd.to_numeric(df.get('item_quantity', 0), errors='coerce').fillna(0)
#     sale_qty   = df.groupby('invoiceNumber')['_qty'].sum().reset_index()
#     sale_qty.columns = ['invoiceNumber', 'sale_qty']

#     raw = df.drop_duplicates(subset='invoiceNumber')[
#         ['invoiceNumber', 'branchName', 'channel', 'date', 'clean_phone'] + available_extra
#     ].copy()
#     raw = raw.merge(mapping_df[['branchName', 'City', 'Format']], on='branchName', how='left')
#     raw = raw.merge(sale_qty, on='invoiceNumber', how='left')
#     raw['City']   = raw['City'].fillna('Unknown')
#     raw['Format'] = raw['Format'].fillna('Unknown')
#     raw['date']   = pd.to_datetime(raw['date'], errors='coerce').dt.date

#     # ── EXCLUDE Cloud Kitchen from Raw Dump too ───────────────────────────────
#     raw = raw[raw['Format'] != 'Cloud Kitchen']

#     for nc in ['netAmount', 'grossAmount']:
#         if nc in raw.columns:
#             raw[nc] = pd.to_numeric(raw[nc], errors='coerce')

#     raw_dump = (
#         raw[raw['clean_phone'].notna()]
#         .merge(order_counts, on='branchName', how='left')
#         .sort_values(['total_orders', 'date', 'branchName'], ascending=[False, True, True])
#         .drop(columns=['total_orders'])
#         .reset_index(drop=True)
#     )

#     ordered_cols = ['invoiceNumber', 'date']
#     for c in ['invoiceDate', 'branchName', 'City', 'Format', 'channel',
#               'customer_name', 'clean_phone', 'customer_email', 'customer_id',
#               'orderNumber', 'sale_qty', 'netAmount', 'grossAmount']:
#         if c in raw_dump.columns:
#             ordered_cols.append(c)
#     raw_dump = raw_dump[ordered_cols]

#     header_labels = {
#         'invoiceNumber': 'Invoice No', 'date': 'Date', 'invoiceDate': 'Invoice Date',
#         'branchName': 'Branch', 'City': 'City', 'Format': 'Format', 'channel': 'Channel',
#         'customer_name': 'Customer Name', 'clean_phone': 'Phone Number',
#         'customer_email': 'Email', 'customer_id': 'Customer ID', 'orderNumber': 'Order No',
#         'sale_qty': 'Sale Qty', 'netAmount': 'Net Amount', 'grossAmount': 'Gross Amount',
#     }

#     numeric_cols_raw = {'netAmount', 'grossAmount', 'sale_qty'}
#     left_cols        = {'branchName', 'customer_name', 'customer_email'}
#     hdr_fill  = PatternFill('solid', start_color="1A3A6B")
#     hdr_font  = Font(name='Arial', bold=True, size=10, color="FFFFFF")
#     hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

#     for ci, col in enumerate(ordered_cols, start=1):
#         cell = ws2.cell(row=1, column=ci, value=header_labels.get(col, col))
#         cell.fill = hdr_fill; cell.font = hdr_font
#         cell.alignment = hdr_align; cell.border = bdr
#     ws2.row_dimensions[1].height = 28

#     for ri, row_data in raw_dump.iterrows():
#         excel_row = ri + 2
#         fill_bg   = "F9FAFB" if ri % 2 == 0 else "FFFFFF"
#         for ci, col in enumerate(ordered_cols, start=1):
#             val  = row_data[col]
#             cell = ws2.cell(row=excel_row, column=ci)
#             cell.fill   = PatternFill('solid', start_color=fill_bg)
#             cell.font   = Font(name='Arial', size=9)
#             cell.border = bdr
#             if col in numeric_cols_raw:
#                 try:
#                     cell.value = float(val) if val is not None and str(val) != 'nan' else None
#                 except (ValueError, TypeError):
#                     cell.value = None
#                 cell.number_format = '#,##0.00' if col in {'netAmount', 'grossAmount'} else '#,##0'
#                 cell.alignment = Alignment(horizontal='right', vertical='center')
#             elif col == 'clean_phone':
#                 cell.value         = str(val) if val is not None else None
#                 cell.number_format = '@'
#                 cell.alignment     = Alignment(horizontal='center', vertical='center')
#             else:
#                 cell.value     = val
#                 cell.alignment = Alignment(
#                     horizontal='left' if col in left_cols else 'center',
#                     vertical='center'
#                 )

#     col_widths2 = {
#         'invoiceNumber': 16, 'date': 12, 'invoiceDate': 14, 'branchName': 26,
#         'City': 12, 'Format': 14, 'channel': 12, 'customer_name': 20,
#         'clean_phone': 15, 'customer_email': 26, 'customer_id': 14,
#         'orderNumber': 14, 'sale_qty': 10, 'netAmount': 14, 'grossAmount': 14,
#     }
#     for ci, col in enumerate(ordered_cols, start=1):
#         ws2.column_dimensions[get_column_letter(ci)].width = col_widths2.get(col, 14)

#     ws2.freeze_panes = 'A2'

#     wb.save(output_xlsx)
#     print(f"✅ Report saved → {output_xlsx}")
#     print(f"   Branches analysed             : {len(branches)}")
#     print(f"   Total invoices (Dine In / TA) : {len(inv)}")
#     print(f"   Raw Dump rows (with phone)    : {len(raw_dump)}")
#     phone_fill_pct = (inv['clean_phone'].nunique() / inv['invoiceNumber'].nunique() * 100) if len(inv) else 0
#     print(f"   Overall phone capture         : {phone_fill_pct:.1f}%")


# # ── Entry point ───────────────────────────────────────────────────────────────
# if __name__ == "__main__":
#     import sys

#     ROLLING_CSV = r"C:/Users/jeryy/OneDrive - BLISS CHOCOLATES INDIA PRIVATE LIMITED/Desktop/Documents/Phone No/Data/sales_rolling.csv"
#     OUTPUT_DIR  = r"C:/Users/jeryy/OneDrive - BLISS CHOCOLATES INDIA PRIVATE LIMITED/Desktop/Documents/Phone No/output"

#     input_file = sys.argv[1] if len(sys.argv) > 1 else ROLLING_CSV
#     if len(sys.argv) > 2:
#         output_file = sys.argv[2]
#     else:
#         os.makedirs(OUTPUT_DIR, exist_ok=True)
#         date_str    = datetime.today().strftime("%Y-%m-%d")
#         output_file = os.path.join(OUTPUT_DIR, f"Phone No Record - {date_str}.xlsx")

#     build_report(input_file, output_file)





# import os
# import pandas as pd
# import re
# from datetime import datetime, date
# import openpyxl
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from openpyxl.formatting.rule import ColorScaleRule
# from openpyxl.utils import get_column_letter

# # ── Branch → City / Format mapping ──────────────────────────────────────────
# mapping_dict = {
#     "1 MG": {"City": "Bangalore", "Format": "Café"},
#     "1 Mg Cafe": {"City": "Bangalore", "Format": "Café"},
#     "1mg cafe": {"City": "Bangalore", "Format": "Café"},
#     "Adarsh Palm Retreat": {"City": "Bangalore", "Format": "Café"},
#     "Arumbakkam Cloud Kitchen": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "arumbakkam ck": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Avadi CH": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Bandra": {"City": "Mumbai", "Format": "Café"},
#     "Banni Square": {"City": "Delhi NCR", "Format": "Kiosk"},
#     "Basaveshwara Nagar CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "BEL Road": {"City": "Bangalore", "Format": "Café"},
#     "Bellandur": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Bhartiya City Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "BLR Airport T2 International Bakery": {"City": "Bangalore", "Format": "Airport"},
#     "BLR Airport Terminal 1": {"City": "Bangalore", "Format": "Airport"},
#     "Central Arcade DLF": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "dlf": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Chennai ECR": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Cummulus": {"City": "Bangalore", "Format": "Café"},
#     "Dwarka Cloud Kitchen": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "dwarka ck": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Electronic City": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Express Avenue SMOOR": {"City": "Chennai", "Format": "Kiosk"},
#     "express avenue": {"City": "Chennai", "Format": "Kiosk"},
#     "Forum Shantiniketan": {"City": "Bangalore", "Format": "Kiosk"},
#     "Forum South Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "Garuda Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "Godrej One": {"City": "Mumbai", "Format": "Kiosk"},
#     "Grant Road CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Head Office": {"City": "Bangalore", "Format": "NA"},
#     "Hesaraghatta": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Hinjewadi": {"City": "Pune", "Format": "Cloud Kitchen"},
#     "HSR": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "IC Colony CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Indiranagar": {"City": "Bangalore", "Format": "Signature"},
#     "Infinity Mall Malad": {"City": "Mumbai", "Format": "Kiosk"},
#     "J.P Nagar": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "j.p nagar": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Jayanagar": {"City": "Bangalore", "Format": "Café"},
#     "Kadugodi CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Kalkaji CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Kamanahalli": {"City": "Bangalore", "Format": "Café"},
#     "kammanahalli": {"City": "Bangalore", "Format": "Café"},
#     "Kharghar CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Koramangala": {"City": "Bangalore", "Format": "Café"},
#     "Koreagaon Park": {"City": "Pune", "Format": "Signature"},
#     "Lavelle Road": {"City": "Bangalore", "Format": "Signature"},
#     "Laxmi Nagar": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Magarpatta": {"City": "Pune", "Format": "Cloud Kitchen"},
#     "Mahadevpura": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Malad": {"City": "Mumbai", "Format": "Kiosk"},
#     "MAP Smoor": {"City": "Bangalore", "Format": "Café"},
#     "Marathahalli CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Medavakkam TN": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Mira Road CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Mukherjee Nagar": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Netaji Shubhash Palace": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "nsp": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Nexus Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "Noida Sec 46": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Noida Sec 83": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Orion Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "Perungudi CK": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Phoenix Mall Pune": {"City": "Pune", "Format": "Kiosk"},
#     "Powai": {"City": "Mumbai", "Format": "Signature"},
#     "Rajouri Garden": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "rajouri": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Ramapuram TN": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "ramapuram": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "RT Nagar CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Sadashivanagar": {"City": "Bangalore", "Format": "Café"},
#     "Sakinaka CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Sanpada": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Sarjapur CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Seawoods Mall": {"City": "Mumbai", "Format": "Café"},
#     "Smoor - Defence Colony CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "defence colony ck": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Test.Rebel": {"City": "NA", "Format": "NA"},
#     "Udyog Vihar CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Undri": {"City": "Pune", "Format": "Cloud Kitchen"},
#     "Uttarahalli CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Vandalur": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Vashi Smoor": {"City": "Mumbai", "Format": "Kiosk"},
#     "Vasundra CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Vega City": {"City": "Bangalore", "Format": "Kiosk"},
#     "Versova": {"City": "Mumbai", "Format": "Café"},
#     "Viviana Lounge Mumbai": {"City": "Mumbai", "Format": "Kiosk"},
#     "viviana": {"City": "Mumbai", "Format": "Kiosk"},
#     "West Tambaram": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Whitefield": {"City": "Bangalore", "Format": "Signature"},
#     "Whitefield Lounge": {"City": "Bangalore", "Format": "Signature"},
#     "Worli": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Yelahanka Cloud Kitchen": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "yelahanka": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Kandivali CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Airoli CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Ambience Mall": {"City": "Delhi NCR", "Format": "Kiosk"},
#     "Bhandup CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Green Park CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Gurugram sector 31 Cloud Kitchen": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Malviya Nagar CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "RR Nagar CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Rohini CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Henjewadi": {"City": "Pune", "Format": "Cloud Kitchen"},
#     "Hesarghatta": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "hessaraghatta": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Embassy Lake": {"City": "Bangalore", "Format": "Café"},
#     "Chembur East CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Hesaraghatta Outlet": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Mahakali CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Noida Sector 46": {"City": "NCR", "Format": "Cloud Kitchen"},
#     "Noida Sector 83": {"City": "NCR", "Format": "Cloud Kitchen"},
#     "Velachery": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Smoor Deloitte": {"City": "Bangalore", "Format": "Kiosk"},
#     "Zirakpur": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "RBI Officers Canteen": {"City": "Bangalore", "Format": "Kiosk"},
# }

# # ── Phone cleaning ────────────────────────────────────────────────────────────
# def clean_phone(raw):
#     if pd.isna(raw):
#         return None
#     phone = str(raw).strip()
#     phone = re.sub(r'[\s\-\(\)\+\.]', '', phone)
#     phone = re.sub(r'\D', '', phone)
#     if not phone:
#         return None
#     if len(phone) == 12 and phone.startswith('91'):
#         phone = phone[2:]
#     if len(phone) == 11 and phone.startswith('0'):
#         phone = phone[1:]
#     if len(phone) != 10:
#         return None
#     for digit in set(phone):
#         if phone.count(digit) > 5:
#             return None
#     if phone[0] in ('0', '1'):
#         return None
#     if phone[0] not in ('6', '7', '8', '9'):
#         return None
#     return phone


# # ── Logging helper ────────────────────────────────────────────────────────────
# class Logger:
#     def __init__(self, log_path: str):
#         self.log_path = log_path
#         self.lines    = []

#     def log(self, text: str = ""):
#         self.lines.append(text)

#     def section(self, title: str):
#         self.log()
#         self.log("=" * 80)
#         self.log(f"  {title}")
#         self.log("=" * 80)

#     def subsection(self, title: str):
#         self.log()
#         self.log(f"  ── {title} ──")
#         self.log()

#     def save(self):
#         with open(self.log_path, 'w', encoding='utf-8') as f:
#             f.write('\n'.join(self.lines))
#         print(f"📋 Log saved → {self.log_path}")


# # ── Main analysis ─────────────────────────────────────────────────────────────
# def build_report(input_csv: str, output_xlsx: str):

#     # Setup logger
#     log_path = output_xlsx.replace('.xlsx', '_log.txt')
#     L = Logger(log_path)
#     run_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

#     L.log("=" * 80)
#     L.log("  PHONE NO RECORD TRACKER  —  DETAILED CALCULATION LOG")
#     L.log(f"  Run timestamp : {run_time}")
#     L.log(f"  Input file    : {input_csv}")
#     L.log(f"  Output file   : {output_xlsx}")
#     L.log("=" * 80)

#     # ── STEP 1: Load & filter ─────────────────────────────────────────────────
#     L.section("STEP 1 — LOAD & CHANNEL FILTER")
#     df_raw = pd.read_csv(input_csv, dtype=str, low_memory=False)
#     L.log(f"  Total rows loaded from CSV          : {len(df_raw):,}")
#     L.log(f"  Unique invoices in raw data         : {df_raw['invoiceNumber'].nunique():,}")

#     channel_counts = df_raw['channel'].str.strip().str.lower().value_counts()
#     L.log()
#     L.log("  Channel distribution in raw data:")
#     for ch, cnt in channel_counts.items():
#         L.log(f"    {ch:<30} : {cnt:,}")

#     df = df_raw[df_raw['channel'].str.strip().str.lower().isin(['dine in', 'take away', 'dine-in', 'takeaway'])].copy()
#     L.log()
#     L.log(f"  Rows KEPT after channel filter      : {len(df):,}")
#     L.log(f"  Rows DROPPED (other channels)       : {len(df_raw) - len(df):,}")
#     L.log(f"  Unique invoices after filter        : {df['invoiceNumber'].nunique():,}")

#     # ── STEP 2: Phone cleaning ────────────────────────────────────────────────
#     L.section("STEP 2 — PHONE NUMBER CLEANING")
#     df['clean_phone'] = df['customer_phoneNumber'].apply(clean_phone)

#     total_rows        = len(df)
#     raw_has_phone     = df['customer_phoneNumber'].notna().sum()
#     cleaned_has_phone = df['clean_phone'].notna().sum()
#     dropped_in_clean  = raw_has_phone - cleaned_has_phone

#     L.log(f"  Total rows (post channel filter)    : {total_rows:,}")
#     L.log(f"  Rows with ANY phone value           : {raw_has_phone:,}")
#     L.log(f"  Rows with VALID cleaned phone       : {cleaned_has_phone:,}")
#     L.log(f"  Rows where phone was REJECTED       : {dropped_in_clean:,}")
#     L.log()
#     L.log("  Phone rejection reasons (applied in order):")
#     L.log("    1. Empty / NaN value")
#     L.log("    2. No digits after stripping special chars")
#     L.log("    3. 12-digit starting with 91 → strip country code")
#     L.log("    4. 11-digit starting with 0  → strip leading zero")
#     L.log("    5. Not exactly 10 digits after cleaning")
#     L.log("    6. Any single digit repeats more than 5 times (e.g. 9999999999)")
#     L.log("    7. First digit is 0 or 1 (invalid Indian mobile)")
#     L.log("    8. First digit not in 6,7,8,9 (invalid Indian mobile prefix)")

#     # ── STEP 3: Deduplicate at invoice level ──────────────────────────────────
#     L.section("STEP 3 — DEDUPLICATE AT INVOICE LEVEL")
#     before_dedup = len(df)
#     inv = df.drop_duplicates(subset='invoiceNumber')[
#         ['invoiceNumber', 'branchName', 'channel', 'date', 'clean_phone']
#     ].copy()
#     inv['date'] = pd.to_datetime(inv['date'], errors='coerce').dt.date
#     after_dedup = len(inv)

#     L.log(f"  Rows before dedup                   : {before_dedup:,}")
#     L.log(f"  Unique invoices after dedup         : {after_dedup:,}")
#     L.log(f"  Duplicate rows removed              : {before_dedup - after_dedup:,}")
#     L.log()
#     L.log("  NOTE: When multiple rows share the same invoiceNumber,")
#     L.log("  only the FIRST occurrence is kept (drop_duplicates keeps first row).")
#     L.log("  This means phone number is taken from the first line item of each invoice.")

#     # ── STEP 4: City / Format mapping ─────────────────────────────────────────
#     L.section("STEP 4 — BRANCH MAPPING (City & Format)")
#     mapping_df = pd.DataFrame.from_dict(mapping_dict, orient='index').reset_index()
#     mapping_df.columns = ['branchName', 'City', 'Format']
#     inv = inv.merge(mapping_df, on='branchName', how='left')
#     inv['City']   = inv['City'].fillna('Unknown')
#     inv['Format'] = inv['Format'].fillna('Unknown')

#     unknown_branches = inv[inv['City'] == 'Unknown']['branchName'].unique()
#     L.log(f"  Branches mapped successfully        : {inv[inv['City'] != 'Unknown']['branchName'].nunique():,}")
#     if len(unknown_branches) > 0:
#         L.log(f"  Branches NOT found in mapping       : {len(unknown_branches)}")
#         for b in unknown_branches:
#             L.log(f"    → '{b}'  (will be DROPPED)")
#     else:
#         L.log("  All branches found in mapping dict  : ✓")

#     inv_before_filter = len(inv)
#     inv = inv[~inv['City'].isin(['Unknown', 'NA'])]
#     inv = inv[inv['Format'] != 'Cloud Kitchen']
#     inv_after_filter  = len(inv)

#     format_dist = inv['Format'].value_counts()
#     L.log()
#     L.log(f"  Invoices dropped (Unknown/NA city)  : {inv_before_filter - inv_after_filter:,}")
#     L.log(f"  Cloud Kitchen invoices excluded     : (filtered out entirely)")
#     L.log(f"  Invoices remaining for analysis     : {inv_after_filter:,}")
#     L.log()
#     L.log("  Format distribution (post-filter):")
#     for fmt, cnt in format_dist.items():
#         L.log(f"    {fmt:<20} : {cnt:,} invoices")

#     # ── STEP 5: Week / Day buckets ────────────────────────────────────────────
#     L.section("STEP 5 — WEEK & DAY BUCKETS")
#     inv['date_dt']  = pd.to_datetime(inv['date'], errors='coerce')
#     inv['iso_week'] = inv['date_dt'].dt.isocalendar().week.astype(int)
#     inv['iso_year'] = inv['date_dt'].dt.isocalendar().year.astype(int)
#     inv['yw']       = inv['iso_year'] * 100 + inv['iso_week']

#     last3_yw          = sorted(inv['yw'].dropna().unique())[-3:]
#     target_week_pairs = [(int(yw // 100), int(yw % 100)) for yw in last3_yw]
#     all_days          = sorted(inv['date'].dropna().unique())
#     target_days       = all_days[-3:] if len(all_days) >= 3 else all_days
#     latest_yw         = last3_yw[-1] if last3_yw else None

#     L.log(f"  Date range in data                  : {min(all_days)} → {max(all_days)}")
#     L.log(f"  Total unique days in data           : {len(all_days)}")
#     L.log()
#     L.log("  Last 3 ISO weeks selected:")
#     for (yr, wk) in target_week_pairs:
#         yw_key = yr * 100 + wk
#         label  = "← WTD (current week)" if yw_key == latest_yw else ""
#         wk_inv = inv[(inv['iso_week'] == wk) & (inv['iso_year'] == yr)]
#         L.log(f"    Year {yr}  Week {wk:02d}  (yw={yw_key})  →  {wk_inv['invoiceNumber'].nunique():,} invoices  {label}")
#     L.log()
#     L.log("  Last 3 days selected:")
#     for d in target_days:
#         day_inv = inv[inv['date'] == d]
#         L.log(f"    {d}  →  {day_inv['invoiceNumber'].nunique():,} invoices across {day_inv['branchName'].nunique()} branches")

#     # ── STEP 6: Branch-level metrics ──────────────────────────────────────────
#     L.section("STEP 6 — BRANCH-LEVEL CAPTURE RATE CALCULATIONS")
#     L.log()
#     L.log("  FORMULA:")
#     L.log("    Capture Rate = Unique valid phone numbers ÷ Unique invoices")
#     L.log("    (scoped to: one branch × one time window)")
#     L.log()

#     order_counts = inv.groupby('branchName')['invoiceNumber'].nunique().reset_index()
#     order_counts.columns = ['branchName', 'total_orders']

#     branches = (
#         inv[['branchName', 'City', 'Format']]
#         .drop_duplicates('branchName')
#         .merge(order_counts, on='branchName', how='left')
#         .sort_values('total_orders', ascending=False)
#         .reset_index(drop=True)
#     )

#     # ── FIXED: Unique valid phone numbers ÷ Unique invoices ──────────────────
#     def capture(sub):
#         total     = sub['invoiceNumber'].nunique()
#         has_phone = sub['clean_phone'].nunique()  # unique valid phone numbers
#         return has_phone / total if total > 0 else None

#     day_fmt   = lambda d: d.strftime('%b %#d') if os.name == 'nt' else d.strftime('%b %-d')

#     # Week cols
#     week_cols = []
#     L.subsection("Weekly Capture Rates by Branch")
#     for (yr, wk) in target_week_pairs:
#         yw_key = yr * 100 + wk
#         col    = f'W{wk} WTD' if yw_key == latest_yw else f'W{wk}'
#         week_cols.append(col)
#         L.log(f"  [{col}]  Year={yr}  Week={wk}")
#         L.log(f"  {'Branch':<30} {'Unique Invoices':>16} {'Unique Phones':>14} {'Capture Rate':>13}")
#         L.log(f"  {'-'*30} {'-'*16} {'-'*14} {'-'*13}")
#         rates = []
#         for _, row in branches.iterrows():
#             sub   = inv[(inv['branchName'] == row['branchName']) &
#                         (inv['iso_week'] == wk) & (inv['iso_year'] == yr)]
#             rate  = capture(sub)
#             rates.append(rate)
#             if len(sub) > 0:
#                 total_inv = sub['invoiceNumber'].nunique()
#                 uniq_ph   = sub['clean_phone'].nunique()
#                 rate_str  = f"{rate:.1%}" if rate is not None else "—"
#                 L.log(f"  {row['branchName']:<30} {total_inv:>16,} {uniq_ph:>14,} {rate_str:>13}")
#         branches[col] = rates
#         L.log()

#     # Day cols
#     day_cols = []
#     L.subsection("Daily Capture Rates by Branch")
#     for d in target_days:
#         col = day_fmt(d)
#         day_cols.append(col)
#         L.log(f"  [{col}]  Date={d}")
#         L.log(f"  {'Branch':<30} {'Unique Invoices':>16} {'Unique Phones':>14} {'Capture Rate':>13}")
#         L.log(f"  {'-'*30} {'-'*16} {'-'*14} {'-'*13}")
#         rates = []
#         for _, row in branches.iterrows():
#             sub   = inv[(inv['branchName'] == row['branchName']) & (inv['date'] == d)]
#             rate  = capture(sub)
#             rates.append(rate)
#             if len(sub) > 0:
#                 total_inv = sub['invoiceNumber'].nunique()
#                 uniq_ph   = sub['clean_phone'].nunique()
#                 rate_str  = f"{rate:.1%}" if rate is not None else "—"
#                 L.log(f"  {row['branchName']:<30} {total_inv:>16,} {uniq_ph:>14,} {rate_str:>13}")
#         branches[col] = rates
#         L.log()

#     metric_cols = week_cols + day_cols

#     # ── STEP 7: Overall summary ───────────────────────────────────────────────
#     L.section("STEP 7 — OVERALL SUMMARY")
#     total_invoices = inv['invoiceNumber'].nunique()
#     total_phones   = inv['clean_phone'].nunique()
#     overall_rate   = total_phones / total_invoices * 100 if total_invoices else 0

#     L.log(f"  Total unique invoices (Dine In/TA)  : {total_invoices:,}")
#     L.log(f"  Total unique valid phone numbers    : {total_phones:,}")
#     L.log(f"  Overall capture rate                : {overall_rate:.1f}%")
#     L.log(f"  Formula: {total_phones:,} ÷ {total_invoices:,} = {overall_rate:.1f}%")

#     L.subsection("Top 15 Most Frequent Phone Numbers (Dummy Number Check)")
#     top_phones = inv['clean_phone'].value_counts().head(15)
#     L.log(f"  {'Phone Number':<15} {'Count':>8}   {'Flag'}")
#     L.log(f"  {'-'*15} {'-'*8}   {'-'*30}")
#     for ph, cnt in top_phones.items():
#         flag = "⚠️  LIKELY DUMMY — appears very frequently" if cnt > 10 else ""
#         L.log(f"  {str(ph):<15} {cnt:>8,}   {flag}")

#     L.subsection("Capture Rate by City")
#     for city in sorted(inv['City'].unique()):
#         city_sub   = inv[inv['City'] == city]
#         city_inv   = city_sub['invoiceNumber'].nunique()
#         city_ph    = city_sub['clean_phone'].nunique()
#         city_rate  = city_ph / city_inv * 100 if city_inv else 0
#         L.log(f"  {city:<15}  Invoices={city_inv:,}  Unique Phones={city_ph:,}  Rate={city_rate:.1f}%")

#     L.subsection("Capture Rate by Format")
#     for fmt in sorted(inv['Format'].unique()):
#         fmt_sub  = inv[inv['Format'] == fmt]
#         fmt_inv  = fmt_sub['invoiceNumber'].nunique()
#         fmt_ph   = fmt_sub['clean_phone'].nunique()
#         fmt_rate = fmt_ph / fmt_inv * 100 if fmt_inv else 0
#         L.log(f"  {fmt:<15}  Invoices={fmt_inv:,}  Unique Phones={fmt_ph:,}  Rate={fmt_rate:.1f}%")

#     L.subsection("All Branches — Overall Capture Rate (sorted by invoices)")
#     L.log(f"  {'Branch':<30} {'City':<12} {'Format':<12} {'Invoices':>10} {'Uniq Phones':>12} {'Rate':>8}")
#     L.log(f"  {'-'*30} {'-'*12} {'-'*12} {'-'*10} {'-'*12} {'-'*8}")
#     for _, row in branches.iterrows():
#         b_sub  = inv[inv['branchName'] == row['branchName']]
#         b_inv  = b_sub['invoiceNumber'].nunique()
#         b_ph   = b_sub['clean_phone'].nunique()
#         b_rate = f"{b_ph/b_inv:.1%}" if b_inv else "—"
#         L.log(f"  {row['branchName']:<30} {row['City']:<12} {row['Format']:<12} {b_inv:>10,} {b_ph:>12,} {b_rate:>8}")

#     L.log()
#     L.log("=" * 80)
#     L.log("  END OF LOG")
#     L.log("=" * 80)
#     L.save()

#     # 7. Build Excel ────────────────────────────────────────────────────────────
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Phone No Record Tracker"

#     ROYAL_BLUE = "1A3A6B"
#     HDR_WEEK   = "2C5282"
#     HDR_DAY    = "1A3A6B"
#     WHITE      = "FFFFFF"
#     LIGHT_ROW  = "F7F9FC"
#     ALT_ROW    = "EEF2F9"
#     BORDER_C   = "C5CFE3"

#     thin = Side(style='thin', color=BORDER_C)
#     bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

#     total_cols      = 3 + len(metric_cols)
#     last_col_letter = get_column_letter(total_cols)

#     # Title row
#     ws.merge_cells(f'A1:{last_col_letter}1')
#     c = ws['A1']
#     c.value     = "Phone No Record Tracker  |  Dine In & Take Away"
#     c.font      = Font(name='Arial', bold=True, size=14, color=WHITE)
#     c.fill      = PatternFill('solid', start_color=ROYAL_BLUE)
#     c.alignment = Alignment(horizontal='center', vertical='center')
#     ws.row_dimensions[1].height = 30

#     ws.merge_cells(f'A2:{last_col_letter}2')
#     ws['A2'].fill = PatternFill('solid', start_color=ROYAL_BLUE)
#     ws.row_dimensions[2].height = 6

#     # Group header row (row 3)
#     for c_idx in [1, 2, 3]:
#         ws.cell(row=3, column=c_idx).fill = PatternFill('solid', start_color=ROYAL_BLUE)

#     ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=6)
#     wk_cell = ws.cell(row=3, column=4, value="← Last 3 Weeks →")
#     wk_cell.font      = Font(name='Arial', bold=True, size=9, color=WHITE)
#     wk_cell.fill      = PatternFill('solid', start_color=HDR_WEEK)
#     wk_cell.alignment = Alignment(horizontal='center')

#     ws.merge_cells(start_row=3, start_column=7, end_row=3, end_column=9)
#     dy_cell = ws.cell(row=3, column=7, value="← Last 3 Days →")
#     dy_cell.font      = Font(name='Arial', bold=True, size=9, color=WHITE)
#     dy_cell.fill      = PatternFill('solid', start_color=HDR_DAY)
#     dy_cell.alignment = Alignment(horizontal='center')

#     # Column label row (row 4)
#     headers  = ['City', 'Branch', 'Format'] + metric_cols
#     col_meta = {}
#     for h in ['City', 'Branch', 'Format']:
#         col_meta[h] = (ROYAL_BLUE, WHITE)
#     for c in week_cols:
#         col_meta[c] = (HDR_WEEK, WHITE)
#     for c in day_cols:
#         col_meta[c] = (HDR_DAY, WHITE)

#     for ci, h in enumerate(headers, start=1):
#         cell = ws.cell(row=4, column=ci, value=h)
#         bg, fg = col_meta.get(h, (ROYAL_BLUE, WHITE))
#         cell.fill      = PatternFill('solid', start_color=bg)
#         cell.font      = Font(name='Arial', bold=True, size=10, color=fg)
#         cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
#         cell.border    = bdr
#     ws.row_dimensions[4].height = 28

#     # Data rows
#     for ri, row in branches.iterrows():
#         excel_row = ri + 5
#         fill_bg   = LIGHT_ROW if ri % 2 == 0 else ALT_ROW
#         vals = [row['City'], row['branchName'], row['Format']] + [row[c] for c in metric_cols]
#         for ci, val in enumerate(vals, start=1):
#             cell        = ws.cell(row=excel_row, column=ci)
#             cell.border = bdr
#             if ci <= 3:
#                 cell.value     = val
#                 cell.font      = Font(name='Arial', size=9)
#                 cell.fill      = PatternFill('solid', start_color=fill_bg)
#                 cell.alignment = Alignment(horizontal='left', vertical='center')
#             else:
#                 if val is None:
#                     cell.value     = "—"
#                     cell.font      = Font(name='Arial', size=9, color='BBBBBB')
#                     cell.fill      = PatternFill('solid', start_color=fill_bg)
#                     cell.alignment = Alignment(horizontal='center')
#                 else:
#                     cell.value         = val
#                     cell.number_format = '0.0%'
#                     cell.font          = Font(name='Arial', size=9)
#                     cell.fill          = PatternFill('solid', start_color=fill_bg)
#                     cell.alignment     = Alignment(horizontal='center')

#     total_data_rows = len(branches)
#     last_row        = 4 + total_data_rows

#     # Conditional formatting: ONLY on day cols
#     day_col_start = 3 + len(week_cols) + 1
#     for ci_offset in range(len(day_cols)):
#         ci         = day_col_start + ci_offset
#         col_letter = get_column_letter(ci)
#         ws.conditional_formatting.add(
#             f"{col_letter}5:{col_letter}{last_row}",
#             ColorScaleRule(
#                 start_type='min', start_color='FFFFFF',
#                 end_type='max',   end_color='C0392B'
#             )
#         )

#     col_widths = [14, 28, 14, 9, 9, 9, 9, 9, 9]
#     for ci, w in enumerate(col_widths, start=1):
#         ws.column_dimensions[get_column_letter(ci)].width = w

#     ws.freeze_panes = 'D5'

#     # 8. Raw Dump sheet ─────────────────────────────────────────────────────────
#     ws2 = wb.create_sheet("Raw Dump")

#     extra_cols_wanted = ['invoiceDate', 'customer_name', 'customer_email',
#                          'customer_id', 'orderNumber', 'netAmount', 'grossAmount']
#     available_extra = [c for c in extra_cols_wanted if c in df.columns]

#     df['_qty'] = pd.to_numeric(df.get('item_quantity', 0), errors='coerce').fillna(0)
#     sale_qty   = df.groupby('invoiceNumber')['_qty'].sum().reset_index()
#     sale_qty.columns = ['invoiceNumber', 'sale_qty']

#     raw = df.drop_duplicates(subset='invoiceNumber')[
#         ['invoiceNumber', 'branchName', 'channel', 'date', 'clean_phone'] + available_extra
#     ].copy()
#     raw = raw.merge(mapping_df[['branchName', 'City', 'Format']], on='branchName', how='left')
#     raw = raw.merge(sale_qty, on='invoiceNumber', how='left')
#     raw['City']   = raw['City'].fillna('Unknown')
#     raw['Format'] = raw['Format'].fillna('Unknown')
#     raw['date']   = pd.to_datetime(raw['date'], errors='coerce').dt.date

#     # ── EXCLUDE Cloud Kitchen from Raw Dump too ───────────────────────────────
#     raw = raw[raw['Format'] != 'Cloud Kitchen']

#     for nc in ['netAmount', 'grossAmount']:
#         if nc in raw.columns:
#             raw[nc] = pd.to_numeric(raw[nc], errors='coerce')

#     raw_dump = (
#         raw[raw['clean_phone'].notna()]
#         .merge(order_counts, on='branchName', how='left')
#         .sort_values(['total_orders', 'date', 'branchName'], ascending=[False, True, True])
#         .drop(columns=['total_orders'])
#         .reset_index(drop=True)
#     )

#     ordered_cols = ['invoiceNumber', 'date']
#     for c in ['invoiceDate', 'branchName', 'City', 'Format', 'channel',
#               'customer_name', 'clean_phone', 'customer_email', 'customer_id',
#               'orderNumber', 'sale_qty', 'netAmount', 'grossAmount']:
#         if c in raw_dump.columns:
#             ordered_cols.append(c)
#     raw_dump = raw_dump[ordered_cols]

#     header_labels = {
#         'invoiceNumber': 'Invoice No', 'date': 'Date', 'invoiceDate': 'Invoice Date',
#         'branchName': 'Branch', 'City': 'City', 'Format': 'Format', 'channel': 'Channel',
#         'customer_name': 'Customer Name', 'clean_phone': 'Phone Number',
#         'customer_email': 'Email', 'customer_id': 'Customer ID', 'orderNumber': 'Order No',
#         'sale_qty': 'Sale Qty', 'netAmount': 'Net Amount', 'grossAmount': 'Gross Amount',
#     }

#     numeric_cols_raw = {'netAmount', 'grossAmount', 'sale_qty'}
#     left_cols        = {'branchName', 'customer_name', 'customer_email'}
#     hdr_fill  = PatternFill('solid', start_color="1A3A6B")
#     hdr_font  = Font(name='Arial', bold=True, size=10, color="FFFFFF")
#     hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

#     for ci, col in enumerate(ordered_cols, start=1):
#         cell = ws2.cell(row=1, column=ci, value=header_labels.get(col, col))
#         cell.fill = hdr_fill; cell.font = hdr_font
#         cell.alignment = hdr_align; cell.border = bdr
#     ws2.row_dimensions[1].height = 28

#     for ri, row_data in raw_dump.iterrows():
#         excel_row = ri + 2
#         fill_bg   = "F9FAFB" if ri % 2 == 0 else "FFFFFF"
#         for ci, col in enumerate(ordered_cols, start=1):
#             val  = row_data[col]
#             cell = ws2.cell(row=excel_row, column=ci)
#             cell.fill   = PatternFill('solid', start_color=fill_bg)
#             cell.font   = Font(name='Arial', size=9)
#             cell.border = bdr
#             if col in numeric_cols_raw:
#                 try:
#                     cell.value = float(val) if val is not None and str(val) != 'nan' else None
#                 except (ValueError, TypeError):
#                     cell.value = None
#                 cell.number_format = '#,##0.00' if col in {'netAmount', 'grossAmount'} else '#,##0'
#                 cell.alignment = Alignment(horizontal='right', vertical='center')
#             elif col == 'clean_phone':
#                 cell.value         = str(val) if val is not None else None
#                 cell.number_format = '@'
#                 cell.alignment     = Alignment(horizontal='center', vertical='center')
#             else:
#                 cell.value     = val
#                 cell.alignment = Alignment(
#                     horizontal='left' if col in left_cols else 'center',
#                     vertical='center'
#                 )

#     col_widths2 = {
#         'invoiceNumber': 16, 'date': 12, 'invoiceDate': 14, 'branchName': 26,
#         'City': 12, 'Format': 14, 'channel': 12, 'customer_name': 20,
#         'clean_phone': 15, 'customer_email': 26, 'customer_id': 14,
#         'orderNumber': 14, 'sale_qty': 10, 'netAmount': 14, 'grossAmount': 14,
#     }
#     for ci, col in enumerate(ordered_cols, start=1):
#         ws2.column_dimensions[get_column_letter(ci)].width = col_widths2.get(col, 14)

#     ws2.freeze_panes = 'A2'

#     wb.save(output_xlsx)
#     phone_fill_pct = (inv['clean_phone'].nunique() / inv['invoiceNumber'].nunique() * 100) if len(inv) else 0
#     print(f"✅ Report saved      → {output_xlsx}")
#     print(f"📋 Log saved         → {log_path}")
#     print(f"   Branches analysed             : {len(branches)}")
#     print(f"   Total invoices (Dine In / TA) : {inv['invoiceNumber'].nunique():,}")
#     print(f"   Raw Dump rows (with phone)    : {len(raw_dump):,}")
#     print(f"   Overall phone capture         : {phone_fill_pct:.1f}%")


# # ── Entry point ───────────────────────────────────────────────────────────────
# if __name__ == "__main__":
#     import sys

#     ROLLING_CSV = r"C:/Users/jeryy/OneDrive - BLISS CHOCOLATES INDIA PRIVATE LIMITED/Desktop/Documents/Phone No/Data/sales_rolling.csv"
#     OUTPUT_DIR  = r"C:/Users/jeryy/OneDrive - BLISS CHOCOLATES INDIA PRIVATE LIMITED/Desktop/Documents/Phone No/output"

#     input_file = sys.argv[1] if len(sys.argv) > 1 else ROLLING_CSV
#     if len(sys.argv) > 2:
#         output_file = sys.argv[2]
#     else:
#         os.makedirs(OUTPUT_DIR, exist_ok=True)
#         date_str    = datetime.today().strftime("%Y-%m-%d")
#         output_file = os.path.join(OUTPUT_DIR, f"Phone No Record - {date_str}.xlsx")

#     build_report(input_file, output_file)



# import os
# import pandas as pd
# import re
# from datetime import datetime, date
# import openpyxl
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from openpyxl.formatting.rule import ColorScaleRule
# from openpyxl.utils import get_column_letter

# # ── Branch → City / Format mapping ──────────────────────────────────────────
# mapping_dict = {
#     "1 MG": {"City": "Bangalore", "Format": "Café"},
#     "1 Mg Cafe": {"City": "Bangalore", "Format": "Café"},
#     "1mg cafe": {"City": "Bangalore", "Format": "Café"},
#     "Adarsh Palm Retreat": {"City": "Bangalore", "Format": "Café"},
#     "Arumbakkam Cloud Kitchen": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "arumbakkam ck": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Avadi CH": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Bandra": {"City": "Mumbai", "Format": "Café"},
#     "Banni Square": {"City": "Delhi NCR", "Format": "Kiosk"},
#     "Basaveshwara Nagar CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "BEL Road": {"City": "Bangalore", "Format": "Café"},
#     "Bellandur": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Bhartiya City Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "BLR Airport T2 International Bakery": {"City": "Bangalore", "Format": "Airport"},
#     "BLR Airport Terminal 1": {"City": "Bangalore", "Format": "Airport"},
#     "Central Arcade DLF": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "dlf": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Chennai ECR": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Cummulus": {"City": "Bangalore", "Format": "Café"},
#     "Dwarka Cloud Kitchen": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "dwarka ck": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Electronic City": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Express Avenue SMOOR": {"City": "Chennai", "Format": "Kiosk"},
#     "express avenue": {"City": "Chennai", "Format": "Kiosk"},
#     "Forum Shantiniketan": {"City": "Bangalore", "Format": "Kiosk"},
#     "Forum South Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "Garuda Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "Godrej One": {"City": "Mumbai", "Format": "Kiosk"},
#     "Grant Road CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Head Office": {"City": "Bangalore", "Format": "NA"},
#     "Hesaraghatta": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Hinjewadi": {"City": "Pune", "Format": "Cloud Kitchen"},
#     "HSR": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "IC Colony CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Indiranagar": {"City": "Bangalore", "Format": "Signature"},
#     "Infinity Mall Malad": {"City": "Mumbai", "Format": "Kiosk"},
#     "J.P Nagar": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "j.p nagar": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Jayanagar": {"City": "Bangalore", "Format": "Café"},
#     "Kadugodi CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Kalkaji CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Kamanahalli": {"City": "Bangalore", "Format": "Café"},
#     "kammanahalli": {"City": "Bangalore", "Format": "Café"},
#     "Kharghar CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Koramangala": {"City": "Bangalore", "Format": "Café"},
#     "Koreagaon Park": {"City": "Pune", "Format": "Signature"},
#     "Lavelle Road": {"City": "Bangalore", "Format": "Signature"},
#     "Laxmi Nagar": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Magarpatta": {"City": "Pune", "Format": "Cloud Kitchen"},
#     "Mahadevpura": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Malad": {"City": "Mumbai", "Format": "Kiosk"},
#     "MAP Smoor": {"City": "Bangalore", "Format": "Café"},
#     "Marathahalli CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Medavakkam TN": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Mira Road CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Mukherjee Nagar": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Netaji Shubhash Palace": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "nsp": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Nexus Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "Noida Sec 46": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Noida Sec 83": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Orion Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "Perungudi CK": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Phoenix Mall Pune": {"City": "Pune", "Format": "Kiosk"},
#     "Powai": {"City": "Mumbai", "Format": "Signature"},
#     "Rajouri Garden": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "rajouri": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Ramapuram TN": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "ramapuram": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "RT Nagar CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Sadashivanagar": {"City": "Bangalore", "Format": "Café"},
#     "Sakinaka CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Sanpada": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Sarjapur CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Seawoods Mall": {"City": "Mumbai", "Format": "Café"},
#     "Smoor - Defence Colony CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "defence colony ck": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Test.Rebel": {"City": "NA", "Format": "NA"},
#     "Udyog Vihar CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Undri": {"City": "Pune", "Format": "Cloud Kitchen"},
#     "Uttarahalli CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Vandalur": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Vashi Smoor": {"City": "Mumbai", "Format": "Kiosk"},
#     "Vasundra CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Vega City": {"City": "Bangalore", "Format": "Kiosk"},
#     "Versova": {"City": "Mumbai", "Format": "Café"},
#     "Viviana Lounge Mumbai": {"City": "Mumbai", "Format": "Kiosk"},
#     "viviana": {"City": "Mumbai", "Format": "Kiosk"},
#     "West Tambaram": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Whitefield": {"City": "Bangalore", "Format": "Signature"},
#     "Whitefield Lounge": {"City": "Bangalore", "Format": "Signature"},
#     "Worli": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Yelahanka Cloud Kitchen": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "yelahanka": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Kandivali CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Airoli CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Ambience Mall": {"City": "Delhi NCR", "Format": "Kiosk"},
#     "Bhandup CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Green Park CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Gurugram sector 31 Cloud Kitchen": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Malviya Nagar CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "RR Nagar CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Rohini CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Henjewadi": {"City": "Pune", "Format": "Cloud Kitchen"},
#     "Hesarghatta": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "hessaraghatta": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Embassy Lake": {"City": "Bangalore", "Format": "Café"},
#     "Chembur East CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Hesaraghatta Outlet": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Mahakali CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Noida Sector 46": {"City": "NCR", "Format": "Cloud Kitchen"},
#     "Noida Sector 83": {"City": "NCR", "Format": "Cloud Kitchen"},
#     "Velachery": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Smoor Deloitte": {"City": "Bangalore", "Format": "Kiosk"},
#     "Zirakpur": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "RBI Officers Canteen": {"City": "Bangalore", "Format": "Kiosk"},
# }

# # ── Phone cleaning ────────────────────────────────────────────────────────────
# def clean_phone(raw):
#     if pd.isna(raw):
#         return None
#     phone = str(raw).strip()
#     phone = re.sub(r'[\s\-\(\)\+\.]', '', phone)
#     phone = re.sub(r'\D', '', phone)
#     if not phone:
#         return None
#     if len(phone) == 12 and phone.startswith('91'):
#         phone = phone[2:]
#     if len(phone) == 11 and phone.startswith('0'):
#         phone = phone[1:]
#     if len(phone) != 10:
#         return None
#     for digit in set(phone):
#         if phone.count(digit) > 5:
#             return None
#     if phone[0] in ('0', '1'):
#         return None
#     if phone[0] not in ('6', '7', '8', '9'):
#         return None
#     return phone


# # ── Logging helper ────────────────────────────────────────────────────────────
# class Logger:
#     def __init__(self, log_path: str):
#         self.log_path = log_path
#         self.lines    = []

#     def log(self, text: str = ""):
#         self.lines.append(text)

#     def section(self, title: str):
#         self.log()
#         self.log("=" * 80)
#         self.log(f"  {title}")
#         self.log("=" * 80)

#     def subsection(self, title: str):
#         self.log()
#         self.log(f"  ── {title} ──")
#         self.log()

#     def save(self):
#         with open(self.log_path, 'w', encoding='utf-8') as f:
#             f.write('\n'.join(self.lines))
#         print(f"📋 Log saved → {self.log_path}")


# # ── Main analysis ─────────────────────────────────────────────────────────────
# def build_report(input_csv: str, output_xlsx: str):

#     # Setup logger
#     log_path = output_xlsx.replace('.xlsx', '_log.txt')
#     L = Logger(log_path)
#     run_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

#     L.log("=" * 80)
#     L.log("  PHONE NO RECORD TRACKER  —  DETAILED CALCULATION LOG")
#     L.log(f"  Run timestamp : {run_time}")
#     L.log(f"  Input file    : {input_csv}")
#     L.log(f"  Output file   : {output_xlsx}")
#     L.log("=" * 80)

#     # ── STEP 1: Load & filter ─────────────────────────────────────────────────
#     L.section("STEP 1 — LOAD & CHANNEL FILTER")
#     df_raw = pd.read_csv(input_csv, dtype=str, low_memory=False)
#     L.log(f"  Total rows loaded from CSV          : {len(df_raw):,}")
#     L.log(f"  Unique invoices in raw data         : {df_raw['invoiceNumber'].nunique():,}")

#     channel_counts = df_raw['channel'].str.strip().str.lower().value_counts()
#     L.log()
#     L.log("  Channel distribution in raw data:")
#     for ch, cnt in channel_counts.items():
#         L.log(f"    {ch:<30} : {cnt:,}")

#     df = df_raw[df_raw['channel'].str.strip().str.lower().isin(['dine in', 'take away', 'dine-in', 'takeaway'])].copy()
#     L.log()
#     L.log(f"  Rows KEPT after channel filter      : {len(df):,}")
#     L.log(f"  Rows DROPPED (other channels)       : {len(df_raw) - len(df):,}")
#     L.log(f"  Unique invoices after filter        : {df['invoiceNumber'].nunique():,}")

#     # ── STEP 2: Phone cleaning ────────────────────────────────────────────────
#     L.section("STEP 2 — PHONE NUMBER CLEANING")
#     df['clean_phone'] = df['customer_phoneNumber'].apply(clean_phone)

#     total_rows        = len(df)
#     raw_has_phone     = df['customer_phoneNumber'].notna().sum()
#     cleaned_has_phone = df['clean_phone'].notna().sum()
#     dropped_in_clean  = raw_has_phone - cleaned_has_phone

#     L.log(f"  Total rows (post channel filter)    : {total_rows:,}")
#     L.log(f"  Rows with ANY phone value           : {raw_has_phone:,}")
#     L.log(f"  Rows with VALID cleaned phone       : {cleaned_has_phone:,}")
#     L.log(f"  Rows where phone was REJECTED       : {dropped_in_clean:,}")
#     L.log()
#     L.log("  Phone rejection reasons (applied in order):")
#     L.log("    1. Empty / NaN value")
#     L.log("    2. No digits after stripping special chars")
#     L.log("    3. 12-digit starting with 91 → strip country code")
#     L.log("    4. 11-digit starting with 0  → strip leading zero")
#     L.log("    5. Not exactly 10 digits after cleaning")
#     L.log("    6. Any single digit repeats more than 5 times (e.g. 9999999999)")
#     L.log("    7. First digit is 0 or 1 (invalid Indian mobile)")
#     L.log("    8. First digit not in 6,7,8,9 (invalid Indian mobile prefix)")

#     # ── STEP 3: Deduplicate at invoice level ──────────────────────────────────
#     L.section("STEP 3 — DEDUPLICATE AT INVOICE LEVEL")
#     before_dedup = len(df)
#     inv = df.drop_duplicates(subset='invoiceNumber')[
#         ['invoiceNumber', 'branchName', 'channel', 'date', 'clean_phone']
#     ].copy()
#     inv['date'] = pd.to_datetime(inv['date'], errors='coerce').dt.date
#     after_dedup = len(inv)

#     L.log(f"  Rows before dedup                   : {before_dedup:,}")
#     L.log(f"  Unique invoices after dedup         : {after_dedup:,}")
#     L.log(f"  Duplicate rows removed              : {before_dedup - after_dedup:,}")
#     L.log()
#     L.log("  NOTE: When multiple rows share the same invoiceNumber,")
#     L.log("  only the FIRST occurrence is kept (drop_duplicates keeps first row).")
#     L.log("  This means phone number is taken from the first line item of each invoice.")

#     # ── STEP 4: City / Format mapping ─────────────────────────────────────────
#     L.section("STEP 4 — BRANCH MAPPING (City & Format)")
#     mapping_df = pd.DataFrame.from_dict(mapping_dict, orient='index').reset_index()
#     mapping_df.columns = ['branchName', 'City', 'Format']
#     inv = inv.merge(mapping_df, on='branchName', how='left')
#     inv['City']   = inv['City'].fillna('Unknown')
#     inv['Format'] = inv['Format'].fillna('Unknown')

#     unknown_branches = inv[inv['City'] == 'Unknown']['branchName'].unique()
#     L.log(f"  Branches mapped successfully        : {inv[inv['City'] != 'Unknown']['branchName'].nunique():,}")
#     if len(unknown_branches) > 0:
#         L.log(f"  Branches NOT found in mapping       : {len(unknown_branches)}")
#         for b in unknown_branches:
#             L.log(f"    → '{b}'  (will be DROPPED)")
#     else:
#         L.log("  All branches found in mapping dict  : ✓")

#     inv_before_filter = len(inv)
#     inv = inv[~inv['City'].isin(['Unknown', 'NA'])]
#     inv = inv[inv['Format'] != 'Cloud Kitchen']
#     inv_after_filter  = len(inv)

#     format_dist = inv['Format'].value_counts()
#     L.log()
#     L.log(f"  Invoices dropped (Unknown/NA city)  : {inv_before_filter - inv_after_filter:,}")
#     L.log(f"  Cloud Kitchen invoices excluded     : (filtered out entirely)")
#     L.log(f"  Invoices remaining for analysis     : {inv_after_filter:,}")
#     L.log()
#     L.log("  Format distribution (post-filter):")
#     for fmt, cnt in format_dist.items():
#         L.log(f"    {fmt:<20} : {cnt:,} invoices")

#     # ── STEP 5: Week / Day buckets ────────────────────────────────────────────
#     L.section("STEP 5 — WEEK & DAY BUCKETS")
#     inv['date_dt']  = pd.to_datetime(inv['date'], errors='coerce')
#     inv['iso_week'] = inv['date_dt'].dt.isocalendar().week.astype(int)
#     inv['iso_year'] = inv['date_dt'].dt.isocalendar().year.astype(int)
#     inv['yw']       = inv['iso_year'] * 100 + inv['iso_week']

#     last3_yw          = sorted(inv['yw'].dropna().unique())[-3:]
#     target_week_pairs = [(int(yw // 100), int(yw % 100)) for yw in last3_yw]
#     all_days          = sorted(inv['date'].dropna().unique())
#     target_days       = all_days[-3:] if len(all_days) >= 3 else all_days
#     latest_yw         = last3_yw[-1] if last3_yw else None

#     L.log(f"  Date range in data                  : {min(all_days)} → {max(all_days)}")
#     L.log(f"  Total unique days in data           : {len(all_days)}")
#     L.log()
#     L.log("  Last 3 ISO weeks selected:")
#     for (yr, wk) in target_week_pairs:
#         yw_key = yr * 100 + wk
#         label  = "← WTD (current week)" if yw_key == latest_yw else ""
#         wk_inv = inv[(inv['iso_week'] == wk) & (inv['iso_year'] == yr)]
#         L.log(f"    Year {yr}  Week {wk:02d}  (yw={yw_key})  →  {wk_inv['invoiceNumber'].nunique():,} invoices  {label}")
#     L.log()
#     L.log("  Last 3 days selected:")
#     for d in target_days:
#         day_inv = inv[inv['date'] == d]
#         L.log(f"    {d}  →  {day_inv['invoiceNumber'].nunique():,} invoices across {day_inv['branchName'].nunique()} branches")

#     # ── STEP 6: Branch-level metrics ──────────────────────────────────────────
#     L.section("STEP 6 — BRANCH-LEVEL CAPTURE RATE CALCULATIONS")
#     L.log()
#     L.log("  FORMULA:")
#     L.log("    Capture Rate = Unique valid phone numbers ÷ Unique invoices")
#     L.log("    (scoped to: one branch × one time window)")
#     L.log()

#     order_counts = inv.groupby('branchName')['invoiceNumber'].nunique().reset_index()
#     order_counts.columns = ['branchName', 'total_orders']

#     branches = (
#         inv[['branchName', 'City', 'Format']]
#         .drop_duplicates('branchName')
#         .merge(order_counts, on='branchName', how='left')
#         .sort_values('total_orders', ascending=False)
#         .reset_index(drop=True)
#     )

#     # ── FIXED: Unique valid phone numbers ÷ Unique invoices ──────────────────
#     def capture(sub):
#         total     = sub['invoiceNumber'].nunique()
#         has_phone = sub['clean_phone'].nunique()  # unique valid phone numbers
#         return has_phone / total if total > 0 else None

#     day_fmt   = lambda d: d.strftime('%b %#d') if os.name == 'nt' else d.strftime('%b %-d')

#     # Week cols
#     week_cols = []
#     L.subsection("Weekly Capture Rates by Branch")
#     for (yr, wk) in target_week_pairs:
#         yw_key = yr * 100 + wk
#         col    = f'W{wk} WTD' if yw_key == latest_yw else f'W{wk}'
#         week_cols.append(col)
#         L.log(f"  [{col}]  Year={yr}  Week={wk}")
#         L.log(f"  {'Branch':<30} {'Unique Invoices':>16} {'Unique Phones':>14} {'Capture Rate':>13}")
#         L.log(f"  {'-'*30} {'-'*16} {'-'*14} {'-'*13}")
#         rates = []
#         for _, row in branches.iterrows():
#             sub   = inv[(inv['branchName'] == row['branchName']) &
#                         (inv['iso_week'] == wk) & (inv['iso_year'] == yr)]
#             rate  = capture(sub)
#             rates.append(rate)
#             if len(sub) > 0:
#                 total_inv = sub['invoiceNumber'].nunique()
#                 uniq_ph   = sub['clean_phone'].nunique()
#                 rate_str  = f"{rate:.1%}" if rate is not None else "—"
#                 L.log(f"  {row['branchName']:<30} {total_inv:>16,} {uniq_ph:>14,} {rate_str:>13}")
#         branches[col] = rates
#         L.log()

#     # Day cols
#     day_cols = []
#     L.subsection("Daily Capture Rates by Branch")
#     for d in target_days:
#         col = day_fmt(d)
#         day_cols.append(col)
#         L.log(f"  [{col}]  Date={d}")
#         L.log(f"  {'Branch':<30} {'Unique Invoices':>16} {'Unique Phones':>14} {'Capture Rate':>13}")
#         L.log(f"  {'-'*30} {'-'*16} {'-'*14} {'-'*13}")
#         rates = []
#         for _, row in branches.iterrows():
#             sub   = inv[(inv['branchName'] == row['branchName']) & (inv['date'] == d)]
#             rate  = capture(sub)
#             rates.append(rate)
#             if len(sub) > 0:
#                 total_inv = sub['invoiceNumber'].nunique()
#                 uniq_ph   = sub['clean_phone'].nunique()
#                 rate_str  = f"{rate:.1%}" if rate is not None else "—"
#                 L.log(f"  {row['branchName']:<30} {total_inv:>16,} {uniq_ph:>14,} {rate_str:>13}")
#         branches[col] = rates
#         L.log()

#     metric_cols = week_cols + day_cols

#     # ── STEP 7: Overall summary ───────────────────────────────────────────────
#     L.section("STEP 7 — OVERALL SUMMARY")
#     total_invoices = inv['invoiceNumber'].nunique()
#     total_phones   = inv['clean_phone'].nunique()
#     overall_rate   = total_phones / total_invoices * 100 if total_invoices else 0

#     L.log(f"  Total unique invoices (Dine In/TA)  : {total_invoices:,}")
#     L.log(f"  Total unique valid phone numbers    : {total_phones:,}")
#     L.log(f"  Overall capture rate                : {overall_rate:.1f}%")
#     L.log(f"  Formula: {total_phones:,} ÷ {total_invoices:,} = {overall_rate:.1f}%")

#     L.subsection("Top 15 Most Frequent Phone Numbers (Dummy Number Check)")
#     top_phones = inv['clean_phone'].value_counts().head(15)
#     L.log(f"  {'Phone Number':<15} {'Count':>8}   {'Flag'}")
#     L.log(f"  {'-'*15} {'-'*8}   {'-'*30}")
#     for ph, cnt in top_phones.items():
#         flag = "⚠️  LIKELY DUMMY — appears very frequently" if cnt > 10 else ""
#         L.log(f"  {str(ph):<15} {cnt:>8,}   {flag}")

#     L.subsection("Capture Rate by City")
#     for city in sorted(inv['City'].unique()):
#         city_sub   = inv[inv['City'] == city]
#         city_inv   = city_sub['invoiceNumber'].nunique()
#         city_ph    = city_sub['clean_phone'].nunique()
#         city_rate  = city_ph / city_inv * 100 if city_inv else 0
#         L.log(f"  {city:<15}  Invoices={city_inv:,}  Unique Phones={city_ph:,}  Rate={city_rate:.1f}%")

#     L.subsection("Capture Rate by Format")
#     for fmt in sorted(inv['Format'].unique()):
#         fmt_sub  = inv[inv['Format'] == fmt]
#         fmt_inv  = fmt_sub['invoiceNumber'].nunique()
#         fmt_ph   = fmt_sub['clean_phone'].nunique()
#         fmt_rate = fmt_ph / fmt_inv * 100 if fmt_inv else 0
#         L.log(f"  {fmt:<15}  Invoices={fmt_inv:,}  Unique Phones={fmt_ph:,}  Rate={fmt_rate:.1f}%")

#     L.subsection("All Branches — Overall Capture Rate (sorted by invoices)")
#     L.log(f"  {'Branch':<30} {'City':<12} {'Format':<12} {'Invoices':>10} {'Uniq Phones':>12} {'Rate':>8}")
#     L.log(f"  {'-'*30} {'-'*12} {'-'*12} {'-'*10} {'-'*12} {'-'*8}")
#     for _, row in branches.iterrows():
#         b_sub  = inv[inv['branchName'] == row['branchName']]
#         b_inv  = b_sub['invoiceNumber'].nunique()
#         b_ph   = b_sub['clean_phone'].nunique()
#         b_rate = f"{b_ph/b_inv:.1%}" if b_inv else "—"
#         L.log(f"  {row['branchName']:<30} {row['City']:<12} {row['Format']:<12} {b_inv:>10,} {b_ph:>12,} {b_rate:>8}")

#     L.log()
#     L.log("=" * 80)
#     L.log("  END OF LOG")
#     L.log("=" * 80)
#     L.save()

#     # 7. Build Excel ────────────────────────────────────────────────────────────
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Phone No Record Tracker"

#     ROYAL_BLUE = "1A3A6B"
#     HDR_WEEK   = "2C5282"
#     HDR_DAY    = "1A3A6B"
#     WHITE      = "FFFFFF"
#     LIGHT_ROW  = "F7F9FC"
#     ALT_ROW    = "EEF2F9"
#     BORDER_C   = "C5CFE3"

#     thin = Side(style='thin', color=BORDER_C)
#     bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

#     total_cols      = 3 + len(metric_cols)
#     last_col_letter = get_column_letter(total_cols)

#     # Title row
#     ws.merge_cells(f'A1:{last_col_letter}1')
#     c = ws['A1']
#     c.value     = "Phone No Record Tracker  |  Dine In & Take Away"
#     c.font      = Font(name='Arial', bold=True, size=14, color=WHITE)
#     c.fill      = PatternFill('solid', start_color=ROYAL_BLUE)
#     c.alignment = Alignment(horizontal='center', vertical='center')
#     ws.row_dimensions[1].height = 30

#     ws.merge_cells(f'A2:{last_col_letter}2')
#     ws['A2'].fill = PatternFill('solid', start_color=ROYAL_BLUE)
#     ws.row_dimensions[2].height = 6

#     # Group header row (row 3)
#     for c_idx in [1, 2, 3]:
#         ws.cell(row=3, column=c_idx).fill = PatternFill('solid', start_color=ROYAL_BLUE)

#     ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=6)
#     wk_cell = ws.cell(row=3, column=4, value="← Last 3 Weeks →")
#     wk_cell.font      = Font(name='Arial', bold=True, size=9, color=WHITE)
#     wk_cell.fill      = PatternFill('solid', start_color=HDR_WEEK)
#     wk_cell.alignment = Alignment(horizontal='center')

#     ws.merge_cells(start_row=3, start_column=7, end_row=3, end_column=9)
#     dy_cell = ws.cell(row=3, column=7, value="← Last 3 Days →")
#     dy_cell.font      = Font(name='Arial', bold=True, size=9, color=WHITE)
#     dy_cell.fill      = PatternFill('solid', start_color=HDR_DAY)
#     dy_cell.alignment = Alignment(horizontal='center')

#     # Column label row (row 4)
#     headers  = ['City', 'Branch', 'Format'] + metric_cols
#     col_meta = {}
#     for h in ['City', 'Branch', 'Format']:
#         col_meta[h] = (ROYAL_BLUE, WHITE)
#     for c in week_cols:
#         col_meta[c] = (HDR_WEEK, WHITE)
#     for c in day_cols:
#         col_meta[c] = (HDR_DAY, WHITE)

#     for ci, h in enumerate(headers, start=1):
#         cell = ws.cell(row=4, column=ci, value=h)
#         bg, fg = col_meta.get(h, (ROYAL_BLUE, WHITE))
#         cell.fill      = PatternFill('solid', start_color=bg)
#         cell.font      = Font(name='Arial', bold=True, size=10, color=fg)
#         cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
#         cell.border    = bdr
#     ws.row_dimensions[4].height = 28

#     # Data rows
#     for ri, row in branches.iterrows():
#         excel_row = ri + 5
#         fill_bg   = LIGHT_ROW if ri % 2 == 0 else ALT_ROW
#         vals = [row['City'], row['branchName'], row['Format']] + [row[c] for c in metric_cols]
#         for ci, val in enumerate(vals, start=1):
#             cell        = ws.cell(row=excel_row, column=ci)
#             cell.border = bdr
#             if ci <= 3:
#                 cell.value     = val
#                 cell.font      = Font(name='Arial', size=9)
#                 cell.fill      = PatternFill('solid', start_color=fill_bg)
#                 cell.alignment = Alignment(horizontal='left', vertical='center')
#             else:
#                 if val is None:
#                     cell.value     = "—"
#                     cell.font      = Font(name='Arial', size=9, color='BBBBBB')
#                     cell.fill      = PatternFill('solid', start_color=fill_bg)
#                     cell.alignment = Alignment(horizontal='center')
#                 else:
#                     cell.value         = val
#                     cell.number_format = '0.0%'
#                     cell.font          = Font(name='Arial', size=9)
#                     cell.fill          = PatternFill('solid', start_color=fill_bg)
#                     cell.alignment     = Alignment(horizontal='center')

#     total_data_rows = len(branches)
#     last_row        = 4 + total_data_rows

#     # Conditional formatting: ONLY on day cols
#     day_col_start = 3 + len(week_cols) + 1
#     for ci_offset in range(len(day_cols)):
#         ci         = day_col_start + ci_offset
#         col_letter = get_column_letter(ci)
#         ws.conditional_formatting.add(
#             f"{col_letter}5:{col_letter}{last_row}",
#             ColorScaleRule(
#                 start_type='min', start_color='C0392B',
#                 end_type='max',   end_color='FFFFFF'
#             )
#         )

#     col_widths = [14, 28, 14, 9, 9, 9, 9, 9, 9]
#     for ci, w in enumerate(col_widths, start=1):
#         ws.column_dimensions[get_column_letter(ci)].width = w

#     ws.freeze_panes = 'D5'

#     # 8. Raw Dump sheet ─────────────────────────────────────────────────────────
#     ws2 = wb.create_sheet("Raw Dump")

#     extra_cols_wanted = ['invoiceDate', 'customer_name', 'customer_email',
#                          'customer_id', 'orderNumber', 'netAmount', 'grossAmount']
#     available_extra = [c for c in extra_cols_wanted if c in df.columns]

#     df['_qty'] = pd.to_numeric(df.get('item_quantity', 0), errors='coerce').fillna(0)
#     sale_qty   = df.groupby('invoiceNumber')['_qty'].sum().reset_index()
#     sale_qty.columns = ['invoiceNumber', 'sale_qty']

#     raw = df.drop_duplicates(subset='invoiceNumber')[
#         ['invoiceNumber', 'branchName', 'channel', 'date', 'clean_phone'] + available_extra
#     ].copy()
#     raw = raw.merge(mapping_df[['branchName', 'City', 'Format']], on='branchName', how='left')
#     raw = raw.merge(sale_qty, on='invoiceNumber', how='left')
#     raw['City']   = raw['City'].fillna('Unknown')
#     raw['Format'] = raw['Format'].fillna('Unknown')
#     raw['date']   = pd.to_datetime(raw['date'], errors='coerce').dt.date

#     # ── EXCLUDE Cloud Kitchen from Raw Dump too ───────────────────────────────
#     raw = raw[raw['Format'] != 'Cloud Kitchen']

#     for nc in ['netAmount', 'grossAmount']:
#         if nc in raw.columns:
#             raw[nc] = pd.to_numeric(raw[nc], errors='coerce')

#     raw_dump = (
#         raw[raw['clean_phone'].notna()]
#         .merge(order_counts, on='branchName', how='left')
#         .sort_values(['total_orders', 'date', 'branchName'], ascending=[False, True, True])
#         .drop_duplicates(subset='clean_phone', keep='first')  # ── one row per unique phone
#         .drop(columns=['total_orders'])
#         .reset_index(drop=True)
#     )

#     ordered_cols = ['invoiceNumber', 'date']
#     for c in ['invoiceDate', 'branchName', 'City', 'Format', 'channel',
#               'customer_name', 'clean_phone', 'customer_email', 'customer_id',
#               'orderNumber', 'sale_qty', 'netAmount', 'grossAmount']:
#         if c in raw_dump.columns:
#             ordered_cols.append(c)
#     raw_dump = raw_dump[ordered_cols]

#     header_labels = {
#         'invoiceNumber': 'Invoice No', 'date': 'Date', 'invoiceDate': 'Invoice Date',
#         'branchName': 'Branch', 'City': 'City', 'Format': 'Format', 'channel': 'Channel',
#         'customer_name': 'Customer Name', 'clean_phone': 'Phone Number',
#         'customer_email': 'Email', 'customer_id': 'Customer ID', 'orderNumber': 'Order No',
#         'sale_qty': 'Sale Qty', 'netAmount': 'Net Amount', 'grossAmount': 'Gross Amount',
#     }

#     numeric_cols_raw = {'netAmount', 'grossAmount', 'sale_qty'}
#     left_cols        = {'branchName', 'customer_name', 'customer_email'}
#     hdr_fill  = PatternFill('solid', start_color="1A3A6B")
#     hdr_font  = Font(name='Arial', bold=True, size=10, color="FFFFFF")
#     hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

#     for ci, col in enumerate(ordered_cols, start=1):
#         cell = ws2.cell(row=1, column=ci, value=header_labels.get(col, col))
#         cell.fill = hdr_fill; cell.font = hdr_font
#         cell.alignment = hdr_align; cell.border = bdr
#     ws2.row_dimensions[1].height = 28

#     for ri, row_data in raw_dump.iterrows():
#         excel_row = ri + 2
#         fill_bg   = "F9FAFB" if ri % 2 == 0 else "FFFFFF"
#         for ci, col in enumerate(ordered_cols, start=1):
#             val  = row_data[col]
#             cell = ws2.cell(row=excel_row, column=ci)
#             cell.fill   = PatternFill('solid', start_color=fill_bg)
#             cell.font   = Font(name='Arial', size=9)
#             cell.border = bdr
#             if col in numeric_cols_raw:
#                 try:
#                     cell.value = float(val) if val is not None and str(val) != 'nan' else None
#                 except (ValueError, TypeError):
#                     cell.value = None
#                 cell.number_format = '#,##0.00' if col in {'netAmount', 'grossAmount'} else '#,##0'
#                 cell.alignment = Alignment(horizontal='right', vertical='center')
#             elif col == 'clean_phone':
#                 cell.value         = str(val) if val is not None else None
#                 cell.number_format = '@'
#                 cell.alignment     = Alignment(horizontal='center', vertical='center')
#             else:
#                 cell.value     = val
#                 cell.alignment = Alignment(
#                     horizontal='left' if col in left_cols else 'center',
#                     vertical='center'
#                 )

#     col_widths2 = {
#         'invoiceNumber': 16, 'date': 12, 'invoiceDate': 14, 'branchName': 26,
#         'City': 12, 'Format': 14, 'channel': 12, 'customer_name': 20,
#         'clean_phone': 15, 'customer_email': 26, 'customer_id': 14,
#         'orderNumber': 14, 'sale_qty': 10, 'netAmount': 14, 'grossAmount': 14,
#     }
#     for ci, col in enumerate(ordered_cols, start=1):
#         ws2.column_dimensions[get_column_letter(ci)].width = col_widths2.get(col, 14)

#     ws2.freeze_panes = 'A2'

#     wb.save(output_xlsx)
#     phone_fill_pct = (inv['clean_phone'].nunique() / inv['invoiceNumber'].nunique() * 100) if len(inv) else 0
#     print(f"✅ Report saved      → {output_xlsx}")
#     print(f"📋 Log saved         → {log_path}")
#     print(f"   Branches analysed             : {len(branches)}")
#     print(f"   Total invoices (Dine In / TA) : {inv['invoiceNumber'].nunique():,}")
#     print(f"   Raw Dump rows (with phone)    : {len(raw_dump):,}")
#     print(f"   Overall phone capture         : {phone_fill_pct:.1f}%")


# # ── Entry point ───────────────────────────────────────────────────────────────
# if __name__ == "__main__":
#     import sys

#     ROLLING_CSV = r"C:/Users/jeryy/OneDrive - BLISS CHOCOLATES INDIA PRIVATE LIMITED/Desktop/Documents/Phone No/Data/sales_rolling.csv"
#     OUTPUT_DIR  = r"C:/Users/jeryy/OneDrive - BLISS CHOCOLATES INDIA PRIVATE LIMITED/Desktop/Documents/Phone No/output"

#     input_file = sys.argv[1] if len(sys.argv) > 1 else ROLLING_CSV
#     if len(sys.argv) > 2:
#         output_file = sys.argv[2]
#     else:
#         os.makedirs(OUTPUT_DIR, exist_ok=True)
#         date_str    = datetime.today().strftime("%Y-%m-%d")
#         output_file = os.path.join(OUTPUT_DIR, f"Phone No Record - {date_str}.xlsx")

#     build_report(input_file, output_file)












# import os
# import pandas as pd
# import re
# from datetime import datetime, date
# import openpyxl
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from openpyxl.formatting.rule import ColorScaleRule
# from openpyxl.utils import get_column_letter

# # ── Branch → City / Format mapping ──────────────────────────────────────────
# mapping_dict = {
#     "1 MG": {"City": "Bangalore", "Format": "Café"},
#     "1 Mg Cafe": {"City": "Bangalore", "Format": "Café"},
#     "1mg cafe": {"City": "Bangalore", "Format": "Café"},
#     "Adarsh Palm Retreat": {"City": "Bangalore", "Format": "Café"},
#     "Arumbakkam Cloud Kitchen": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "arumbakkam ck": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Avadi CH": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Bandra": {"City": "Mumbai", "Format": "Café"},
#     "Banni Square": {"City": "Delhi NCR", "Format": "Kiosk"},
#     "Basaveshwara Nagar CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "BEL Road": {"City": "Bangalore", "Format": "Café"},
#     "Bellandur": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Bhartiya City Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "BLR Airport T2 International Bakery": {"City": "Bangalore", "Format": "Airport"},
#     "BLR Airport Terminal 1": {"City": "Bangalore", "Format": "Airport"},
#     "Central Arcade DLF": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "dlf": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Chennai ECR": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Cummulus": {"City": "Bangalore", "Format": "Café"},
#     "Dwarka Cloud Kitchen": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "dwarka ck": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Electronic City": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Express Avenue SMOOR": {"City": "Chennai", "Format": "Kiosk"},
#     "express avenue": {"City": "Chennai", "Format": "Kiosk"},
#     "Forum Shantiniketan": {"City": "Bangalore", "Format": "Kiosk"},
#     "Forum South Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "Garuda Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "Godrej One": {"City": "Mumbai", "Format": "Kiosk"},
#     "Grant Road CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Head Office": {"City": "Bangalore", "Format": "NA"},
#     "Hesaraghatta": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Hinjewadi": {"City": "Pune", "Format": "Cloud Kitchen"},
#     "HSR": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "IC Colony CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Indiranagar": {"City": "Bangalore", "Format": "Signature"},
#     "Infinity Mall Malad": {"City": "Mumbai", "Format": "Kiosk"},
#     "J.P Nagar": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "j.p nagar": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Jayanagar": {"City": "Bangalore", "Format": "Café"},
#     "Kadugodi CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Kalkaji CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Kamanahalli": {"City": "Bangalore", "Format": "Café"},
#     "kammanahalli": {"City": "Bangalore", "Format": "Café"},
#     "Kharghar CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Koramangala": {"City": "Bangalore", "Format": "Café"},
#     "Koreagaon Park": {"City": "Pune", "Format": "Signature"},
#     "Lavelle Road": {"City": "Bangalore", "Format": "Signature"},
#     "Laxmi Nagar": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Magarpatta": {"City": "Pune", "Format": "Cloud Kitchen"},
#     "Mahadevpura": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Malad": {"City": "Mumbai", "Format": "Kiosk"},
#     "MAP Smoor": {"City": "Bangalore", "Format": "Café"},
#     "Marathahalli CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Medavakkam TN": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Mira Road CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Mukherjee Nagar": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Netaji Shubhash Palace": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "nsp": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Nexus Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "Noida Sec 46": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Noida Sec 83": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Orion Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "Perungudi CK": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Phoenix Mall Pune": {"City": "Pune", "Format": "Kiosk"},
#     "Powai": {"City": "Mumbai", "Format": "Signature"},
#     "Rajouri Garden": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "rajouri": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Ramapuram TN": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "ramapuram": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "RT Nagar CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Sadashivanagar": {"City": "Bangalore", "Format": "Café"},
#     "Sakinaka CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Sanpada": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Sarjapur CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Seawoods Mall": {"City": "Mumbai", "Format": "Café"},
#     "Smoor - Defence Colony CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "defence colony ck": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Test.Rebel": {"City": "NA", "Format": "NA"},
#     "Udyog Vihar CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Undri": {"City": "Pune", "Format": "Cloud Kitchen"},
#     "Uttarahalli CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Vandalur": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Vashi Smoor": {"City": "Mumbai", "Format": "Kiosk"},
#     "Vasundra CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Vega City": {"City": "Bangalore", "Format": "Kiosk"},
#     "Versova": {"City": "Mumbai", "Format": "Café"},
#     "Viviana Lounge Mumbai": {"City": "Mumbai", "Format": "Kiosk"},
#     "viviana": {"City": "Mumbai", "Format": "Kiosk"},
#     "West Tambaram": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Whitefield": {"City": "Bangalore", "Format": "Signature"},
#     "Whitefield Lounge": {"City": "Bangalore", "Format": "Signature"},
#     "Worli": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Yelahanka Cloud Kitchen": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "yelahanka": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Kandivali CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Airoli CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Ambience Mall": {"City": "Delhi NCR", "Format": "Kiosk"},
#     "Bhandup CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Green Park CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Gurugram sector 31 Cloud Kitchen": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Malviya Nagar CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "RR Nagar CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Rohini CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Henjewadi": {"City": "Pune", "Format": "Cloud Kitchen"},
#     "Hesarghatta": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "hessaraghatta": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Embassy Lake": {"City": "Bangalore", "Format": "Café"},
#     "Chembur East CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Hesaraghatta Outlet": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Mahakali CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Noida Sector 46": {"City": "NCR", "Format": "Cloud Kitchen"},
#     "Noida Sector 83": {"City": "NCR", "Format": "Cloud Kitchen"},
#     "Velachery": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Smoor Deloitte": {"City": "Bangalore", "Format": "Kiosk"},
#     "Zirakpur": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "RBI Officers Canteen": {"City": "Bangalore", "Format": "Kiosk"},
# }

# # ── Phone cleaning ────────────────────────────────────────────────────────────
# def clean_phone(raw):
#     if pd.isna(raw):
#         return None
#     phone = str(raw).strip()
#     phone = re.sub(r'[\s\-\(\)\+\.]', '', phone)
#     phone = re.sub(r'\D', '', phone)
#     if not phone:
#         return None
#     if len(phone) == 12 and phone.startswith('91'):
#         phone = phone[2:]
#     if len(phone) == 11 and phone.startswith('0'):
#         phone = phone[1:]
#     if len(phone) != 10:
#         return None
#     for digit in set(phone):
#         if phone.count(digit) > 5:
#             return None
#     if phone[0] in ('0', '1'):
#         return None
#     if phone[0] not in ('6', '7', '8', '9'):
#         return None
#     return phone


# # ── Logging helper ────────────────────────────────────────────────────────────
# class Logger:
#     def __init__(self, log_path: str):
#         self.log_path = log_path
#         self.lines    = []

#     def log(self, text: str = ""):
#         self.lines.append(text)

#     def section(self, title: str):
#         self.log()
#         self.log("=" * 80)
#         self.log(f"  {title}")
#         self.log("=" * 80)

#     def subsection(self, title: str):
#         self.log()
#         self.log(f"  ── {title} ──")
#         self.log()

#     def save(self):
#         with open(self.log_path, 'w', encoding='utf-8') as f:
#             f.write('\n'.join(self.lines))
#         print(f"📋 Log saved → {self.log_path}")


# # ── Main analysis ─────────────────────────────────────────────────────────────
# def build_report(input_csv: str, output_xlsx: str):

#     # Setup logger
#     log_path = output_xlsx.replace('.xlsx', '_log.txt')
#     L = Logger(log_path)
#     run_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

#     L.log("=" * 80)
#     L.log("  PHONE NO RECORD TRACKER  —  DETAILED CALCULATION LOG")
#     L.log(f"  Run timestamp : {run_time}")
#     L.log(f"  Input file    : {input_csv}")
#     L.log(f"  Output file   : {output_xlsx}")
#     L.log("=" * 80)

#     # ── STEP 1: Load & filter ─────────────────────────────────────────────────
#     L.section("STEP 1 — LOAD & CHANNEL FILTER")
#     df_raw = pd.read_csv(input_csv, dtype=str, low_memory=False)
#     L.log(f"  Total rows loaded from CSV          : {len(df_raw):,}")
#     L.log(f"  Unique invoices in raw data         : {df_raw['invoiceNumber'].nunique():,}")

#     channel_counts = df_raw['channel'].str.strip().str.lower().value_counts()
#     L.log()
#     L.log("  Channel distribution in raw data:")
#     for ch, cnt in channel_counts.items():
#         L.log(f"    {ch:<30} : {cnt:,}")

#     df = df_raw[df_raw['channel'].str.strip().str.lower().isin(['dine in', 'take away', 'dine-in', 'takeaway'])].copy()
#     L.log()
#     L.log(f"  Rows KEPT after channel filter      : {len(df):,}")
#     L.log(f"  Rows DROPPED (other channels)       : {len(df_raw) - len(df):,}")
#     L.log(f"  Unique invoices after filter        : {df['invoiceNumber'].nunique():,}")

#     # ── STEP 2: Phone cleaning ────────────────────────────────────────────────
#     L.section("STEP 2 — PHONE NUMBER CLEANING")
#     df['clean_phone'] = df['customer_phoneNumber'].apply(clean_phone)

#     total_rows        = len(df)
#     raw_has_phone     = df['customer_phoneNumber'].notna().sum()
#     cleaned_has_phone = df['clean_phone'].notna().sum()
#     dropped_in_clean  = raw_has_phone - cleaned_has_phone

#     L.log(f"  Total rows (post channel filter)    : {total_rows:,}")
#     L.log(f"  Rows with ANY phone value           : {raw_has_phone:,}")
#     L.log(f"  Rows with VALID cleaned phone       : {cleaned_has_phone:,}")
#     L.log(f"  Rows where phone was REJECTED       : {dropped_in_clean:,}")
#     L.log()
#     L.log("  Phone rejection reasons (applied in order):")
#     L.log("    1. Empty / NaN value")
#     L.log("    2. No digits after stripping special chars")
#     L.log("    3. 12-digit starting with 91 → strip country code")
#     L.log("    4. 11-digit starting with 0  → strip leading zero")
#     L.log("    5. Not exactly 10 digits after cleaning")
#     L.log("    6. Any single digit repeats more than 5 times (e.g. 9999999999)")
#     L.log("    7. First digit is 0 or 1 (invalid Indian mobile)")
#     L.log("    8. First digit not in 6,7,8,9 (invalid Indian mobile prefix)")

#     # ── STEP 3: Deduplicate at invoice level ──────────────────────────────────
#     L.section("STEP 3 — DEDUPLICATE AT INVOICE LEVEL")
#     before_dedup = len(df)
#     inv = df.drop_duplicates(subset='invoiceNumber')[
#         ['invoiceNumber', 'branchName', 'channel', 'date', 'clean_phone']
#     ].copy()
#     inv['date'] = pd.to_datetime(inv['date'], errors='coerce').dt.date
#     after_dedup = len(inv)

#     L.log(f"  Rows before dedup                   : {before_dedup:,}")
#     L.log(f"  Unique invoices after dedup         : {after_dedup:,}")
#     L.log(f"  Duplicate rows removed              : {before_dedup - after_dedup:,}")
#     L.log()
#     L.log("  NOTE: When multiple rows share the same invoiceNumber,")
#     L.log("  only the FIRST occurrence is kept (drop_duplicates keeps first row).")
#     L.log("  This means phone number is taken from the first line item of each invoice.")

#     # ── STEP 4: City / Format mapping ─────────────────────────────────────────
#     L.section("STEP 4 — BRANCH MAPPING (City & Format)")
#     mapping_df = pd.DataFrame.from_dict(mapping_dict, orient='index').reset_index()
#     mapping_df.columns = ['branchName', 'City', 'Format']
#     inv = inv.merge(mapping_df, on='branchName', how='left')
#     inv['City']   = inv['City'].fillna('Unknown')
#     inv['Format'] = inv['Format'].fillna('Unknown')

#     unknown_branches = inv[inv['City'] == 'Unknown']['branchName'].unique()
#     L.log(f"  Branches mapped successfully        : {inv[inv['City'] != 'Unknown']['branchName'].nunique():,}")
#     if len(unknown_branches) > 0:
#         L.log(f"  Branches NOT found in mapping       : {len(unknown_branches)}")
#         for b in unknown_branches:
#             L.log(f"    → '{b}'  (will be DROPPED)")
#     else:
#         L.log("  All branches found in mapping dict  : ✓")

#     inv_before_filter = len(inv)
#     inv = inv[~inv['City'].isin(['Unknown', 'NA'])]
#     inv = inv[inv['Format'] != 'Cloud Kitchen']
#     inv = inv[inv['branchName'] != 'RBI Officers Canteen']  # ── excluded outlet
#     inv_after_filter  = len(inv)

#     format_dist = inv['Format'].value_counts()
#     L.log()
#     L.log(f"  Invoices dropped (Unknown/NA city)  : {inv_before_filter - inv_after_filter:,}")
#     L.log(f"  Cloud Kitchen invoices excluded     : (filtered out entirely)")
#     L.log(f"  Invoices remaining for analysis     : {inv_after_filter:,}")
#     L.log()
#     L.log("  Format distribution (post-filter):")
#     for fmt, cnt in format_dist.items():
#         L.log(f"    {fmt:<20} : {cnt:,} invoices")

#     # ── STEP 5: Week / Day buckets ────────────────────────────────────────────
#     L.section("STEP 5 — WEEK & DAY BUCKETS")
#     inv['date_dt']  = pd.to_datetime(inv['date'], errors='coerce')
#     inv['iso_week'] = inv['date_dt'].dt.isocalendar().week.astype(int)
#     inv['iso_year'] = inv['date_dt'].dt.isocalendar().year.astype(int)
#     inv['yw']       = inv['iso_year'] * 100 + inv['iso_week']

#     last3_yw          = sorted(inv['yw'].dropna().unique())[-3:]
#     target_week_pairs = [(int(yw // 100), int(yw % 100)) for yw in last3_yw]
#     all_days          = sorted(inv['date'].dropna().unique())
#     target_days       = all_days[-3:] if len(all_days) >= 3 else all_days
#     latest_yw         = last3_yw[-1] if last3_yw else None

#     L.log(f"  Date range in data                  : {min(all_days)} → {max(all_days)}")
#     L.log(f"  Total unique days in data           : {len(all_days)}")
#     L.log()
#     L.log("  Last 3 ISO weeks selected:")
#     for (yr, wk) in target_week_pairs:
#         yw_key = yr * 100 + wk
#         label  = "← WTD (current week)" if yw_key == latest_yw else ""
#         wk_inv = inv[(inv['iso_week'] == wk) & (inv['iso_year'] == yr)]
#         L.log(f"    Year {yr}  Week {wk:02d}  (yw={yw_key})  →  {wk_inv['invoiceNumber'].nunique():,} invoices  {label}")
#     L.log()
#     L.log("  Last 3 days selected:")
#     for d in target_days:
#         day_inv = inv[inv['date'] == d]
#         L.log(f"    {d}  →  {day_inv['invoiceNumber'].nunique():,} invoices across {day_inv['branchName'].nunique()} branches")

#     # ── STEP 6: Branch-level metrics ──────────────────────────────────────────
#     L.section("STEP 6 — BRANCH-LEVEL CAPTURE RATE CALCULATIONS")
#     L.log()
#     L.log("  FORMULA:")
#     L.log("    Capture Rate = Unique valid phone numbers ÷ Unique invoices")
#     L.log("    (scoped to: one branch × one time window)")
#     L.log()

#     order_counts = inv.groupby('branchName')['invoiceNumber'].nunique().reset_index()
#     order_counts.columns = ['branchName', 'total_orders']

#     branches = (
#         inv[['branchName', 'City', 'Format']]
#         .drop_duplicates('branchName')
#         .merge(order_counts, on='branchName', how='left')
#         .sort_values('total_orders', ascending=False)
#         .reset_index(drop=True)
#     )

#     # ── FIXED: Unique valid phone numbers ÷ Unique invoices ──────────────────
#     def capture(sub):
#         total     = sub['invoiceNumber'].nunique()
#         has_phone = sub['clean_phone'].nunique()  # unique valid phone numbers
#         return has_phone / total if total > 0 else None

#     day_fmt   = lambda d: d.strftime('%b %#d') if os.name == 'nt' else d.strftime('%b %-d')

#     # Week cols
#     week_cols = []
#     L.subsection("Weekly Capture Rates by Branch")
#     for (yr, wk) in target_week_pairs:
#         yw_key = yr * 100 + wk
#         col    = f'W{wk} WTD' if yw_key == latest_yw else f'W{wk}'
#         week_cols.append(col)
#         L.log(f"  [{col}]  Year={yr}  Week={wk}")
#         L.log(f"  {'Branch':<30} {'Unique Invoices':>16} {'Unique Phones':>14} {'Capture Rate':>13}")
#         L.log(f"  {'-'*30} {'-'*16} {'-'*14} {'-'*13}")
#         rates = []
#         for _, row in branches.iterrows():
#             sub   = inv[(inv['branchName'] == row['branchName']) &
#                         (inv['iso_week'] == wk) & (inv['iso_year'] == yr)]
#             rate  = capture(sub)
#             rates.append(rate)
#             if len(sub) > 0:
#                 total_inv = sub['invoiceNumber'].nunique()
#                 uniq_ph   = sub['clean_phone'].nunique()
#                 rate_str  = f"{rate:.1%}" if rate is not None else "—"
#                 L.log(f"  {row['branchName']:<30} {total_inv:>16,} {uniq_ph:>14,} {rate_str:>13}")
#         branches[col] = rates
#         L.log()

#     # Day cols
#     day_cols = []
#     L.subsection("Daily Capture Rates by Branch")
#     for d in target_days:
#         col = day_fmt(d)
#         day_cols.append(col)
#         L.log(f"  [{col}]  Date={d}")
#         L.log(f"  {'Branch':<30} {'Unique Invoices':>16} {'Unique Phones':>14} {'Capture Rate':>13}")
#         L.log(f"  {'-'*30} {'-'*16} {'-'*14} {'-'*13}")
#         rates = []
#         for _, row in branches.iterrows():
#             sub   = inv[(inv['branchName'] == row['branchName']) & (inv['date'] == d)]
#             rate  = capture(sub)
#             rates.append(rate)
#             if len(sub) > 0:
#                 total_inv = sub['invoiceNumber'].nunique()
#                 uniq_ph   = sub['clean_phone'].nunique()
#                 rate_str  = f"{rate:.1%}" if rate is not None else "—"
#                 L.log(f"  {row['branchName']:<30} {total_inv:>16,} {uniq_ph:>14,} {rate_str:>13}")
#         branches[col] = rates
#         L.log()

#     metric_cols = week_cols + day_cols

#     # ── STEP 7: Overall summary ───────────────────────────────────────────────
#     L.section("STEP 7 — OVERALL SUMMARY")
#     total_invoices = inv['invoiceNumber'].nunique()
#     total_phones   = inv['clean_phone'].nunique()
#     overall_rate   = total_phones / total_invoices * 100 if total_invoices else 0

#     L.log(f"  Total unique invoices (Dine In/TA)  : {total_invoices:,}")
#     L.log(f"  Total unique valid phone numbers    : {total_phones:,}")
#     L.log(f"  Overall capture rate                : {overall_rate:.1f}%")
#     L.log(f"  Formula: {total_phones:,} ÷ {total_invoices:,} = {overall_rate:.1f}%")

#     L.subsection("Top 15 Most Frequent Phone Numbers (Dummy Number Check)")
#     top_phones = inv['clean_phone'].value_counts().head(15)
#     L.log(f"  {'Phone Number':<15} {'Count':>8}   {'Flag'}")
#     L.log(f"  {'-'*15} {'-'*8}   {'-'*30}")
#     for ph, cnt in top_phones.items():
#         flag = "⚠️  LIKELY DUMMY — appears very frequently" if cnt > 10 else ""
#         L.log(f"  {str(ph):<15} {cnt:>8,}   {flag}")

#     L.subsection("Capture Rate by City")
#     for city in sorted(inv['City'].unique()):
#         city_sub   = inv[inv['City'] == city]
#         city_inv   = city_sub['invoiceNumber'].nunique()
#         city_ph    = city_sub['clean_phone'].nunique()
#         city_rate  = city_ph / city_inv * 100 if city_inv else 0
#         L.log(f"  {city:<15}  Invoices={city_inv:,}  Unique Phones={city_ph:,}  Rate={city_rate:.1f}%")

#     L.subsection("Capture Rate by Format")
#     for fmt in sorted(inv['Format'].unique()):
#         fmt_sub  = inv[inv['Format'] == fmt]
#         fmt_inv  = fmt_sub['invoiceNumber'].nunique()
#         fmt_ph   = fmt_sub['clean_phone'].nunique()
#         fmt_rate = fmt_ph / fmt_inv * 100 if fmt_inv else 0
#         L.log(f"  {fmt:<15}  Invoices={fmt_inv:,}  Unique Phones={fmt_ph:,}  Rate={fmt_rate:.1f}%")

#     L.subsection("All Branches — Overall Capture Rate (sorted by invoices)")
#     L.log(f"  {'Branch':<30} {'City':<12} {'Format':<12} {'Invoices':>10} {'Uniq Phones':>12} {'Rate':>8}")
#     L.log(f"  {'-'*30} {'-'*12} {'-'*12} {'-'*10} {'-'*12} {'-'*8}")
#     for _, row in branches.iterrows():
#         b_sub  = inv[inv['branchName'] == row['branchName']]
#         b_inv  = b_sub['invoiceNumber'].nunique()
#         b_ph   = b_sub['clean_phone'].nunique()
#         b_rate = f"{b_ph/b_inv:.1%}" if b_inv else "—"
#         L.log(f"  {row['branchName']:<30} {row['City']:<12} {row['Format']:<12} {b_inv:>10,} {b_ph:>12,} {b_rate:>8}")

#     L.log()
#     L.log("=" * 80)
#     L.log("  END OF LOG")
#     L.log("=" * 80)
#     L.save()

#     # 7. Build Excel ────────────────────────────────────────────────────────────
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Phone No Record Tracker"

#     ROYAL_BLUE = "1A3A6B"
#     HDR_WEEK   = "2C5282"
#     HDR_DAY    = "1A3A6B"
#     WHITE      = "FFFFFF"
#     LIGHT_ROW  = "F7F9FC"
#     ALT_ROW    = "EEF2F9"
#     BORDER_C   = "C5CFE3"

#     thin = Side(style='thin', color=BORDER_C)
#     bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

#     total_cols      = 3 + len(metric_cols)
#     last_col_letter = get_column_letter(total_cols)

#     # Title row
#     ws.merge_cells(f'A1:{last_col_letter}1')
#     c = ws['A1']
#     c.value     = "Phone No Record Tracker  |  Dine In & Take Away"
#     c.font      = Font(name='Arial', bold=True, size=14, color=WHITE)
#     c.fill      = PatternFill('solid', start_color=ROYAL_BLUE)
#     c.alignment = Alignment(horizontal='center', vertical='center')
#     ws.row_dimensions[1].height = 30

#     ws.merge_cells(f'A2:{last_col_letter}2')
#     ws['A2'].fill = PatternFill('solid', start_color=ROYAL_BLUE)
#     ws.row_dimensions[2].height = 6

#     # Group header row (row 3)
#     for c_idx in [1, 2, 3]:
#         ws.cell(row=3, column=c_idx).fill = PatternFill('solid', start_color=ROYAL_BLUE)

#     ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=6)
#     wk_cell = ws.cell(row=3, column=4, value="← Last 3 Weeks →")
#     wk_cell.font      = Font(name='Arial', bold=True, size=9, color=WHITE)
#     wk_cell.fill      = PatternFill('solid', start_color=HDR_WEEK)
#     wk_cell.alignment = Alignment(horizontal='center')

#     ws.merge_cells(start_row=3, start_column=7, end_row=3, end_column=9)
#     dy_cell = ws.cell(row=3, column=7, value="← Last 3 Days →")
#     dy_cell.font      = Font(name='Arial', bold=True, size=9, color=WHITE)
#     dy_cell.fill      = PatternFill('solid', start_color=HDR_DAY)
#     dy_cell.alignment = Alignment(horizontal='center')

#     # Column label row (row 4)
#     headers  = ['City', 'Branch', 'Format'] + metric_cols
#     col_meta = {}
#     for h in ['City', 'Branch', 'Format']:
#         col_meta[h] = (ROYAL_BLUE, WHITE)
#     for c in week_cols:
#         col_meta[c] = (HDR_WEEK, WHITE)
#     for c in day_cols:
#         col_meta[c] = (HDR_DAY, WHITE)

#     for ci, h in enumerate(headers, start=1):
#         cell = ws.cell(row=4, column=ci, value=h)
#         bg, fg = col_meta.get(h, (ROYAL_BLUE, WHITE))
#         cell.fill      = PatternFill('solid', start_color=bg)
#         cell.font      = Font(name='Arial', bold=True, size=10, color=fg)
#         cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
#         cell.border    = bdr
#     ws.row_dimensions[4].height = 28

#     # Data rows
#     for ri, row in branches.iterrows():
#         excel_row = ri + 5
#         fill_bg   = LIGHT_ROW if ri % 2 == 0 else ALT_ROW
#         vals = [row['City'], row['branchName'], row['Format']] + [row[c] for c in metric_cols]
#         for ci, val in enumerate(vals, start=1):
#             cell        = ws.cell(row=excel_row, column=ci)
#             cell.border = bdr
#             if ci <= 3:
#                 cell.value     = val
#                 cell.font      = Font(name='Arial', size=9)
#                 cell.fill      = PatternFill('solid', start_color=fill_bg)
#                 cell.alignment = Alignment(horizontal='left', vertical='center')
#             else:
#                 if val is None:
#                     cell.value     = "—"
#                     cell.font      = Font(name='Arial', size=9, color='BBBBBB')
#                     cell.fill      = PatternFill('solid', start_color=fill_bg)
#                     cell.alignment = Alignment(horizontal='center')
#                 else:
#                     cell.value         = val
#                     cell.number_format = '0.0%'
#                     cell.font          = Font(name='Arial', size=9)
#                     cell.fill          = PatternFill('solid', start_color=fill_bg)
#                     cell.alignment     = Alignment(horizontal='center')

#     total_data_rows = len(branches)
#     last_row        = 4 + total_data_rows

#     # Conditional formatting: ONLY on day cols
#     day_col_start = 3 + len(week_cols) + 1
#     for ci_offset in range(len(day_cols)):
#         ci         = day_col_start + ci_offset
#         col_letter = get_column_letter(ci)
#         ws.conditional_formatting.add(
#             f"{col_letter}5:{col_letter}{last_row}",
#             ColorScaleRule(
#                 start_type='min', start_color='C0392B',
#                 end_type='max',   end_color='FFFFFF'
#             )
#         )

#     col_widths = [14, 28, 14, 9, 9, 9, 9, 9, 9]
#     for ci, w in enumerate(col_widths, start=1):
#         ws.column_dimensions[get_column_letter(ci)].width = w

#     ws.freeze_panes = 'D5'

#     # 8. Raw Dump sheet ─────────────────────────────────────────────────────────
#     ws2 = wb.create_sheet("Raw Dump")

#     extra_cols_wanted = ['invoiceDate', 'customer_name', 'customer_email',
#                          'customer_id', 'orderNumber', 'netAmount', 'grossAmount']
#     available_extra = [c for c in extra_cols_wanted if c in df.columns]

#     df['_qty'] = pd.to_numeric(df.get('item_quantity', 0), errors='coerce').fillna(0)
#     sale_qty   = df.groupby('invoiceNumber')['_qty'].sum().reset_index()
#     sale_qty.columns = ['invoiceNumber', 'sale_qty']

#     raw = df.drop_duplicates(subset='invoiceNumber')[
#         ['invoiceNumber', 'branchName', 'channel', 'date', 'clean_phone'] + available_extra
#     ].copy()
#     raw = raw.merge(mapping_df[['branchName', 'City', 'Format']], on='branchName', how='left')
#     raw = raw.merge(sale_qty, on='invoiceNumber', how='left')
#     raw['City']   = raw['City'].fillna('Unknown')
#     raw['Format'] = raw['Format'].fillna('Unknown')
#     raw['date']   = pd.to_datetime(raw['date'], errors='coerce').dt.date

#     # ── EXCLUDE Cloud Kitchen and specific outlets from Raw Dump ────────────
#     raw = raw[raw['Format'] != 'Cloud Kitchen']
#     raw = raw[raw['branchName'] != 'RBI Officers Canteen']  # ── excluded outlet

#     for nc in ['netAmount', 'grossAmount']:
#         if nc in raw.columns:
#             raw[nc] = pd.to_numeric(raw[nc], errors='coerce')

#     raw_dump = (
#         raw[raw['clean_phone'].notna()]
#         .merge(order_counts, on='branchName', how='left')
#         .sort_values(['total_orders', 'date', 'branchName'], ascending=[False, True, True])
#         .drop_duplicates(subset='clean_phone', keep='first')  # ── one row per unique phone
#         .drop(columns=['total_orders'])
#         .reset_index(drop=True)
#     )

#     ordered_cols = ['invoiceNumber', 'date']
#     for c in ['invoiceDate', 'branchName', 'City', 'Format', 'channel',
#               'customer_name', 'clean_phone', 'customer_email', 'customer_id',
#               'orderNumber', 'sale_qty', 'netAmount', 'grossAmount']:
#         if c in raw_dump.columns:
#             ordered_cols.append(c)
#     raw_dump = raw_dump[ordered_cols]

#     header_labels = {
#         'invoiceNumber': 'Invoice No', 'date': 'Date', 'invoiceDate': 'Invoice Date',
#         'branchName': 'Branch', 'City': 'City', 'Format': 'Format', 'channel': 'Channel',
#         'customer_name': 'Customer Name', 'clean_phone': 'Phone Number',
#         'customer_email': 'Email', 'customer_id': 'Customer ID', 'orderNumber': 'Order No',
#         'sale_qty': 'Sale Qty', 'netAmount': 'Net Amount', 'grossAmount': 'Gross Amount',
#     }

#     numeric_cols_raw = {'netAmount', 'grossAmount', 'sale_qty'}
#     left_cols        = {'branchName', 'customer_name', 'customer_email'}
#     hdr_fill  = PatternFill('solid', start_color="1A3A6B")
#     hdr_font  = Font(name='Arial', bold=True, size=10, color="FFFFFF")
#     hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

#     for ci, col in enumerate(ordered_cols, start=1):
#         cell = ws2.cell(row=1, column=ci, value=header_labels.get(col, col))
#         cell.fill = hdr_fill; cell.font = hdr_font
#         cell.alignment = hdr_align; cell.border = bdr
#     ws2.row_dimensions[1].height = 28

#     for ri, row_data in raw_dump.iterrows():
#         excel_row = ri + 2
#         fill_bg   = "F9FAFB" if ri % 2 == 0 else "FFFFFF"
#         for ci, col in enumerate(ordered_cols, start=1):
#             val  = row_data[col]
#             cell = ws2.cell(row=excel_row, column=ci)
#             cell.fill   = PatternFill('solid', start_color=fill_bg)
#             cell.font   = Font(name='Arial', size=9)
#             cell.border = bdr
#             if col in numeric_cols_raw:
#                 try:
#                     cell.value = float(val) if val is not None and str(val) != 'nan' else None
#                 except (ValueError, TypeError):
#                     cell.value = None
#                 cell.number_format = '#,##0.00' if col in {'netAmount', 'grossAmount'} else '#,##0'
#                 cell.alignment = Alignment(horizontal='right', vertical='center')
#             elif col == 'clean_phone':
#                 cell.value         = str(val) if val is not None else None
#                 cell.number_format = '@'
#                 cell.alignment     = Alignment(horizontal='center', vertical='center')
#             else:
#                 cell.value     = val
#                 cell.alignment = Alignment(
#                     horizontal='left' if col in left_cols else 'center',
#                     vertical='center'
#                 )

#     col_widths2 = {
#         'invoiceNumber': 16, 'date': 12, 'invoiceDate': 14, 'branchName': 26,
#         'City': 12, 'Format': 14, 'channel': 12, 'customer_name': 20,
#         'clean_phone': 15, 'customer_email': 26, 'customer_id': 14,
#         'orderNumber': 14, 'sale_qty': 10, 'netAmount': 14, 'grossAmount': 14,
#     }
#     for ci, col in enumerate(ordered_cols, start=1):
#         ws2.column_dimensions[get_column_letter(ci)].width = col_widths2.get(col, 14)

#     ws2.freeze_panes = 'A2'

#     wb.save(output_xlsx)
#     phone_fill_pct = (inv['clean_phone'].nunique() / inv['invoiceNumber'].nunique() * 100) if len(inv) else 0
#     print(f"✅ Report saved      → {output_xlsx}")
#     print(f"📋 Log saved         → {log_path}")
#     print(f"   Branches analysed             : {len(branches)}")
#     print(f"   Total invoices (Dine In / TA) : {inv['invoiceNumber'].nunique():,}")
#     print(f"   Raw Dump rows (with phone)    : {len(raw_dump):,}")
#     print(f"   Overall phone capture         : {phone_fill_pct:.1f}%")


# # ── Entry point ───────────────────────────────────────────────────────────────
# if __name__ == "__main__":
#     import sys

#     ROLLING_CSV = r"C:/Users/jeryy/OneDrive - BLISS CHOCOLATES INDIA PRIVATE LIMITED/Desktop/Documents/Phone No/Data/sales_rolling.csv"
#     OUTPUT_DIR  = r"C:/Users/jeryy/OneDrive - BLISS CHOCOLATES INDIA PRIVATE LIMITED/Desktop/Documents/Phone No/output"

#     input_file = sys.argv[1] if len(sys.argv) > 1 else ROLLING_CSV
#     if len(sys.argv) > 2:
#         output_file = sys.argv[2]
#     else:
#         os.makedirs(OUTPUT_DIR, exist_ok=True)
#         from datetime import datetime, timedelta

#         date_str = (datetime.today() - timedelta(days=1)).strftime("%Y-%m-%d")
#         output_file = os.path.join(OUTPUT_DIR, f"Phone No Record - {date_str}.xlsx")

#     build_report(input_file, output_file)







# import os
# import pandas as pd
# import re
# from datetime import datetime, date
# import openpyxl
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from openpyxl.formatting.rule import ColorScaleRule
# from openpyxl.utils import get_column_letter

# # ── Branch → City / Format mapping ──────────────────────────────────────────
# mapping_dict = {
#     "1 MG": {"City": "Bangalore", "Format": "Café"},
#     "1 Mg Cafe": {"City": "Bangalore", "Format": "Café"},
#     "1mg cafe": {"City": "Bangalore", "Format": "Café"},
#     "Adarsh Palm Retreat": {"City": "Bangalore", "Format": "Café"},
#     "Arumbakkam Cloud Kitchen": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "arumbakkam ck": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Avadi CH": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Bandra": {"City": "Mumbai", "Format": "Café"},
#     "Banni Square": {"City": "Delhi NCR", "Format": "Kiosk"},
#     "Basaveshwara Nagar CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "BEL Road": {"City": "Bangalore", "Format": "Café"},
#     "Bellandur": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Bhartiya City Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "BLR Airport T2 International Bakery": {"City": "Bangalore", "Format": "Airport"},
#     "BLR Airport Terminal 1": {"City": "Bangalore", "Format": "Airport"},
#     "Central Arcade DLF": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "dlf": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Chennai ECR": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Cummulus": {"City": "Bangalore", "Format": "Café"},
#     "Dwarka Cloud Kitchen": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "dwarka ck": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Electronic City": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Express Avenue SMOOR": {"City": "Chennai", "Format": "Kiosk"},
#     "express avenue": {"City": "Chennai", "Format": "Kiosk"},
#     "Forum Shantiniketan": {"City": "Bangalore", "Format": "Kiosk"},
#     "Forum South Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "Garuda Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "Godrej One": {"City": "Mumbai", "Format": "Kiosk"},
#     "Grant Road CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Head Office": {"City": "Bangalore", "Format": "NA"},
#     "Hesaraghatta": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Hinjewadi": {"City": "Pune", "Format": "Cloud Kitchen"},
#     "HSR": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "IC Colony CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Indiranagar": {"City": "Bangalore", "Format": "Signature"},
#     "Infinity Mall Malad": {"City": "Mumbai", "Format": "Kiosk"},
#     "J.P Nagar": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "j.p nagar": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Jayanagar": {"City": "Bangalore", "Format": "Café"},
#     "Kadugodi CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Kalkaji CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Kamanahalli": {"City": "Bangalore", "Format": "Café"},
#     "kammanahalli": {"City": "Bangalore", "Format": "Café"},
#     "Kharghar CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Koramangala": {"City": "Bangalore", "Format": "Café"},
#     "Koreagaon Park": {"City": "Pune", "Format": "Signature"},
#     "Lavelle Road": {"City": "Bangalore", "Format": "Signature"},
#     "Laxmi Nagar": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Magarpatta": {"City": "Pune", "Format": "Cloud Kitchen"},
#     "Mahadevpura": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Malad": {"City": "Mumbai", "Format": "Kiosk"},
#     "MAP Smoor": {"City": "Bangalore", "Format": "Café"},
#     "Marathahalli CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Medavakkam TN": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Mira Road CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Mukherjee Nagar": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Netaji Shubhash Palace": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "nsp": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Nexus Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "Noida Sec 46": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Noida Sec 83": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Orion Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "Perungudi CK": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Phoenix Mall Pune": {"City": "Pune", "Format": "Kiosk"},
#     "Powai": {"City": "Mumbai", "Format": "Signature"},
#     "Rajouri Garden": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "rajouri": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Ramapuram TN": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "ramapuram": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "RT Nagar CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Sadashivanagar": {"City": "Bangalore", "Format": "Café"},
#     "Sakinaka CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Sanpada": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Sarjapur CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Seawoods Mall": {"City": "Mumbai", "Format": "Café"},
#     "Smoor - Defence Colony CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "defence colony ck": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Test.Rebel": {"City": "NA", "Format": "NA"},
#     "Udyog Vihar CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Undri": {"City": "Pune", "Format": "Cloud Kitchen"},
#     "Uttarahalli CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Vandalur": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Vashi Smoor": {"City": "Mumbai", "Format": "Kiosk"},
#     "Vasundra CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Vega City": {"City": "Bangalore", "Format": "Kiosk"},
#     "Versova": {"City": "Mumbai", "Format": "Café"},
#     "Viviana Lounge Mumbai": {"City": "Mumbai", "Format": "Kiosk"},
#     "viviana": {"City": "Mumbai", "Format": "Kiosk"},
#     "West Tambaram": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Whitefield": {"City": "Bangalore", "Format": "Signature"},
#     "Whitefield Lounge": {"City": "Bangalore", "Format": "Signature"},
#     "Worli": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Yelahanka Cloud Kitchen": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "yelahanka": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Kandivali CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Airoli CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Ambience Mall": {"City": "Delhi NCR", "Format": "Kiosk"},
#     "Bhandup CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Green Park CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Gurugram sector 31 Cloud Kitchen": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Malviya Nagar CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "RR Nagar CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Rohini CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Henjewadi": {"City": "Pune", "Format": "Cloud Kitchen"},
#     "Hesarghatta": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "hessaraghatta": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Embassy Lake": {"City": "Bangalore", "Format": "Café"},
#     "Chembur East CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Hesaraghatta Outlet": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Mahakali CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Noida Sector 46": {"City": "NCR", "Format": "Cloud Kitchen"},
#     "Noida Sector 83": {"City": "NCR", "Format": "Cloud Kitchen"},
#     "Velachery": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Smoor Deloitte": {"City": "Bangalore", "Format": "Kiosk"},
#     "Zirakpur": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "RBI Officers Canteen": {"City": "Bangalore", "Format": "Kiosk"},
# }

# # ── Phone cleaning ────────────────────────────────────────────────────────────
# def clean_phone(raw):
#     if pd.isna(raw):
#         return None
#     phone = str(raw).strip()
#     phone = re.sub(r'[\s\-\(\)\+\.]', '', phone)
#     phone = re.sub(r'\D', '', phone)
#     if not phone:
#         return None
#     if len(phone) == 12 and phone.startswith('91'):
#         phone = phone[2:]
#     if len(phone) == 11 and phone.startswith('0'):
#         phone = phone[1:]
#     if len(phone) != 10:
#         return None
#     for digit in set(phone):
#         if phone.count(digit) > 5:
#             return None
#     if phone[0] in ('0', '1'):
#         return None
#     if phone[0] not in ('6', '7', '8', '9'):
#         return None
#     return phone


# # ── Logging helper ────────────────────────────────────────────────────────────
# class Logger:
#     def __init__(self, log_path: str):
#         self.log_path = log_path
#         self.lines    = []

#     def log(self, text: str = ""):
#         self.lines.append(text)

#     def section(self, title: str):
#         self.log()
#         self.log("=" * 80)
#         self.log(f"  {title}")
#         self.log("=" * 80)

#     def subsection(self, title: str):
#         self.log()
#         self.log(f"  ── {title} ──")
#         self.log()

#     def save(self):
#         with open(self.log_path, 'w', encoding='utf-8') as f:
#             f.write('\n'.join(self.lines))
#         print(f"📋 Log saved → {self.log_path}")


# # ── Main analysis ─────────────────────────────────────────────────────────────
# def build_report(input_csv: str, output_xlsx: str):

#     # Setup logger
#     log_path = output_xlsx.replace('.xlsx', '_log.txt')
#     L = Logger(log_path)
#     run_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

#     L.log("=" * 80)
#     L.log("  PHONE NO RECORD TRACKER  —  DETAILED CALCULATION LOG")
#     L.log(f"  Run timestamp : {run_time}")
#     L.log(f"  Input file    : {input_csv}")
#     L.log(f"  Output file   : {output_xlsx}")
#     L.log("=" * 80)

#     # ── STEP 1: Load & filter ─────────────────────────────────────────────────
#     L.section("STEP 1 — LOAD & CHANNEL FILTER")
#     df_raw = pd.read_csv(input_csv, dtype=str, low_memory=False)
#     L.log(f"  Total rows loaded from CSV          : {len(df_raw):,}")
#     L.log(f"  Unique invoices in raw data         : {df_raw['invoiceNumber'].nunique():,}")

#     channel_counts = df_raw['channel'].str.strip().str.lower().value_counts()
#     L.log()
#     L.log("  Channel distribution in raw data:")
#     for ch, cnt in channel_counts.items():
#         L.log(f"    {ch:<30} : {cnt:,}")

#     df = df_raw[df_raw['channel'].str.strip().str.lower().isin(['dine in', 'take away', 'dine-in', 'takeaway'])].copy()
#     L.log()
#     L.log(f"  Rows KEPT after channel filter      : {len(df):,}")
#     L.log(f"  Rows DROPPED (other channels)       : {len(df_raw) - len(df):,}")
#     L.log(f"  Unique invoices after filter        : {df['invoiceNumber'].nunique():,}")

#     # ── STEP 2: Phone cleaning ────────────────────────────────────────────────
#     L.section("STEP 2 — PHONE NUMBER CLEANING")
#     df['clean_phone'] = df['customer_phoneNumber'].apply(clean_phone)

#     total_rows        = len(df)
#     raw_has_phone     = df['customer_phoneNumber'].notna().sum()
#     cleaned_has_phone = df['clean_phone'].notna().sum()
#     dropped_in_clean  = raw_has_phone - cleaned_has_phone

#     L.log(f"  Total rows (post channel filter)    : {total_rows:,}")
#     L.log(f"  Rows with ANY phone value           : {raw_has_phone:,}")
#     L.log(f"  Rows with VALID cleaned phone       : {cleaned_has_phone:,}")
#     L.log(f"  Rows where phone was REJECTED       : {dropped_in_clean:,}")
#     L.log()
#     L.log("  Phone rejection reasons (applied in order):")
#     L.log("    1. Empty / NaN value")
#     L.log("    2. No digits after stripping special chars")
#     L.log("    3. 12-digit starting with 91 → strip country code")
#     L.log("    4. 11-digit starting with 0  → strip leading zero")
#     L.log("    5. Not exactly 10 digits after cleaning")
#     L.log("    6. Any single digit repeats more than 5 times (e.g. 9999999999)")
#     L.log("    7. First digit is 0 or 1 (invalid Indian mobile)")
#     L.log("    8. First digit not in 6,7,8,9 (invalid Indian mobile prefix)")

#     # ── STEP 3: Deduplicate at invoice level ──────────────────────────────────
#     L.section("STEP 3 — DEDUPLICATE AT INVOICE LEVEL")
#     before_dedup = len(df)
#     inv = df.drop_duplicates(subset='invoiceNumber')[
#         ['invoiceNumber', 'branchName', 'channel', 'date', 'clean_phone']
#     ].copy()
#     inv['date'] = pd.to_datetime(inv['date'], errors='coerce').dt.date
#     after_dedup = len(inv)

#     L.log(f"  Rows before dedup                   : {before_dedup:,}")
#     L.log(f"  Unique invoices after dedup         : {after_dedup:,}")
#     L.log(f"  Duplicate rows removed              : {before_dedup - after_dedup:,}")
#     L.log()
#     L.log("  NOTE: When multiple rows share the same invoiceNumber,")
#     L.log("  only the FIRST occurrence is kept (drop_duplicates keeps first row).")
#     L.log("  This means phone number is taken from the first line item of each invoice.")

#     # ── STEP 4: City / Format mapping ─────────────────────────────────────────
#     L.section("STEP 4 — BRANCH MAPPING (City & Format)")
#     mapping_df = pd.DataFrame.from_dict(mapping_dict, orient='index').reset_index()
#     mapping_df.columns = ['branchName', 'City', 'Format']
#     inv = inv.merge(mapping_df, on='branchName', how='left')
#     inv['City']   = inv['City'].fillna('Unknown')
#     inv['Format'] = inv['Format'].fillna('Unknown')

#     unknown_branches = inv[inv['City'] == 'Unknown']['branchName'].unique()
#     L.log(f"  Branches mapped successfully        : {inv[inv['City'] != 'Unknown']['branchName'].nunique():,}")
#     if len(unknown_branches) > 0:
#         L.log(f"  Branches NOT found in mapping       : {len(unknown_branches)}")
#         for b in unknown_branches:
#             L.log(f"    → '{b}'  (will be DROPPED)")
#     else:
#         L.log("  All branches found in mapping dict  : ✓")

#     inv_before_filter = len(inv)
#     inv = inv[~inv['City'].isin(['Unknown', 'NA'])]
#     inv = inv[inv['Format'] != 'Cloud Kitchen']
#     inv = inv[inv['branchName'] != 'RBI Officers Canteen']  # ── excluded outlet
#     inv_after_filter  = len(inv)

#     format_dist = inv['Format'].value_counts()
#     L.log()
#     L.log(f"  Invoices dropped (Unknown/NA city)  : {inv_before_filter - inv_after_filter:,}")
#     L.log(f"  Cloud Kitchen invoices excluded     : (filtered out entirely)")
#     L.log(f"  Invoices remaining for analysis     : {inv_after_filter:,}")
#     L.log()
#     L.log("  Format distribution (post-filter):")
#     for fmt, cnt in format_dist.items():
#         L.log(f"    {fmt:<20} : {cnt:,} invoices")

#     # ── STEP 5: Week / Day buckets ────────────────────────────────────────────
#     L.section("STEP 5 — WEEK & DAY BUCKETS")
#     inv['date_dt']  = pd.to_datetime(inv['date'], errors='coerce')
#     inv['iso_week'] = inv['date_dt'].dt.isocalendar().week.astype(int)
#     inv['iso_year'] = inv['date_dt'].dt.isocalendar().year.astype(int)
#     inv['yw']       = inv['iso_year'] * 100 + inv['iso_week']

#     last3_yw          = sorted(inv['yw'].dropna().unique())[-3:]
#     target_week_pairs = [(int(yw // 100), int(yw % 100)) for yw in last3_yw]
#     all_days          = sorted(inv['date'].dropna().unique())
#     target_days       = all_days[-3:] if len(all_days) >= 3 else all_days
#     latest_yw         = last3_yw[-1] if last3_yw else None

#     L.log(f"  Date range in data                  : {min(all_days)} → {max(all_days)}")
#     L.log(f"  Total unique days in data           : {len(all_days)}")
#     L.log()
#     L.log("  Last 3 ISO weeks selected:")
#     for (yr, wk) in target_week_pairs:
#         yw_key = yr * 100 + wk
#         label  = "← WTD (current week)" if yw_key == latest_yw else ""
#         wk_inv = inv[(inv['iso_week'] == wk) & (inv['iso_year'] == yr)]
#         L.log(f"    Year {yr}  Week {wk:02d}  (yw={yw_key})  →  {wk_inv['invoiceNumber'].nunique():,} invoices  {label}")
#     L.log()
#     L.log("  Last 3 days selected:")
#     for d in target_days:
#         day_inv = inv[inv['date'] == d]
#         L.log(f"    {d}  →  {day_inv['invoiceNumber'].nunique():,} invoices across {day_inv['branchName'].nunique()} branches")

#     # ── STEP 6: Branch-level metrics ──────────────────────────────────────────
#     L.section("STEP 6 — BRANCH-LEVEL CAPTURE RATE CALCULATIONS")
#     L.log()
#     L.log("  FORMULA:")
#     L.log("    Capture Rate = Unique valid phone numbers ÷ Unique invoices")
#     L.log("    (scoped to: one branch × one time window)")
#     L.log()

#     order_counts = inv.groupby('branchName')['invoiceNumber'].nunique().reset_index()
#     order_counts.columns = ['branchName', 'total_orders']

#     branches = (
#         inv[['branchName', 'City', 'Format']]
#         .drop_duplicates('branchName')
#         .merge(order_counts, on='branchName', how='left')
#         .sort_values('total_orders', ascending=False)
#         .reset_index(drop=True)
#     )

#     # ── FIXED: Unique valid phone numbers ÷ Unique invoices ──────────────────
#     def capture(sub):
#         total     = sub['invoiceNumber'].nunique()
#         has_phone = sub['clean_phone'].nunique()  # unique valid phone numbers
#         return has_phone / total if total > 0 else None

#     day_fmt   = lambda d: d.strftime('%b %#d') if os.name == 'nt' else d.strftime('%b %-d')

#     # Week cols
#     week_cols = []
#     L.subsection("Weekly Capture Rates by Branch")
#     for (yr, wk) in target_week_pairs:
#         yw_key = yr * 100 + wk
#         col    = f'W{wk} WTD' if yw_key == latest_yw else f'W{wk}'
#         week_cols.append(col)
#         L.log(f"  [{col}]  Year={yr}  Week={wk}")
#         L.log(f"  {'Branch':<30} {'Unique Invoices':>16} {'Unique Phones':>14} {'Capture Rate':>13}")
#         L.log(f"  {'-'*30} {'-'*16} {'-'*14} {'-'*13}")
#         rates = []
#         for _, row in branches.iterrows():
#             sub   = inv[(inv['branchName'] == row['branchName']) &
#                         (inv['iso_week'] == wk) & (inv['iso_year'] == yr)]
#             rate  = capture(sub)
#             rates.append(rate)
#             if len(sub) > 0:
#                 total_inv = sub['invoiceNumber'].nunique()
#                 uniq_ph   = sub['clean_phone'].nunique()
#                 rate_str  = f"{rate:.1%}" if rate is not None else "—"
#                 L.log(f"  {row['branchName']:<30} {total_inv:>16,} {uniq_ph:>14,} {rate_str:>13}")
#         branches[col] = rates
#         L.log()

#     # Day cols
#     day_cols = []
#     L.subsection("Daily Capture Rates by Branch")
#     for d in target_days:
#         col = day_fmt(d)
#         day_cols.append(col)
#         L.log(f"  [{col}]  Date={d}")
#         L.log(f"  {'Branch':<30} {'Unique Invoices':>16} {'Unique Phones':>14} {'Capture Rate':>13}")
#         L.log(f"  {'-'*30} {'-'*16} {'-'*14} {'-'*13}")
#         rates = []
#         for _, row in branches.iterrows():
#             sub   = inv[(inv['branchName'] == row['branchName']) & (inv['date'] == d)]
#             rate  = capture(sub)
#             rates.append(rate)
#             if len(sub) > 0:
#                 total_inv = sub['invoiceNumber'].nunique()
#                 uniq_ph   = sub['clean_phone'].nunique()
#                 rate_str  = f"{rate:.1%}" if rate is not None else "—"
#                 L.log(f"  {row['branchName']:<30} {total_inv:>16,} {uniq_ph:>14,} {rate_str:>13}")
#         branches[col] = rates
#         L.log()

#     metric_cols = week_cols + day_cols

#     # ── Compute totals row: unique phones ÷ unique invoices per time window ──
#     totals = {'City': 'ALL', 'branchName': 'TOTAL', 'Format': ''}
#     for (yr, wk) in target_week_pairs:
#         yw_key = yr * 100 + wk
#         col    = f'W{wk} WTD' if yw_key == latest_yw else f'W{wk}'
#         sub    = inv[(inv['iso_week'] == wk) & (inv['iso_year'] == yr)]
#         t_inv  = sub['invoiceNumber'].nunique()
#         t_ph   = sub['clean_phone'].nunique()
#         totals[col] = (t_ph / t_inv) if t_inv > 0 else None
#     for d in target_days:
#         col   = day_fmt(d)
#         sub   = inv[inv['date'] == d]
#         t_inv = sub['invoiceNumber'].nunique()
#         t_ph  = sub['clean_phone'].nunique()
#         totals[col] = (t_ph / t_inv) if t_inv > 0 else None

#     # ── STEP 7: Overall summary ───────────────────────────────────────────────
#     L.section("STEP 7 — OVERALL SUMMARY")
#     total_invoices = inv['invoiceNumber'].nunique()
#     total_phones   = inv['clean_phone'].nunique()
#     overall_rate   = total_phones / total_invoices * 100 if total_invoices else 0

#     L.log(f"  Total unique invoices (Dine In/TA)  : {total_invoices:,}")
#     L.log(f"  Total unique valid phone numbers    : {total_phones:,}")
#     L.log(f"  Overall capture rate                : {overall_rate:.1f}%")
#     L.log(f"  Formula: {total_phones:,} ÷ {total_invoices:,} = {overall_rate:.1f}%")

#     L.subsection("Top 15 Most Frequent Phone Numbers (Dummy Number Check)")
#     top_phones = inv['clean_phone'].value_counts().head(15)
#     L.log(f"  {'Phone Number':<15} {'Count':>8}   {'Flag'}")
#     L.log(f"  {'-'*15} {'-'*8}   {'-'*30}")
#     for ph, cnt in top_phones.items():
#         flag = "⚠️  LIKELY DUMMY — appears very frequently" if cnt > 10 else ""
#         L.log(f"  {str(ph):<15} {cnt:>8,}   {flag}")

#     L.subsection("Capture Rate by City")
#     for city in sorted(inv['City'].unique()):
#         city_sub   = inv[inv['City'] == city]
#         city_inv   = city_sub['invoiceNumber'].nunique()
#         city_ph    = city_sub['clean_phone'].nunique()
#         city_rate  = city_ph / city_inv * 100 if city_inv else 0
#         L.log(f"  {city:<15}  Invoices={city_inv:,}  Unique Phones={city_ph:,}  Rate={city_rate:.1f}%")

#     L.subsection("Capture Rate by Format")
#     for fmt in sorted(inv['Format'].unique()):
#         fmt_sub  = inv[inv['Format'] == fmt]
#         fmt_inv  = fmt_sub['invoiceNumber'].nunique()
#         fmt_ph   = fmt_sub['clean_phone'].nunique()
#         fmt_rate = fmt_ph / fmt_inv * 100 if fmt_inv else 0
#         L.log(f"  {fmt:<15}  Invoices={fmt_inv:,}  Unique Phones={fmt_ph:,}  Rate={fmt_rate:.1f}%")

#     L.subsection("All Branches — Overall Capture Rate (sorted by invoices)")
#     L.log(f"  {'Branch':<30} {'City':<12} {'Format':<12} {'Invoices':>10} {'Uniq Phones':>12} {'Rate':>8}")
#     L.log(f"  {'-'*30} {'-'*12} {'-'*12} {'-'*10} {'-'*12} {'-'*8}")
#     for _, row in branches.iterrows():
#         b_sub  = inv[inv['branchName'] == row['branchName']]
#         b_inv  = b_sub['invoiceNumber'].nunique()
#         b_ph   = b_sub['clean_phone'].nunique()
#         b_rate = f"{b_ph/b_inv:.1%}" if b_inv else "—"
#         L.log(f"  {row['branchName']:<30} {row['City']:<12} {row['Format']:<12} {b_inv:>10,} {b_ph:>12,} {b_rate:>8}")

#     L.log()
#     L.log("=" * 80)
#     L.log("  END OF LOG")
#     L.log("=" * 80)
#     L.save()

#     # 7. Build Excel ────────────────────────────────────────────────────────────
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Phone No Record Tracker"

#     ROYAL_BLUE = "1A3A6B"
#     HDR_WEEK   = "2C5282"
#     HDR_DAY    = "1A3A6B"
#     WHITE      = "FFFFFF"
#     LIGHT_ROW  = "F7F9FC"
#     ALT_ROW    = "EEF2F9"
#     BORDER_C   = "C5CFE3"

#     thin = Side(style='thin', color=BORDER_C)
#     bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

#     total_cols      = 3 + len(metric_cols)
#     last_col_letter = get_column_letter(total_cols)

#     # Title row
#     ws.merge_cells(f'A1:{last_col_letter}1')
#     c = ws['A1']
#     c.value     = "Phone No Record Tracker  |  Dine In & Take Away"
#     c.font      = Font(name='Arial', bold=True, size=14, color=WHITE)
#     c.fill      = PatternFill('solid', start_color=ROYAL_BLUE)
#     c.alignment = Alignment(horizontal='center', vertical='center')
#     ws.row_dimensions[1].height = 30

#     ws.merge_cells(f'A2:{last_col_letter}2')
#     ws['A2'].fill = PatternFill('solid', start_color=ROYAL_BLUE)
#     ws.row_dimensions[2].height = 6

#     # Group header row (row 3)
#     for c_idx in [1, 2, 3]:
#         ws.cell(row=3, column=c_idx).fill = PatternFill('solid', start_color=ROYAL_BLUE)

#     ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=6)
#     wk_cell = ws.cell(row=3, column=4, value="← Last 3 Weeks →")
#     wk_cell.font      = Font(name='Arial', bold=True, size=9, color=WHITE)
#     wk_cell.fill      = PatternFill('solid', start_color=HDR_WEEK)
#     wk_cell.alignment = Alignment(horizontal='center')

#     ws.merge_cells(start_row=3, start_column=7, end_row=3, end_column=9)
#     dy_cell = ws.cell(row=3, column=7, value="← Last 3 Days →")
#     dy_cell.font      = Font(name='Arial', bold=True, size=9, color=WHITE)
#     dy_cell.fill      = PatternFill('solid', start_color=HDR_DAY)
#     dy_cell.alignment = Alignment(horizontal='center')

#     # Column label row (row 4)
#     headers  = ['City', 'Branch', 'Format'] + metric_cols
#     col_meta = {}
#     for h in ['City', 'Branch', 'Format']:
#         col_meta[h] = (ROYAL_BLUE, WHITE)
#     for c in week_cols:
#         col_meta[c] = (HDR_WEEK, WHITE)
#     for c in day_cols:
#         col_meta[c] = (HDR_DAY, WHITE)

#     for ci, h in enumerate(headers, start=1):
#         cell = ws.cell(row=4, column=ci, value=h)
#         bg, fg = col_meta.get(h, (ROYAL_BLUE, WHITE))
#         cell.fill      = PatternFill('solid', start_color=bg)
#         cell.font      = Font(name='Arial', bold=True, size=10, color=fg)
#         cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
#         cell.border    = bdr
#     ws.row_dimensions[4].height = 28

#     # Data rows
#     for ri, row in branches.iterrows():
#         excel_row = ri + 5
#         fill_bg   = LIGHT_ROW if ri % 2 == 0 else ALT_ROW
#         vals = [row['City'], row['branchName'], row['Format']] + [row[c] for c in metric_cols]
#         for ci, val in enumerate(vals, start=1):
#             cell        = ws.cell(row=excel_row, column=ci)
#             cell.border = bdr
#             if ci <= 3:
#                 cell.value     = val
#                 cell.font      = Font(name='Arial', size=9)
#                 cell.fill      = PatternFill('solid', start_color=fill_bg)
#                 cell.alignment = Alignment(horizontal='left', vertical='center')
#             else:
#                 if val is None:
#                     cell.value     = "—"
#                     cell.font      = Font(name='Arial', size=9, color='BBBBBB')
#                     cell.fill      = PatternFill('solid', start_color=fill_bg)
#                     cell.alignment = Alignment(horizontal='center')
#                 else:
#                     cell.value         = val
#                     cell.number_format = '0.0%'
#                     cell.font          = Font(name='Arial', size=9)
#                     cell.fill          = PatternFill('solid', start_color=fill_bg)
#                     cell.alignment     = Alignment(horizontal='center')

#     total_data_rows = len(branches)

#     # ── TOTAL row ─────────────────────────────────────────────────────────────
#     total_row_num = 5 + total_data_rows
#     TOTAL_BG      = "1A3A6B"
#     TOTAL_FG      = "FFFFFF"

#     # Merge first 3 cols for "TOTAL" label
#     ws.merge_cells(start_row=total_row_num, start_column=1, end_row=total_row_num, end_column=3)
#     lbl_cell            = ws.cell(row=total_row_num, column=1, value="TOTAL")
#     lbl_cell.font       = Font(name='Arial', bold=True, size=10, color=TOTAL_FG)
#     lbl_cell.fill       = PatternFill('solid', start_color=TOTAL_BG)
#     lbl_cell.alignment  = Alignment(horizontal='center', vertical='center')
#     lbl_cell.border     = bdr
#     # border on merged cells 2 & 3
#     for c_idx in [2, 3]:
#         ws.cell(row=total_row_num, column=c_idx).border = bdr

#     for ci, col in enumerate(metric_cols, start=4):
#         val  = totals.get(col)
#         cell = ws.cell(row=total_row_num, column=ci)
#         cell.fill   = PatternFill('solid', start_color=TOTAL_BG)
#         cell.border = bdr
#         if val is None:
#             cell.value     = "—"
#             cell.font      = Font(name='Arial', bold=True, size=10, color="AAAAAA")
#             cell.alignment = Alignment(horizontal='center', vertical='center')
#         else:
#             cell.value         = val
#             cell.number_format = '0.0%'
#             cell.font          = Font(name='Arial', bold=True, size=10, color=TOTAL_FG)
#             cell.alignment     = Alignment(horizontal='center', vertical='center')
#     ws.row_dimensions[total_row_num].height = 22

#     last_row = total_row_num

#     # Conditional formatting: ONLY on day cols
#     day_col_start = 3 + len(week_cols) + 1
#     for ci_offset in range(len(day_cols)):
#         ci         = day_col_start + ci_offset
#         col_letter = get_column_letter(ci)
#         ws.conditional_formatting.add(
#             f"{col_letter}5:{col_letter}{last_row}",
#             ColorScaleRule(
#                 start_type='min', start_color='C0392B',
#                 end_type='max',   end_color='FFFFFF'
#             )
#         )

#     col_widths = [14, 28, 14, 9, 9, 9, 9, 9, 9]
#     for ci, w in enumerate(col_widths, start=1):
#         ws.column_dimensions[get_column_letter(ci)].width = w

#     ws.freeze_panes = 'D5'

#     # 8. Raw Dump sheet ─────────────────────────────────────────────────────────
#     ws2 = wb.create_sheet("Raw Dump")

#     extra_cols_wanted = ['invoiceDate', 'customer_name', 'customer_email',
#                          'customer_id', 'orderNumber', 'netAmount', 'grossAmount']
#     available_extra = [c for c in extra_cols_wanted if c in df.columns]

#     df['_qty'] = pd.to_numeric(df.get('item_quantity', 0), errors='coerce').fillna(0)
#     sale_qty   = df.groupby('invoiceNumber')['_qty'].sum().reset_index()
#     sale_qty.columns = ['invoiceNumber', 'sale_qty']

#     raw = df.drop_duplicates(subset='invoiceNumber')[
#         ['invoiceNumber', 'branchName', 'channel', 'date', 'clean_phone'] + available_extra
#     ].copy()
#     raw = raw.merge(mapping_df[['branchName', 'City', 'Format']], on='branchName', how='left')
#     raw = raw.merge(sale_qty, on='invoiceNumber', how='left')
#     raw['City']   = raw['City'].fillna('Unknown')
#     raw['Format'] = raw['Format'].fillna('Unknown')
#     raw['date']   = pd.to_datetime(raw['date'], errors='coerce').dt.date

#     # ── EXCLUDE Cloud Kitchen and specific outlets from Raw Dump ────────────
#     raw = raw[raw['Format'] != 'Cloud Kitchen']
#     raw = raw[raw['branchName'] != 'RBI Officers Canteen']  # ── excluded outlet

#     for nc in ['netAmount', 'grossAmount']:
#         if nc in raw.columns:
#             raw[nc] = pd.to_numeric(raw[nc], errors='coerce')

#     raw_dump = (
#         raw[raw['clean_phone'].notna()]
#         .merge(order_counts, on='branchName', how='left')
#         .sort_values(['total_orders', 'date', 'branchName'], ascending=[False, True, True])
#         .drop_duplicates(subset='clean_phone', keep='first')  # ── one row per unique phone
#         .drop(columns=['total_orders'])
#         .reset_index(drop=True)
#     )

#     ordered_cols = ['invoiceNumber', 'date']
#     for c in ['invoiceDate', 'branchName', 'City', 'Format', 'channel',
#               'customer_name', 'clean_phone', 'customer_email', 'customer_id',
#               'orderNumber', 'sale_qty', 'netAmount', 'grossAmount']:
#         if c in raw_dump.columns:
#             ordered_cols.append(c)
#     raw_dump = raw_dump[ordered_cols]

#     header_labels = {
#         'invoiceNumber': 'Invoice No', 'date': 'Date', 'invoiceDate': 'Invoice Date',
#         'branchName': 'Branch', 'City': 'City', 'Format': 'Format', 'channel': 'Channel',
#         'customer_name': 'Customer Name', 'clean_phone': 'Phone Number',
#         'customer_email': 'Email', 'customer_id': 'Customer ID', 'orderNumber': 'Order No',
#         'sale_qty': 'Sale Qty', 'netAmount': 'Net Amount', 'grossAmount': 'Gross Amount',
#     }

#     numeric_cols_raw = {'netAmount', 'grossAmount', 'sale_qty'}
#     left_cols        = {'branchName', 'customer_name', 'customer_email'}
#     hdr_fill  = PatternFill('solid', start_color="1A3A6B")
#     hdr_font  = Font(name='Arial', bold=True, size=10, color="FFFFFF")
#     hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

#     for ci, col in enumerate(ordered_cols, start=1):
#         cell = ws2.cell(row=1, column=ci, value=header_labels.get(col, col))
#         cell.fill = hdr_fill; cell.font = hdr_font
#         cell.alignment = hdr_align; cell.border = bdr
#     ws2.row_dimensions[1].height = 28

#     for ri, row_data in raw_dump.iterrows():
#         excel_row = ri + 2
#         fill_bg   = "F9FAFB" if ri % 2 == 0 else "FFFFFF"
#         for ci, col in enumerate(ordered_cols, start=1):
#             val  = row_data[col]
#             cell = ws2.cell(row=excel_row, column=ci)
#             cell.fill   = PatternFill('solid', start_color=fill_bg)
#             cell.font   = Font(name='Arial', size=9)
#             cell.border = bdr
#             if col in numeric_cols_raw:
#                 try:
#                     cell.value = float(val) if val is not None and str(val) != 'nan' else None
#                 except (ValueError, TypeError):
#                     cell.value = None
#                 cell.number_format = '#,##0.00' if col in {'netAmount', 'grossAmount'} else '#,##0'
#                 cell.alignment = Alignment(horizontal='right', vertical='center')
#             elif col == 'clean_phone':
#                 cell.value         = str(val) if val is not None else None
#                 cell.number_format = '@'
#                 cell.alignment     = Alignment(horizontal='center', vertical='center')
#             else:
#                 cell.value     = val
#                 cell.alignment = Alignment(
#                     horizontal='left' if col in left_cols else 'center',
#                     vertical='center'
#                 )

#     col_widths2 = {
#         'invoiceNumber': 16, 'date': 12, 'invoiceDate': 14, 'branchName': 26,
#         'City': 12, 'Format': 14, 'channel': 12, 'customer_name': 20,
#         'clean_phone': 15, 'customer_email': 26, 'customer_id': 14,
#         'orderNumber': 14, 'sale_qty': 10, 'netAmount': 14, 'grossAmount': 14,
#     }
#     for ci, col in enumerate(ordered_cols, start=1):
#         ws2.column_dimensions[get_column_letter(ci)].width = col_widths2.get(col, 14)

#     ws2.freeze_panes = 'A2'

#     wb.save(output_xlsx)
#     phone_fill_pct = (inv['clean_phone'].nunique() / inv['invoiceNumber'].nunique() * 100) if len(inv) else 0
#     print(f"✅ Report saved      → {output_xlsx}")
#     print(f"📋 Log saved         → {log_path}")
#     print(f"   Branches analysed             : {len(branches)}")
#     print(f"   Total invoices (Dine In / TA) : {inv['invoiceNumber'].nunique():,}")
#     print(f"   Raw Dump rows (with phone)    : {len(raw_dump):,}")
#     print(f"   Overall phone capture         : {phone_fill_pct:.1f}%")


# # ── Entry point ───────────────────────────────────────────────────────────────
# if __name__ == "__main__":
#     import sys

#     ROLLING_CSV = r"C:/Users/jeryy/OneDrive - BLISS CHOCOLATES INDIA PRIVATE LIMITED/Desktop/Documents/Phone No/Data/sales_rolling.csv"
#     OUTPUT_DIR  = r"C:/Users/jeryy/OneDrive - BLISS CHOCOLATES INDIA PRIVATE LIMITED/Desktop/Documents/Phone No/output"

#     input_file = sys.argv[1] if len(sys.argv) > 1 else ROLLING_CSV
#     if len(sys.argv) > 2:
#         output_file = sys.argv[2]
#     else:
#         os.makedirs(OUTPUT_DIR, exist_ok=True)
#         from datetime import datetime, timedelta

#         date_str = (datetime.today() - timedelta(days=1)).strftime("%Y-%m-%d")
#         output_file = os.path.join(OUTPUT_DIR, f"Phone No Record - {date_str}.xlsx")
#     build_report(input_file, output_file)






# import os
# import pandas as pd
# import re
# from datetime import datetime, date
# import openpyxl
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from openpyxl.formatting.rule import ColorScaleRule
# from openpyxl.utils import get_column_letter

# # ── Branch → City / Format mapping ──────────────────────────────────────────
# mapping_dict = {
#     "1 MG": {"City": "Bangalore", "Format": "Café"},
#     "1 Mg Cafe": {"City": "Bangalore", "Format": "Café"},
#     "1mg cafe": {"City": "Bangalore", "Format": "Café"},
#     "Adarsh Palm Retreat": {"City": "Bangalore", "Format": "Café"},
#     "Arumbakkam Cloud Kitchen": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "arumbakkam ck": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Avadi CH": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Bandra": {"City": "Mumbai", "Format": "Café"},
#     "Banni Square": {"City": "Delhi NCR", "Format": "Kiosk"},
#     "Basaveshwara Nagar CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "BEL Road": {"City": "Bangalore", "Format": "Café"},
#     "Bellandur": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Bhartiya City Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "BLR Airport T2 International Bakery": {"City": "Bangalore", "Format": "Airport"},
#     "BLR Airport Terminal 1": {"City": "Bangalore", "Format": "Airport"},
#     "Central Arcade DLF": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "dlf": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Chennai ECR": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Cummulus": {"City": "Bangalore", "Format": "Café"},
#     "Dwarka Cloud Kitchen": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "dwarka ck": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Electronic City": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Express Avenue SMOOR": {"City": "Chennai", "Format": "Kiosk"},
#     "express avenue": {"City": "Chennai", "Format": "Kiosk"},
#     "Forum Shantiniketan": {"City": "Bangalore", "Format": "Kiosk"},
#     "Forum South Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "Garuda Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "Godrej One": {"City": "Mumbai", "Format": "Kiosk"},
#     "Grant Road CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Head Office": {"City": "Bangalore", "Format": "NA"},
#     "Hesaraghatta": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Hinjewadi": {"City": "Pune", "Format": "Cloud Kitchen"},
#     "HSR": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "IC Colony CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Indiranagar": {"City": "Bangalore", "Format": "Signature"},
#     "Infinity Mall Malad": {"City": "Mumbai", "Format": "Kiosk"},
#     "J.P Nagar": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "j.p nagar": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Jayanagar": {"City": "Bangalore", "Format": "Café"},
#     "Kadugodi CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Kalkaji CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Kamanahalli": {"City": "Bangalore", "Format": "Café"},
#     "kammanahalli": {"City": "Bangalore", "Format": "Café"},
#     "Kharghar CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Koramangala": {"City": "Bangalore", "Format": "Café"},
#     "Koreagaon Park": {"City": "Pune", "Format": "Signature"},
#     "Lavelle Road": {"City": "Bangalore", "Format": "Signature"},
#     "Laxmi Nagar": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Magarpatta": {"City": "Pune", "Format": "Cloud Kitchen"},
#     "Mahadevpura": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Malad": {"City": "Mumbai", "Format": "Kiosk"},
#     "MAP Smoor": {"City": "Bangalore", "Format": "Café"},
#     "Marathahalli CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Medavakkam TN": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Mira Road CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Mukherjee Nagar": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Netaji Shubhash Palace": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "nsp": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Nexus Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "Noida Sec 46": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Noida Sec 83": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Orion Mall": {"City": "Bangalore", "Format": "Kiosk"},
#     "Perungudi CK": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Phoenix Mall Pune": {"City": "Pune", "Format": "Kiosk"},
#     "Powai": {"City": "Mumbai", "Format": "Signature"},
#     "Rajouri Garden": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "rajouri": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Ramapuram TN": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "ramapuram": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "RT Nagar CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Sadashivanagar": {"City": "Bangalore", "Format": "Café"},
#     "Sakinaka CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Sanpada": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Sarjapur CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Seawoods Mall": {"City": "Mumbai", "Format": "Café"},
#     "Smoor - Defence Colony CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "defence colony ck": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Test.Rebel": {"City": "NA", "Format": "NA"},
#     "Udyog Vihar CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Undri": {"City": "Pune", "Format": "Cloud Kitchen"},
#     "Uttarahalli CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Vandalur": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Vashi Smoor": {"City": "Mumbai", "Format": "Kiosk"},
#     "Vasundra CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Vega City": {"City": "Bangalore", "Format": "Kiosk"},
#     "Versova": {"City": "Mumbai", "Format": "Café"},
#     "Viviana Lounge Mumbai": {"City": "Mumbai", "Format": "Kiosk"},
#     "viviana": {"City": "Mumbai", "Format": "Kiosk"},
#     "West Tambaram": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Whitefield": {"City": "Bangalore", "Format": "Signature"},
#     "Whitefield Lounge": {"City": "Bangalore", "Format": "Signature"},
#     "Worli": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Yelahanka Cloud Kitchen": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "yelahanka": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Kandivali CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Airoli CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Ambience Mall": {"City": "Delhi NCR", "Format": "Kiosk"},
#     "Bhandup CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Green Park CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Gurugram sector 31 Cloud Kitchen": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Malviya Nagar CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "RR Nagar CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Rohini CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "Henjewadi": {"City": "Pune", "Format": "Cloud Kitchen"},
#     "Hesarghatta": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "hessaraghatta": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Embassy Lake": {"City": "Bangalore", "Format": "Café"},
#     "Chembur East CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Hesaraghatta Outlet": {"City": "Bangalore", "Format": "Cloud Kitchen"},
#     "Mahakali CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
#     "Noida Sector 46": {"City": "NCR", "Format": "Cloud Kitchen"},
#     "Noida Sector 83": {"City": "NCR", "Format": "Cloud Kitchen"},
#     "Velachery": {"City": "Chennai", "Format": "Cloud Kitchen"},
#     "Smoor Deloitte": {"City": "Bangalore", "Format": "Kiosk"},
#     "Zirakpur": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
#     "RBI Officers Canteen": {"City": "Bangalore", "Format": "Kiosk"},
# }

# # ── Phone cleaning ────────────────────────────────────────────────────────────
# def clean_phone(raw):
#     if pd.isna(raw):
#         return None
#     phone = str(raw).strip()
#     phone = re.sub(r'[\s\-\(\)\+\.]', '', phone)
#     phone = re.sub(r'\D', '', phone)
#     if not phone:
#         return None
#     if len(phone) == 12 and phone.startswith('91'):
#         phone = phone[2:]
#     if len(phone) == 11 and phone.startswith('0'):
#         phone = phone[1:]
#     if len(phone) != 10:
#         return None
#     for digit in set(phone):
#         if phone.count(digit) > 5:
#             return None
#     if phone[0] in ('0', '1'):
#         return None
#     if phone[0] not in ('6', '7', '8', '9'):
#         return None
#     return phone


# # ── Logging helper ────────────────────────────────────────────────────────────
# class Logger:
#     def __init__(self, log_path: str):
#         self.log_path = log_path
#         self.lines    = []

#     def log(self, text: str = ""):
#         self.lines.append(text)

#     def section(self, title: str):
#         self.log()
#         self.log("=" * 80)
#         self.log(f"  {title}")
#         self.log("=" * 80)

#     def subsection(self, title: str):
#         self.log()
#         self.log(f"  ── {title} ──")
#         self.log()

#     def save(self):
#         with open(self.log_path, 'w', encoding='utf-8') as f:
#             f.write('\n'.join(self.lines))
#         print(f"📋 Log saved → {self.log_path}")


# # ── Main analysis ─────────────────────────────────────────────────────────────
# def build_report(input_csv: str, output_xlsx: str):

#     # Setup logger
#     log_path = output_xlsx.replace('.xlsx', '_log.txt')
#     L = Logger(log_path)
#     run_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

#     L.log("=" * 80)
#     L.log("  PHONE NO RECORD TRACKER  —  DETAILED CALCULATION LOG")
#     L.log(f"  Run timestamp : {run_time}")
#     L.log(f"  Input file    : {input_csv}")
#     L.log(f"  Output file   : {output_xlsx}")
#     L.log("=" * 80)

#     # ── STEP 1: Load & filter ─────────────────────────────────────────────────
#     L.section("STEP 1 — LOAD & CHANNEL FILTER")
#     df_raw = pd.read_csv(input_csv, dtype=str, low_memory=False)
#     L.log(f"  Total rows loaded from CSV          : {len(df_raw):,}")
#     L.log(f"  Unique invoices in raw data         : {df_raw['invoiceNumber'].nunique():,}")

#     channel_counts = df_raw['channel'].str.strip().str.lower().value_counts()
#     L.log()
#     L.log("  Channel distribution in raw data:")
#     for ch, cnt in channel_counts.items():
#         L.log(f"    {ch:<30} : {cnt:,}")

#     df = df_raw[df_raw['channel'].str.strip().str.lower().isin(['dine in', 'take away', 'dine-in', 'takeaway'])].copy()
#     L.log()
#     L.log(f"  Rows KEPT after channel filter      : {len(df):,}")
#     L.log(f"  Rows DROPPED (other channels)       : {len(df_raw) - len(df):,}")
#     L.log(f"  Unique invoices after filter        : {df['invoiceNumber'].nunique():,}")

#     # ── STEP 2: Phone cleaning ────────────────────────────────────────────────
#     L.section("STEP 2 — PHONE NUMBER CLEANING")
#     df['clean_phone'] = df['customer_phoneNumber'].apply(clean_phone)

#     total_rows        = len(df)
#     raw_has_phone     = df['customer_phoneNumber'].notna().sum()
#     cleaned_has_phone = df['clean_phone'].notna().sum()
#     dropped_in_clean  = raw_has_phone - cleaned_has_phone

#     L.log(f"  Total rows (post channel filter)    : {total_rows:,}")
#     L.log(f"  Rows with ANY phone value           : {raw_has_phone:,}")
#     L.log(f"  Rows with VALID cleaned phone       : {cleaned_has_phone:,}")
#     L.log(f"  Rows where phone was REJECTED       : {dropped_in_clean:,}")
#     L.log()
#     L.log("  Phone rejection reasons (applied in order):")
#     L.log("    1. Empty / NaN value")
#     L.log("    2. No digits after stripping special chars")
#     L.log("    3. 12-digit starting with 91 → strip country code")
#     L.log("    4. 11-digit starting with 0  → strip leading zero")
#     L.log("    5. Not exactly 10 digits after cleaning")
#     L.log("    6. Any single digit repeats more than 5 times (e.g. 9999999999)")
#     L.log("    7. First digit is 0 or 1 (invalid Indian mobile)")
#     L.log("    8. First digit not in 6,7,8,9 (invalid Indian mobile prefix)")

#     # ── STEP 3: Deduplicate at invoice level ──────────────────────────────────
#     L.section("STEP 3 — DEDUPLICATE AT INVOICE LEVEL")
#     before_dedup = len(df)
#     inv = df.drop_duplicates(subset='invoiceNumber')[
#         ['invoiceNumber', 'branchName', 'channel', 'date', 'clean_phone']
#     ].copy()
#     inv['date'] = pd.to_datetime(inv['date'], errors='coerce').dt.date
#     after_dedup = len(inv)

#     L.log(f"  Rows before dedup                   : {before_dedup:,}")
#     L.log(f"  Unique invoices after dedup         : {after_dedup:,}")
#     L.log(f"  Duplicate rows removed              : {before_dedup - after_dedup:,}")
#     L.log()
#     L.log("  NOTE: When multiple rows share the same invoiceNumber,")
#     L.log("  only the FIRST occurrence is kept (drop_duplicates keeps first row).")
#     L.log("  This means phone number is taken from the first line item of each invoice.")

#     # ── STEP 4: City / Format mapping ─────────────────────────────────────────
#     L.section("STEP 4 — BRANCH MAPPING (City & Format)")
#     mapping_df = pd.DataFrame.from_dict(mapping_dict, orient='index').reset_index()
#     mapping_df.columns = ['branchName', 'City', 'Format']
#     inv = inv.merge(mapping_df, on='branchName', how='left')
#     inv['City']   = inv['City'].fillna('Unknown')
#     inv['Format'] = inv['Format'].fillna('Unknown')

#     unknown_branches = inv[inv['City'] == 'Unknown']['branchName'].unique()
#     L.log(f"  Branches mapped successfully        : {inv[inv['City'] != 'Unknown']['branchName'].nunique():,}")
#     if len(unknown_branches) > 0:
#         L.log(f"  Branches NOT found in mapping       : {len(unknown_branches)}")
#         for b in unknown_branches:
#             L.log(f"    → '{b}'  (will be DROPPED)")
#     else:
#         L.log("  All branches found in mapping dict  : ✓")

#     inv_before_filter = len(inv)
#     inv = inv[~inv['City'].isin(['Unknown', 'NA'])]
#     inv = inv[inv['Format'] != 'Cloud Kitchen']
#     inv = inv[inv['branchName'] != 'RBI Officers Canteen']  # ── excluded outlet
#     inv_after_filter  = len(inv)

#     format_dist = inv['Format'].value_counts()
#     L.log()
#     L.log(f"  Invoices dropped (Unknown/NA city)  : {inv_before_filter - inv_after_filter:,}")
#     L.log(f"  Cloud Kitchen invoices excluded     : (filtered out entirely)")
#     L.log(f"  Invoices remaining for analysis     : {inv_after_filter:,}")
#     L.log()
#     L.log("  Format distribution (post-filter):")
#     for fmt, cnt in format_dist.items():
#         L.log(f"    {fmt:<20} : {cnt:,} invoices")

#     # ── STEP 5: Week / Day buckets ────────────────────────────────────────────
#     L.section("STEP 5 — WEEK & DAY BUCKETS")
#     inv['date_dt']  = pd.to_datetime(inv['date'], errors='coerce')
#     inv['iso_week'] = inv['date_dt'].dt.isocalendar().week.astype(int)
#     inv['iso_year'] = inv['date_dt'].dt.isocalendar().year.astype(int)
#     inv['yw']       = inv['iso_year'] * 100 + inv['iso_week']

#     last3_yw          = sorted(inv['yw'].dropna().unique())[-3:]
#     target_week_pairs = [(int(yw // 100), int(yw % 100)) for yw in last3_yw]
#     all_days          = sorted(inv['date'].dropna().unique())
#     target_days       = all_days[-3:] if len(all_days) >= 3 else all_days
#     latest_yw         = last3_yw[-1] if last3_yw else None

#     L.log(f"  Date range in data                  : {min(all_days)} → {max(all_days)}")
#     L.log(f"  Total unique days in data           : {len(all_days)}")
#     L.log()
#     L.log("  Last 3 ISO weeks selected:")
#     for (yr, wk) in target_week_pairs:
#         yw_key = yr * 100 + wk
#         label  = "← WTD (current week)" if yw_key == latest_yw else ""
#         wk_inv = inv[(inv['iso_week'] == wk) & (inv['iso_year'] == yr)]
#         L.log(f"    Year {yr}  Week {wk:02d}  (yw={yw_key})  →  {wk_inv['invoiceNumber'].nunique():,} invoices  {label}")
#     L.log()
#     L.log("  Last 3 days selected:")
#     for d in target_days:
#         day_inv = inv[inv['date'] == d]
#         L.log(f"    {d}  →  {day_inv['invoiceNumber'].nunique():,} invoices across {day_inv['branchName'].nunique()} branches")

#     # ── STEP 6: Branch-level metrics ──────────────────────────────────────────
#     L.section("STEP 6 — BRANCH-LEVEL CAPTURE RATE CALCULATIONS")
#     L.log()
#     L.log("  FORMULA:")
#     L.log("    Capture Rate = Unique valid phone numbers ÷ Unique invoices")
#     L.log("    (scoped to: one branch × one time window)")
#     L.log()

#     order_counts = inv.groupby('branchName')['invoiceNumber'].nunique().reset_index()
#     order_counts.columns = ['branchName', 'total_orders']

#     branches = (
#         inv[['branchName', 'City', 'Format']]
#         .drop_duplicates('branchName')
#         .merge(order_counts, on='branchName', how='left')
#         .sort_values('total_orders', ascending=False)
#         .reset_index(drop=True)
#     )

#     # ── FIXED: Unique valid phone numbers ÷ Unique invoices ──────────────────
#     def capture(sub):
#         total     = sub['invoiceNumber'].nunique()
#         has_phone = sub['clean_phone'].nunique()  # unique valid phone numbers
#         return has_phone / total if total > 0 else None

#     day_fmt   = lambda d: d.strftime('%b %#d') if os.name == 'nt' else d.strftime('%b %-d')

#     # Week cols
#     week_cols = []
#     L.subsection("Weekly Capture Rates by Branch")
#     for (yr, wk) in target_week_pairs:
#         yw_key = yr * 100 + wk
#         col    = f'W{wk} WTD' if yw_key == latest_yw else f'W{wk}'
#         week_cols.append(col)
#         L.log(f"  [{col}]  Year={yr}  Week={wk}")
#         L.log(f"  {'Branch':<30} {'Unique Invoices':>16} {'Unique Phones':>14} {'Capture Rate':>13}")
#         L.log(f"  {'-'*30} {'-'*16} {'-'*14} {'-'*13}")
#         rates = []
#         for _, row in branches.iterrows():
#             sub   = inv[(inv['branchName'] == row['branchName']) &
#                         (inv['iso_week'] == wk) & (inv['iso_year'] == yr)]
#             rate  = capture(sub)
#             rates.append(rate)
#             if len(sub) > 0:
#                 total_inv = sub['invoiceNumber'].nunique()
#                 uniq_ph   = sub['clean_phone'].nunique()
#                 rate_str  = f"{rate:.1%}" if rate is not None else "—"
#                 L.log(f"  {row['branchName']:<30} {total_inv:>16,} {uniq_ph:>14,} {rate_str:>13}")
#         branches[col] = rates
#         L.log()

#     # Day cols
#     day_cols = []
#     L.subsection("Daily Capture Rates by Branch")
#     for d in target_days:
#         col = day_fmt(d)
#         day_cols.append(col)
#         L.log(f"  [{col}]  Date={d}")
#         L.log(f"  {'Branch':<30} {'Unique Invoices':>16} {'Unique Phones':>14} {'Capture Rate':>13}")
#         L.log(f"  {'-'*30} {'-'*16} {'-'*14} {'-'*13}")
#         rates = []
#         for _, row in branches.iterrows():
#             sub   = inv[(inv['branchName'] == row['branchName']) & (inv['date'] == d)]
#             rate  = capture(sub)
#             rates.append(rate)
#             if len(sub) > 0:
#                 total_inv = sub['invoiceNumber'].nunique()
#                 uniq_ph   = sub['clean_phone'].nunique()
#                 rate_str  = f"{rate:.1%}" if rate is not None else "—"
#                 L.log(f"  {row['branchName']:<30} {total_inv:>16,} {uniq_ph:>14,} {rate_str:>13}")
#         branches[col] = rates
#         L.log()

#     metric_cols = week_cols + day_cols

#     # ── Compute totals row: unique phones ÷ unique invoices per time window ──
#     totals = {'City': 'ALL', 'branchName': 'TOTAL', 'Format': ''}
#     for (yr, wk) in target_week_pairs:
#         yw_key = yr * 100 + wk
#         col    = f'W{wk} WTD' if yw_key == latest_yw else f'W{wk}'
#         sub    = inv[(inv['iso_week'] == wk) & (inv['iso_year'] == yr)]
#         t_inv  = sub['invoiceNumber'].nunique()
#         t_ph   = sub['clean_phone'].nunique()
#         totals[col] = (t_ph / t_inv) if t_inv > 0 else None
#     for d in target_days:
#         col   = day_fmt(d)
#         sub   = inv[inv['date'] == d]
#         t_inv = sub['invoiceNumber'].nunique()
#         t_ph  = sub['clean_phone'].nunique()
#         totals[col] = (t_ph / t_inv) if t_inv > 0 else None

#     # ── STEP 7: Overall summary ───────────────────────────────────────────────
#     L.section("STEP 7 — OVERALL SUMMARY")
#     total_invoices = inv['invoiceNumber'].nunique()
#     total_phones   = inv['clean_phone'].nunique()
#     overall_rate   = total_phones / total_invoices * 100 if total_invoices else 0

#     L.log(f"  Total unique invoices (Dine In/TA)  : {total_invoices:,}")
#     L.log(f"  Total unique valid phone numbers    : {total_phones:,}")
#     L.log(f"  Overall capture rate                : {overall_rate:.1f}%")
#     L.log(f"  Formula: {total_phones:,} ÷ {total_invoices:,} = {overall_rate:.1f}%")

#     L.subsection("Top 15 Most Frequent Phone Numbers (Dummy Number Check)")
#     top_phones = inv['clean_phone'].value_counts().head(15)
#     L.log(f"  {'Phone Number':<15} {'Count':>8}   {'Flag'}")
#     L.log(f"  {'-'*15} {'-'*8}   {'-'*30}")
#     for ph, cnt in top_phones.items():
#         flag = "⚠️  LIKELY DUMMY — appears very frequently" if cnt > 10 else ""
#         L.log(f"  {str(ph):<15} {cnt:>8,}   {flag}")

#     L.subsection("Capture Rate by City")
#     for city in sorted(inv['City'].unique()):
#         city_sub   = inv[inv['City'] == city]
#         city_inv   = city_sub['invoiceNumber'].nunique()
#         city_ph    = city_sub['clean_phone'].nunique()
#         city_rate  = city_ph / city_inv * 100 if city_inv else 0
#         L.log(f"  {city:<15}  Invoices={city_inv:,}  Unique Phones={city_ph:,}  Rate={city_rate:.1f}%")

#     L.subsection("Capture Rate by Format")
#     for fmt in sorted(inv['Format'].unique()):
#         fmt_sub  = inv[inv['Format'] == fmt]
#         fmt_inv  = fmt_sub['invoiceNumber'].nunique()
#         fmt_ph   = fmt_sub['clean_phone'].nunique()
#         fmt_rate = fmt_ph / fmt_inv * 100 if fmt_inv else 0
#         L.log(f"  {fmt:<15}  Invoices={fmt_inv:,}  Unique Phones={fmt_ph:,}  Rate={fmt_rate:.1f}%")

#     L.subsection("All Branches — Overall Capture Rate (sorted by invoices)")
#     L.log(f"  {'Branch':<30} {'City':<12} {'Format':<12} {'Invoices':>10} {'Uniq Phones':>12} {'Rate':>8}")
#     L.log(f"  {'-'*30} {'-'*12} {'-'*12} {'-'*10} {'-'*12} {'-'*8}")
#     for _, row in branches.iterrows():
#         b_sub  = inv[inv['branchName'] == row['branchName']]
#         b_inv  = b_sub['invoiceNumber'].nunique()
#         b_ph   = b_sub['clean_phone'].nunique()
#         b_rate = f"{b_ph/b_inv:.1%}" if b_inv else "—"
#         L.log(f"  {row['branchName']:<30} {row['City']:<12} {row['Format']:<12} {b_inv:>10,} {b_ph:>12,} {b_rate:>8}")

#     L.log()
#     L.log("=" * 80)
#     L.log("  END OF LOG")
#     L.log("=" * 80)
#     L.save()

#     # 7. Build Excel ────────────────────────────────────────────────────────────
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Phone No Record Tracker"

#     ROYAL_BLUE = "1A3A6B"
#     HDR_WEEK   = "2C5282"
#     HDR_DAY    = "1A3A6B"
#     WHITE      = "FFFFFF"
#     LIGHT_ROW  = "F7F9FC"
#     ALT_ROW    = "EEF2F9"
#     BORDER_C   = "C5CFE3"

#     thin = Side(style='thin', color=BORDER_C)
#     bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

#     total_cols      = 3 + len(metric_cols)
#     last_col_letter = get_column_letter(total_cols)

#     # Title row
#     ws.merge_cells(f'A1:{last_col_letter}1')
#     c = ws['A1']
#     c.value     = "Phone No Record Tracker  |  Dine In & Take Away"
#     c.font      = Font(name='Arial', bold=True, size=14, color=WHITE)
#     c.fill      = PatternFill('solid', start_color=ROYAL_BLUE)
#     c.alignment = Alignment(horizontal='center', vertical='center')
#     ws.row_dimensions[1].height = 30

#     ws.merge_cells(f'A2:{last_col_letter}2')
#     ws['A2'].fill = PatternFill('solid', start_color=ROYAL_BLUE)
#     ws.row_dimensions[2].height = 6

#     # Group header row (row 3)
#     for c_idx in [1, 2, 3]:
#         ws.cell(row=3, column=c_idx).fill = PatternFill('solid', start_color=ROYAL_BLUE)

#     ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=6)
#     wk_cell = ws.cell(row=3, column=4, value="← Last 3 Weeks →")
#     wk_cell.font      = Font(name='Arial', bold=True, size=9, color=WHITE)
#     wk_cell.fill      = PatternFill('solid', start_color=HDR_WEEK)
#     wk_cell.alignment = Alignment(horizontal='center')

#     ws.merge_cells(start_row=3, start_column=7, end_row=3, end_column=9)
#     dy_cell = ws.cell(row=3, column=7, value="← Last 3 Days →")
#     dy_cell.font      = Font(name='Arial', bold=True, size=9, color=WHITE)
#     dy_cell.fill      = PatternFill('solid', start_color=HDR_DAY)
#     dy_cell.alignment = Alignment(horizontal='center')

#     # Column label row (row 4)
#     headers  = ['City', 'Branch', 'Format'] + metric_cols
#     col_meta = {}
#     for h in ['City', 'Branch', 'Format']:
#         col_meta[h] = (ROYAL_BLUE, WHITE)
#     for c in week_cols:
#         col_meta[c] = (HDR_WEEK, WHITE)
#     for c in day_cols:
#         col_meta[c] = (HDR_DAY, WHITE)

#     for ci, h in enumerate(headers, start=1):
#         cell = ws.cell(row=4, column=ci, value=h)
#         bg, fg = col_meta.get(h, (ROYAL_BLUE, WHITE))
#         cell.fill      = PatternFill('solid', start_color=bg)
#         cell.font      = Font(name='Arial', bold=True, size=10, color=fg)
#         cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
#         cell.border    = bdr
#     ws.row_dimensions[4].height = 28

#     # Data rows
#     for ri, row in branches.iterrows():
#         excel_row = ri + 5
#         fill_bg   = LIGHT_ROW if ri % 2 == 0 else ALT_ROW
#         vals = [row['City'], row['branchName'], row['Format']] + [row[c] for c in metric_cols]
#         for ci, val in enumerate(vals, start=1):
#             cell        = ws.cell(row=excel_row, column=ci)
#             cell.border = bdr
#             if ci <= 3:
#                 cell.value     = val
#                 cell.font      = Font(name='Arial', size=9)
#                 cell.fill      = PatternFill('solid', start_color=fill_bg)
#                 cell.alignment = Alignment(horizontal='left', vertical='center')
#             else:
#                 if val is None:
#                     cell.value     = "—"
#                     cell.font      = Font(name='Arial', size=9, color='BBBBBB')
#                     cell.fill      = PatternFill('solid', start_color=fill_bg)
#                     cell.alignment = Alignment(horizontal='center')
#                 else:
#                     cell.value         = val
#                     cell.number_format = '0.0%'
#                     cell.font          = Font(name='Arial', size=9)
#                     cell.fill          = PatternFill('solid', start_color=fill_bg)
#                     cell.alignment     = Alignment(horizontal='center')

#     total_data_rows = len(branches)

#     # ── TOTAL row ─────────────────────────────────────────────────────────────
#     total_row_num = 5 + total_data_rows
#     TOTAL_BG      = "1A3A6B"
#     TOTAL_FG      = "FFFFFF"

#     # Merge first 3 cols for "TOTAL" label
#     ws.merge_cells(start_row=total_row_num, start_column=1, end_row=total_row_num, end_column=3)
#     lbl_cell            = ws.cell(row=total_row_num, column=1, value="TOTAL")
#     lbl_cell.font       = Font(name='Arial', bold=True, size=10, color=TOTAL_FG)
#     lbl_cell.fill       = PatternFill('solid', start_color=TOTAL_BG)
#     lbl_cell.alignment  = Alignment(horizontal='center', vertical='center')
#     lbl_cell.border     = bdr
#     # border on merged cells 2 & 3
#     for c_idx in [2, 3]:
#         ws.cell(row=total_row_num, column=c_idx).border = bdr

#     # Collect all non-None total values to compute min/max for color scale
#     total_vals = {col: v for col, v in totals.items() if isinstance(v, float)}
#     t_min = min(total_vals.values()) if total_vals else 0.0
#     t_max = max(total_vals.values()) if total_vals else 1.0
#     t_range = t_max - t_min if t_max != t_min else 1.0

#     def total_pct_to_bg(val):
#         # Same scale as email: #F8696B (light red) at min → #FFFFFF (white) at max
#         t = max(0.0, min(1.0, (val - t_min) / t_range))
#         r = int(248 + (255 - 248) * t)   # 248 → 255
#         g = int(105 + (255 - 105) * t)   # 105 → 255
#         b = int(107 + (255 - 107) * t)   # 107 → 255
#         return f"{r:02X}{g:02X}{b:02X}"  # no # prefix for openpyxl

#     for ci, col in enumerate(metric_cols, start=4):
#         val  = totals.get(col)
#         cell = ws.cell(row=total_row_num, column=ci)
#         cell.border = bdr
#         if val is None:
#             cell.value     = "—"
#             cell.fill      = PatternFill('solid', start_color=TOTAL_BG)
#             cell.font      = Font(name='Arial', bold=True, size=10, color="AAAAAA")
#             cell.alignment = Alignment(horizontal='center', vertical='center')
#         else:
#             cell_bg            = total_pct_to_bg(val)
#             cell.value         = val
#             cell.number_format = '0.0%'
#             cell.fill          = PatternFill('solid', start_color=cell_bg)
#             # Dark text on light bg, white text on dark bg
#             text_color         = "1A1A1A" if val > (t_min + t_range * 0.4) else "FFFFFF"
#             cell.font          = Font(name='Arial', bold=True, size=10, color=text_color)
#             cell.alignment     = Alignment(horizontal='center', vertical='center')
#     ws.row_dimensions[total_row_num].height = 22

#     last_row     = total_row_num
#     cf_last_row  = total_row_num - 1  # ── exclude TOTAL row from conditional formatting

#     # Conditional formatting: ONLY on day cols, ONLY on data rows (not total)
#     day_col_start = 3 + len(week_cols) + 1
#     for ci_offset in range(len(day_cols)):
#         ci         = day_col_start + ci_offset
#         col_letter = get_column_letter(ci)
#         ws.conditional_formatting.add(
#             f"{col_letter}5:{col_letter}{cf_last_row}",
#             ColorScaleRule(
#                 start_type='min', start_color='F8696B',
#                 end_type='max',   end_color='FFFFFF'
#             )
#         )

#     col_widths = [14, 28, 14, 9, 9, 9, 9, 9, 9]
#     for ci, w in enumerate(col_widths, start=1):
#         ws.column_dimensions[get_column_letter(ci)].width = w

#     ws.freeze_panes = 'D5'

#     # 8. Raw Dump sheet ─────────────────────────────────────────────────────────
#     ws2 = wb.create_sheet("Raw Dump")

#     extra_cols_wanted = ['invoiceDate', 'customer_name', 'customer_email',
#                          'customer_id', 'orderNumber', 'netAmount', 'grossAmount']
#     available_extra = [c for c in extra_cols_wanted if c in df.columns]

#     df['_qty'] = pd.to_numeric(df.get('item_quantity', 0), errors='coerce').fillna(0)
#     sale_qty   = df.groupby('invoiceNumber')['_qty'].sum().reset_index()
#     sale_qty.columns = ['invoiceNumber', 'sale_qty']

#     raw = df.drop_duplicates(subset='invoiceNumber')[
#         ['invoiceNumber', 'branchName', 'channel', 'date', 'clean_phone'] + available_extra
#     ].copy()
#     raw = raw.merge(mapping_df[['branchName', 'City', 'Format']], on='branchName', how='left')
#     raw = raw.merge(sale_qty, on='invoiceNumber', how='left')
#     raw['City']   = raw['City'].fillna('Unknown')
#     raw['Format'] = raw['Format'].fillna('Unknown')
#     raw['date']   = pd.to_datetime(raw['date'], errors='coerce').dt.date

#     # ── EXCLUDE Cloud Kitchen and specific outlets from Raw Dump ────────────
#     raw = raw[raw['Format'] != 'Cloud Kitchen']
#     raw = raw[raw['branchName'] != 'RBI Officers Canteen']  # ── excluded outlet

#     for nc in ['netAmount', 'grossAmount']:
#         if nc in raw.columns:
#             raw[nc] = pd.to_numeric(raw[nc], errors='coerce')

#     raw_dump = (
#         raw[raw['clean_phone'].notna()]
#         .merge(order_counts, on='branchName', how='left')
#         .sort_values(['total_orders', 'date', 'branchName'], ascending=[False, True, True])
#         .drop_duplicates(subset='clean_phone', keep='first')  # ── one row per unique phone
#         .drop(columns=['total_orders'])
#         .reset_index(drop=True)
#     )

#     ordered_cols = ['invoiceNumber', 'date']
#     for c in ['invoiceDate', 'branchName', 'City', 'Format', 'channel',
#               'customer_name', 'clean_phone', 'customer_email', 'customer_id',
#               'orderNumber', 'sale_qty', 'netAmount', 'grossAmount']:
#         if c in raw_dump.columns:
#             ordered_cols.append(c)
#     raw_dump = raw_dump[ordered_cols]

#     header_labels = {
#         'invoiceNumber': 'Invoice No', 'date': 'Date', 'invoiceDate': 'Invoice Date',
#         'branchName': 'Branch', 'City': 'City', 'Format': 'Format', 'channel': 'Channel',
#         'customer_name': 'Customer Name', 'clean_phone': 'Phone Number',
#         'customer_email': 'Email', 'customer_id': 'Customer ID', 'orderNumber': 'Order No',
#         'sale_qty': 'Sale Qty', 'netAmount': 'Net Amount', 'grossAmount': 'Gross Amount',
#     }

#     numeric_cols_raw = {'netAmount', 'grossAmount', 'sale_qty'}
#     left_cols        = {'branchName', 'customer_name', 'customer_email'}
#     hdr_fill  = PatternFill('solid', start_color="1A3A6B")
#     hdr_font  = Font(name='Arial', bold=True, size=10, color="FFFFFF")
#     hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

#     for ci, col in enumerate(ordered_cols, start=1):
#         cell = ws2.cell(row=1, column=ci, value=header_labels.get(col, col))
#         cell.fill = hdr_fill; cell.font = hdr_font
#         cell.alignment = hdr_align; cell.border = bdr
#     ws2.row_dimensions[1].height = 28

#     for ri, row_data in raw_dump.iterrows():
#         excel_row = ri + 2
#         fill_bg   = "F9FAFB" if ri % 2 == 0 else "FFFFFF"
#         for ci, col in enumerate(ordered_cols, start=1):
#             val  = row_data[col]
#             cell = ws2.cell(row=excel_row, column=ci)
#             cell.fill   = PatternFill('solid', start_color=fill_bg)
#             cell.font   = Font(name='Arial', size=9)
#             cell.border = bdr
#             if col in numeric_cols_raw:
#                 try:
#                     cell.value = float(val) if val is not None and str(val) != 'nan' else None
#                 except (ValueError, TypeError):
#                     cell.value = None
#                 cell.number_format = '#,##0.00' if col in {'netAmount', 'grossAmount'} else '#,##0'
#                 cell.alignment = Alignment(horizontal='right', vertical='center')
#             elif col == 'clean_phone':
#                 cell.value         = str(val) if val is not None else None
#                 cell.number_format = '@'
#                 cell.alignment     = Alignment(horizontal='center', vertical='center')
#             else:
#                 cell.value     = val
#                 cell.alignment = Alignment(
#                     horizontal='left' if col in left_cols else 'center',
#                     vertical='center'
#                 )

#     col_widths2 = {
#         'invoiceNumber': 16, 'date': 12, 'invoiceDate': 14, 'branchName': 26,
#         'City': 12, 'Format': 14, 'channel': 12, 'customer_name': 20,
#         'clean_phone': 15, 'customer_email': 26, 'customer_id': 14,
#         'orderNumber': 14, 'sale_qty': 10, 'netAmount': 14, 'grossAmount': 14,
#     }
#     for ci, col in enumerate(ordered_cols, start=1):
#         ws2.column_dimensions[get_column_letter(ci)].width = col_widths2.get(col, 14)

#     ws2.freeze_panes = 'A2'

#     wb.save(output_xlsx)
#     phone_fill_pct = (inv['clean_phone'].nunique() / inv['invoiceNumber'].nunique() * 100) if len(inv) else 0
#     print(f"✅ Report saved      → {output_xlsx}")
#     print(f"📋 Log saved         → {log_path}")
#     print(f"   Branches analysed             : {len(branches)}")
#     print(f"   Total invoices (Dine In / TA) : {inv['invoiceNumber'].nunique():,}")
#     print(f"   Raw Dump rows (with phone)    : {len(raw_dump):,}")
#     print(f"   Overall phone capture         : {phone_fill_pct:.1f}%")


# # ── Entry point ───────────────────────────────────────────────────────────────
# if __name__ == "__main__":
#     import sys

#     ROLLING_CSV = r"C:/Users/jeryy/OneDrive - BLISS CHOCOLATES INDIA PRIVATE LIMITED/Desktop/Documents/Phone No/Data/sales_rolling.csv"
#     OUTPUT_DIR  = r"C:/Users/jeryy/OneDrive - BLISS CHOCOLATES INDIA PRIVATE LIMITED/Desktop/Documents/Phone No/output"

#     input_file = sys.argv[1] if len(sys.argv) > 1 else ROLLING_CSV
#     if len(sys.argv) > 2:
#         output_file = sys.argv[2]
#     else:
#         os.makedirs(OUTPUT_DIR, exist_ok=True)
#         from datetime import timedelta
#         date_str    = (datetime.today() - timedelta(days=1)).strftime("%Y-%m-%d")
#         output_file = os.path.join(OUTPUT_DIR, f"Phone No Record - {date_str}.xlsx")

#     build_report(input_file, output_file)

import os
import pandas as pd
import re
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

# ── Branch → City / Format mapping ──────────────────────────────────────────
mapping_dict = {
    "1 MG": {"City": "Bangalore", "Format": "Café"},
    "1 Mg Cafe": {"City": "Bangalore", "Format": "Café"},
    "1mg cafe": {"City": "Bangalore", "Format": "Café"},
    "Adarsh Palm Retreat": {"City": "Bangalore", "Format": "Café"},
    "Arumbakkam Cloud Kitchen": {"City": "Chennai", "Format": "Cloud Kitchen"},
    "arumbakkam ck": {"City": "Chennai", "Format": "Cloud Kitchen"},
    "Avadi CH": {"City": "Chennai", "Format": "Cloud Kitchen"},
    "Bandra": {"City": "Mumbai", "Format": "Café"},
    "Banni Square": {"City": "Delhi NCR", "Format": "Kiosk"},
    "Basaveshwara Nagar CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
    "BEL Road": {"City": "Bangalore", "Format": "Café"},
    "Bellandur": {"City": "Bangalore", "Format": "Cloud Kitchen"},
    "Bhartiya City Mall": {"City": "Bangalore", "Format": "Kiosk"},
    "BLR Airport T2 International Bakery": {"City": "Bangalore", "Format": "Airport"},
    "BLR Airport Terminal 1": {"City": "Bangalore", "Format": "Airport"},
    "Central Arcade DLF": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
    "dlf": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
    "Chennai ECR": {"City": "Chennai", "Format": "Cloud Kitchen"},
    "Cummulus": {"City": "Bangalore", "Format": "Café"},
    "Dwarka Cloud Kitchen": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
    "dwarka ck": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
    "Electronic City": {"City": "Bangalore", "Format": "Cloud Kitchen"},
    "Express Avenue SMOOR": {"City": "Chennai", "Format": "Kiosk"},
    "express avenue": {"City": "Chennai", "Format": "Kiosk"},
    "Forum Shantiniketan": {"City": "Bangalore", "Format": "Kiosk"},
    "Forum South Mall": {"City": "Bangalore", "Format": "Kiosk"},
    "Garuda Mall": {"City": "Bangalore", "Format": "Kiosk"},
    "Godrej One": {"City": "Mumbai", "Format": "Kiosk"},
    "Grant Road CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
    "Head Office": {"City": "Bangalore", "Format": "NA"},
    "Hesaraghatta": {"City": "Bangalore", "Format": "Cloud Kitchen"},
    "Hinjewadi": {"City": "Pune", "Format": "Cloud Kitchen"},
    "HSR": {"City": "Bangalore", "Format": "Cloud Kitchen"},
    "IC Colony CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
    "Indiranagar": {"City": "Bangalore", "Format": "Signature"},
    "Infinity Mall Malad": {"City": "Mumbai", "Format": "Kiosk"},
    "J.P Nagar": {"City": "Bangalore", "Format": "Cloud Kitchen"},
    "j.p nagar": {"City": "Bangalore", "Format": "Cloud Kitchen"},
    "Jayanagar": {"City": "Bangalore", "Format": "Café"},
    "Kadugodi CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
    "Kalkaji CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
    "Kamanahalli": {"City": "Bangalore", "Format": "Café"},
    "kammanahalli": {"City": "Bangalore", "Format": "Café"},
    "Kharghar CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
    "Koramangala": {"City": "Bangalore", "Format": "Café"},
    "Koreagaon Park": {"City": "Pune", "Format": "Signature"},
    "Lavelle Road": {"City": "Bangalore", "Format": "Signature"},
    "Laxmi Nagar": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
    "Magarpatta": {"City": "Pune", "Format": "Cloud Kitchen"},
    "Mahadevpura": {"City": "Bangalore", "Format": "Cloud Kitchen"},
    "Malad": {"City": "Mumbai", "Format": "Kiosk"},
    "MAP Smoor": {"City": "Bangalore", "Format": "Café"},
    "Marathahalli CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
    "Medavakkam TN": {"City": "Chennai", "Format": "Cloud Kitchen"},
    "Mira Road CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
    "Mukherjee Nagar": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
    "Netaji Shubhash Palace": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
    "nsp": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
    "Nexus Mall": {"City": "Bangalore", "Format": "Kiosk"},
    "Noida Sec 46": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
    "Noida Sec 83": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
    "Orion Mall": {"City": "Bangalore", "Format": "Kiosk"},
    "Perungudi CK": {"City": "Chennai", "Format": "Cloud Kitchen"},
    "Phoenix Mall Pune": {"City": "Pune", "Format": "Kiosk"},
    "Powai": {"City": "Mumbai", "Format": "Signature"},
    "Rajouri Garden": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
    "rajouri": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
    "Ramapuram TN": {"City": "Chennai", "Format": "Cloud Kitchen"},
    "ramapuram": {"City": "Chennai", "Format": "Cloud Kitchen"},
    "RT Nagar CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
    "Sadashivanagar": {"City": "Bangalore", "Format": "Café"},
    "Sakinaka CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
    "Sanpada": {"City": "Mumbai", "Format": "Cloud Kitchen"},
    "Sarjapur CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
    "Seawoods Mall": {"City": "Mumbai", "Format": "Café"},
    "Smoor - Defence Colony CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
    "defence colony ck": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
    "Test.Rebel": {"City": "NA", "Format": "NA"},
    "Udyog Vihar CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
    "Undri": {"City": "Pune", "Format": "Cloud Kitchen"},
    "Uttarahalli CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
    "Vandalur": {"City": "Chennai", "Format": "Cloud Kitchen"},
    "Vashi Smoor": {"City": "Mumbai", "Format": "Kiosk"},
    "Vasundra CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
    "Vega City": {"City": "Bangalore", "Format": "Kiosk"},
    "Versova": {"City": "Mumbai", "Format": "Café"},
    "Viviana Lounge Mumbai": {"City": "Mumbai", "Format": "Kiosk"},
    "viviana": {"City": "Mumbai", "Format": "Kiosk"},
    "West Tambaram": {"City": "Chennai", "Format": "Cloud Kitchen"},
    "Whitefield": {"City": "Bangalore", "Format": "Signature"},
    "Whitefield Lounge": {"City": "Bangalore", "Format": "Signature"},
    "Worli": {"City": "Mumbai", "Format": "Cloud Kitchen"},
    "Yelahanka Cloud Kitchen": {"City": "Bangalore", "Format": "Cloud Kitchen"},
    "yelahanka": {"City": "Bangalore", "Format": "Cloud Kitchen"},
    "Kandivali CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
    "Airoli CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
    "Ambience Mall": {"City": "Delhi NCR", "Format": "Kiosk"},
    "Bhandup CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
    "Green Park CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
    "Gurugram sector 31 Cloud Kitchen": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
    "Malviya Nagar CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
    "RR Nagar CK": {"City": "Bangalore", "Format": "Cloud Kitchen"},
    "Rohini CK": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
    "Henjewadi": {"City": "Pune", "Format": "Cloud Kitchen"},
    "Hesarghatta": {"City": "Bangalore", "Format": "Cloud Kitchen"},
    "hessaraghatta": {"City": "Bangalore", "Format": "Cloud Kitchen"},
    "Embassy Lake": {"City": "Bangalore", "Format": "Café"},
    "Chembur East CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
    "Hesaraghatta Outlet": {"City": "Bangalore", "Format": "Cloud Kitchen"},
    "Mahakali CK": {"City": "Mumbai", "Format": "Cloud Kitchen"},
    "Noida Sector 46": {"City": "NCR", "Format": "Cloud Kitchen"},
    "Noida Sector 83": {"City": "NCR", "Format": "Cloud Kitchen"},
    "Velachery": {"City": "Chennai", "Format": "Cloud Kitchen"},
    "Smoor Deloitte": {"City": "Bangalore", "Format": "Kiosk"},
    "Zirakpur": {"City": "Delhi NCR", "Format": "Cloud Kitchen"},
    "RBI Officers Canteen": {"City": "Bangalore", "Format": "Kiosk"},
}

# ── Phone cleaning ────────────────────────────────────────────────────────────
def clean_phone(raw):
    if pd.isna(raw):
        return None
    phone = str(raw).strip()
    phone = re.sub(r'[\s\-\(\)\+\.]', '', phone)
    phone = re.sub(r'\D', '', phone)
    if not phone:
        return None
    if len(phone) == 12 and phone.startswith('91'):
        phone = phone[2:]
    if len(phone) == 11 and phone.startswith('0'):
        phone = phone[1:]
    if len(phone) != 10:
        return None
    for digit in set(phone):
        if phone.count(digit) > 5:
            return None
    if phone[0] in ('0', '1'):
        return None
    if phone[0] not in ('6', '7', '8', '9'):
        return None
    return phone


# ── Logging helper ────────────────────────────────────────────────────────────
class Logger:
    def __init__(self, log_path: str):
        self.log_path = log_path
        self.lines    = []

    def log(self, text: str = ""):
        self.lines.append(text)

    def section(self, title: str):
        self.log()
        self.log("=" * 80)
        self.log(f"  {title}")
        self.log("=" * 80)

    def subsection(self, title: str):
        self.log()
        self.log(f"  ── {title} ──")
        self.log()

    def save(self):
        with open(self.log_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(self.lines))
        print(f"📋 Log saved → {self.log_path}")


# ── Main analysis ─────────────────────────────────────────────────────────────
def build_report(input_csv: str, output_xlsx: str):

    # Setup logger
    log_path = output_xlsx.replace('.xlsx', '_log.txt')
    L = Logger(log_path)
    run_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    L.log("=" * 80)
    L.log("  PHONE NO RECORD TRACKER  —  DETAILED CALCULATION LOG")
    L.log(f"  Run timestamp : {run_time}")
    L.log(f"  Input file    : {input_csv}")
    L.log(f"  Output file   : {output_xlsx}")
    L.log("=" * 80)

    # ── STEP 1: Load & filter ─────────────────────────────────────────────────
    L.section("STEP 1 — LOAD & CHANNEL FILTER")
    df_raw = pd.read_csv(input_csv, dtype=str, low_memory=False)
    L.log(f"  Total rows loaded from CSV          : {len(df_raw):,}")
    L.log(f"  Unique invoices in raw data         : {df_raw['invoiceNumber'].nunique():,}")

    channel_counts = df_raw['channel'].str.strip().str.lower().value_counts()
    L.log()
    L.log("  Channel distribution in raw data:")
    for ch, cnt in channel_counts.items():
        L.log(f"    {ch:<30} : {cnt:,}")

    df = df_raw[df_raw['channel'].str.strip().str.lower().isin(['dine in', 'take away', 'dine-in', 'takeaway'])].copy()
    L.log()
    L.log(f"  Rows KEPT after channel filter      : {len(df):,}")
    L.log(f"  Rows DROPPED (other channels)       : {len(df_raw) - len(df):,}")
    L.log(f"  Unique invoices after filter        : {df['invoiceNumber'].nunique():,}")

    # ── STEP 2: Phone cleaning ────────────────────────────────────────────────
    L.section("STEP 2 — PHONE NUMBER CLEANING")
    df['clean_phone'] = df['customer_phoneNumber'].apply(clean_phone)

    total_rows        = len(df)
    raw_has_phone     = df['customer_phoneNumber'].notna().sum()
    cleaned_has_phone = df['clean_phone'].notna().sum()
    dropped_in_clean  = raw_has_phone - cleaned_has_phone

    L.log(f"  Total rows (post channel filter)    : {total_rows:,}")
    L.log(f"  Rows with ANY phone value           : {raw_has_phone:,}")
    L.log(f"  Rows with VALID cleaned phone       : {cleaned_has_phone:,}")
    L.log(f"  Rows where phone was REJECTED       : {dropped_in_clean:,}")
    L.log()
    L.log("  Phone rejection reasons (applied in order):")
    L.log("    1. Empty / NaN value")
    L.log("    2. No digits after stripping special chars")
    L.log("    3. 12-digit starting with 91 → strip country code")
    L.log("    4. 11-digit starting with 0  → strip leading zero")
    L.log("    5. Not exactly 10 digits after cleaning")
    L.log("    6. Any single digit repeats more than 5 times (e.g. 9999999999)")
    L.log("    7. First digit is 0 or 1 (invalid Indian mobile)")
    L.log("    8. First digit not in 6,7,8,9 (invalid Indian mobile prefix)")

    # ── STEP 3: Deduplicate at invoice level ──────────────────────────────────
    L.section("STEP 3 — DEDUPLICATE AT INVOICE LEVEL")
    before_dedup = len(df)
    inv = df.drop_duplicates(subset='invoiceNumber')[
        ['invoiceNumber', 'branchName', 'channel', 'date', 'clean_phone']
    ].copy()
    inv['date'] = pd.to_datetime(inv['date'], errors='coerce').dt.date
    after_dedup = len(inv)

    L.log(f"  Rows before dedup                   : {before_dedup:,}")
    L.log(f"  Unique invoices after dedup         : {after_dedup:,}")
    L.log(f"  Duplicate rows removed              : {before_dedup - after_dedup:,}")
    L.log()
    L.log("  NOTE: When multiple rows share the same invoiceNumber,")
    L.log("  only the FIRST occurrence is kept (drop_duplicates keeps first row).")
    L.log("  This means phone number is taken from the first line item of each invoice.")

    # ── STEP 4: City / Format mapping ─────────────────────────────────────────
    L.section("STEP 4 — BRANCH MAPPING (City & Format)")
    mapping_df = pd.DataFrame.from_dict(mapping_dict, orient='index').reset_index()
    mapping_df.columns = ['branchName', 'City', 'Format']
    inv = inv.merge(mapping_df, on='branchName', how='left')
    inv['City']   = inv['City'].fillna('Unknown')
    inv['Format'] = inv['Format'].fillna('Unknown')

    unknown_branches = inv[inv['City'] == 'Unknown']['branchName'].unique()
    L.log(f"  Branches mapped successfully        : {inv[inv['City'] != 'Unknown']['branchName'].nunique():,}")
    if len(unknown_branches) > 0:
        L.log(f"  Branches NOT found in mapping       : {len(unknown_branches)}")
        for b in unknown_branches:
            L.log(f"    → '{b}'  (will be DROPPED)")
    else:
        L.log("  All branches found in mapping dict  : ✓")

    inv_before_filter = len(inv)
    inv = inv[~inv['City'].isin(['Unknown', 'NA'])]
    inv = inv[inv['Format'] != 'Cloud Kitchen']
    inv = inv[inv['branchName'] != 'RBI Officers Canteen']  # ── excluded outlet
    inv_after_filter  = len(inv)

    format_dist = inv['Format'].value_counts()
    L.log()
    L.log(f"  Invoices dropped (Unknown/NA city)  : {inv_before_filter - inv_after_filter:,}")
    L.log(f"  Cloud Kitchen invoices excluded     : (filtered out entirely)")
    L.log(f"  Invoices remaining for analysis     : {inv_after_filter:,}")
    L.log()
    L.log("  Format distribution (post-filter):")
    for fmt, cnt in format_dist.items():
        L.log(f"    {fmt:<20} : {cnt:,} invoices")

    # ── STEP 5: Week / Day buckets ────────────────────────────────────────────
    L.section("STEP 5 — WEEK & DAY BUCKETS")
    inv['date_dt']  = pd.to_datetime(inv['date'], errors='coerce')
    inv['iso_week'] = inv['date_dt'].dt.isocalendar().week.astype(int)
    inv['iso_year'] = inv['date_dt'].dt.isocalendar().year.astype(int)
    inv['yw']       = inv['iso_year'] * 100 + inv['iso_week']

    last3_yw          = sorted(inv['yw'].dropna().unique())[-3:]
    target_week_pairs = [(int(yw // 100), int(yw % 100)) for yw in last3_yw]
    all_days          = sorted(inv['date'].dropna().unique())
    target_days       = all_days[-3:] if len(all_days) >= 3 else all_days
    latest_yw         = last3_yw[-1] if last3_yw else None

    L.log(f"  Date range in data                  : {min(all_days)} → {max(all_days)}")
    L.log(f"  Total unique days in data           : {len(all_days)}")
    L.log()
    L.log("  Last 3 ISO weeks selected:")
    for (yr, wk) in target_week_pairs:
        yw_key = yr * 100 + wk
        label  = "← WTD (current week)" if yw_key == latest_yw else ""
        wk_inv = inv[(inv['iso_week'] == wk) & (inv['iso_year'] == yr)]
        L.log(f"    Year {yr}  Week {wk:02d}  (yw={yw_key})  →  {wk_inv['invoiceNumber'].nunique():,} invoices  {label}")
    L.log()
    L.log("  Last 3 days selected:")
    for d in target_days:
        day_inv = inv[inv['date'] == d]
        L.log(f"    {d}  →  {day_inv['invoiceNumber'].nunique():,} invoices across {day_inv['branchName'].nunique()} branches")

    # ── STEP 6: Branch-level metrics ──────────────────────────────────────────
    L.section("STEP 6 — BRANCH-LEVEL CAPTURE RATE CALCULATIONS")
    L.log()
    L.log("  FORMULA:")
    L.log("    Capture Rate = Unique valid phone numbers ÷ Unique invoices")
    L.log("    (scoped to: one branch × one time window)")
    L.log()

    order_counts = inv.groupby('branchName')['invoiceNumber'].nunique().reset_index()
    order_counts.columns = ['branchName', 'total_orders']

    branches = (
        inv[['branchName', 'City', 'Format']]
        .drop_duplicates('branchName')
        .merge(order_counts, on='branchName', how='left')
        .sort_values('total_orders', ascending=False)
        .reset_index(drop=True)
    )

    # ── FIXED: Unique valid phone numbers ÷ Unique invoices ──────────────────
    def capture(sub):
        total     = sub['invoiceNumber'].nunique()
        has_phone = sub['clean_phone'].nunique()  # unique valid phone numbers
        return has_phone / total if total > 0 else None

    day_fmt   = lambda d: d.strftime('%b %#d') if os.name == 'nt' else d.strftime('%b %-d')

    # Week cols
    week_cols = []
    L.subsection("Weekly Capture Rates by Branch")
    for (yr, wk) in target_week_pairs:
        yw_key = yr * 100 + wk
        col    = f'W{wk} WTD' if yw_key == latest_yw else f'W{wk}'
        week_cols.append(col)
        L.log(f"  [{col}]  Year={yr}  Week={wk}")
        L.log(f"  {'Branch':<30} {'Unique Invoices':>16} {'Unique Phones':>14} {'Capture Rate':>13}")
        L.log(f"  {'-'*30} {'-'*16} {'-'*14} {'-'*13}")
        rates = []
        for _, row in branches.iterrows():
            sub   = inv[(inv['branchName'] == row['branchName']) &
                        (inv['iso_week'] == wk) & (inv['iso_year'] == yr)]
            rate  = capture(sub)
            rates.append(rate)
            if len(sub) > 0:
                total_inv = sub['invoiceNumber'].nunique()
                uniq_ph   = sub['clean_phone'].nunique()
                rate_str  = f"{rate:.1%}" if rate is not None else "—"
                L.log(f"  {row['branchName']:<30} {total_inv:>16,} {uniq_ph:>14,} {rate_str:>13}")
        branches[col] = rates
        L.log()

    # Day cols
    day_cols = []
    L.subsection("Daily Capture Rates by Branch")
    for d in target_days:
        col = day_fmt(d)
        day_cols.append(col)
        L.log(f"  [{col}]  Date={d}")
        L.log(f"  {'Branch':<30} {'Unique Invoices':>16} {'Unique Phones':>14} {'Capture Rate':>13}")
        L.log(f"  {'-'*30} {'-'*16} {'-'*14} {'-'*13}")
        rates = []
        for _, row in branches.iterrows():
            sub   = inv[(inv['branchName'] == row['branchName']) & (inv['date'] == d)]
            rate  = capture(sub)
            rates.append(rate)
            if len(sub) > 0:
                total_inv = sub['invoiceNumber'].nunique()
                uniq_ph   = sub['clean_phone'].nunique()
                rate_str  = f"{rate:.1%}" if rate is not None else "—"
                L.log(f"  {row['branchName']:<30} {total_inv:>16,} {uniq_ph:>14,} {rate_str:>13}")
        branches[col] = rates
        L.log()

    metric_cols = week_cols + day_cols

    # ── Compute totals row: unique phones ÷ unique invoices per time window ──
    totals = {'City': 'ALL', 'branchName': 'TOTAL', 'Format': ''}
    for (yr, wk) in target_week_pairs:
        yw_key = yr * 100 + wk
        col    = f'W{wk} WTD' if yw_key == latest_yw else f'W{wk}'
        sub    = inv[(inv['iso_week'] == wk) & (inv['iso_year'] == yr)]
        t_inv  = sub['invoiceNumber'].nunique()
        t_ph   = sub['clean_phone'].nunique()
        totals[col] = (t_ph / t_inv) if t_inv > 0 else None
    for d in target_days:
        col   = day_fmt(d)
        sub   = inv[inv['date'] == d]
        t_inv = sub['invoiceNumber'].nunique()
        t_ph  = sub['clean_phone'].nunique()
        totals[col] = (t_ph / t_inv) if t_inv > 0 else None

    # ── STEP 7: Overall summary ───────────────────────────────────────────────
    L.section("STEP 7 — OVERALL SUMMARY")
    total_invoices = inv['invoiceNumber'].nunique()
    total_phones   = inv['clean_phone'].nunique()
    overall_rate   = total_phones / total_invoices * 100 if total_invoices else 0

    L.log(f"  Total unique invoices (Dine In/TA)  : {total_invoices:,}")
    L.log(f"  Total unique valid phone numbers    : {total_phones:,}")
    L.log(f"  Overall capture rate                : {overall_rate:.1f}%")
    L.log(f"  Formula: {total_phones:,} ÷ {total_invoices:,} = {overall_rate:.1f}%")

    L.subsection("Top 15 Most Frequent Phone Numbers (Dummy Number Check)")
    top_phones = inv['clean_phone'].value_counts().head(15)
    L.log(f"  {'Phone Number':<15} {'Count':>8}   {'Flag'}")
    L.log(f"  {'-'*15} {'-'*8}   {'-'*30}")
    for ph, cnt in top_phones.items():
        flag = "⚠️  LIKELY DUMMY — appears very frequently" if cnt > 10 else ""
        L.log(f"  {str(ph):<15} {cnt:>8,}   {flag}")

    L.subsection("Capture Rate by City")
    for city in sorted(inv['City'].unique()):
        city_sub   = inv[inv['City'] == city]
        city_inv   = city_sub['invoiceNumber'].nunique()
        city_ph    = city_sub['clean_phone'].nunique()
        city_rate  = city_ph / city_inv * 100 if city_inv else 0
        L.log(f"  {city:<15}  Invoices={city_inv:,}  Unique Phones={city_ph:,}  Rate={city_rate:.1f}%")

    L.subsection("Capture Rate by Format")
    for fmt in sorted(inv['Format'].unique()):
        fmt_sub  = inv[inv['Format'] == fmt]
        fmt_inv  = fmt_sub['invoiceNumber'].nunique()
        fmt_ph   = fmt_sub['clean_phone'].nunique()
        fmt_rate = fmt_ph / fmt_inv * 100 if fmt_inv else 0
        L.log(f"  {fmt:<15}  Invoices={fmt_inv:,}  Unique Phones={fmt_ph:,}  Rate={fmt_rate:.1f}%")

    L.subsection("All Branches — Overall Capture Rate (sorted by invoices)")
    L.log(f"  {'Branch':<30} {'City':<12} {'Format':<12} {'Invoices':>10} {'Uniq Phones':>12} {'Rate':>8}")
    L.log(f"  {'-'*30} {'-'*12} {'-'*12} {'-'*10} {'-'*12} {'-'*8}")
    for _, row in branches.iterrows():
        b_sub  = inv[inv['branchName'] == row['branchName']]
        b_inv  = b_sub['invoiceNumber'].nunique()
        b_ph   = b_sub['clean_phone'].nunique()
        b_rate = f"{b_ph/b_inv:.1%}" if b_inv else "—"
        L.log(f"  {row['branchName']:<30} {row['City']:<12} {row['Format']:<12} {b_inv:>10,} {b_ph:>12,} {b_rate:>8}")

    L.log()
    L.log("=" * 80)
    L.log("  END OF LOG")
    L.log("=" * 80)
    L.save()

    # 7. Build Excel ────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Phone No Record Tracker"

    ROYAL_BLUE = "1A3A6B"
    HDR_WEEK   = "2C5282"
    HDR_DAY    = "1A3A6B"
    WHITE      = "FFFFFF"
    LIGHT_ROW  = "F7F9FC"
    ALT_ROW    = "EEF2F9"
    BORDER_C   = "C5CFE3"

    thin = Side(style='thin', color=BORDER_C)
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    total_cols      = 3 + len(metric_cols)
    last_col_letter = get_column_letter(total_cols)

    # Title row
    ws.merge_cells(f'A1:{last_col_letter}1')
    c = ws['A1']
    c.value     = "Phone No Record Tracker  |  Dine In & Take Away"
    c.font      = Font(name='Arial', bold=True, size=14, color=WHITE)
    c.fill      = PatternFill('solid', start_color=ROYAL_BLUE)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f'A2:{last_col_letter}2')
    ws['A2'].fill = PatternFill('solid', start_color=ROYAL_BLUE)
    ws.row_dimensions[2].height = 6

    # Group header row (row 3)
    for c_idx in [1, 2, 3]:
        ws.cell(row=3, column=c_idx).fill = PatternFill('solid', start_color=ROYAL_BLUE)

    ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=6)
    wk_cell = ws.cell(row=3, column=4, value="← Last 3 Weeks →")
    wk_cell.font      = Font(name='Arial', bold=True, size=9, color=WHITE)
    wk_cell.fill      = PatternFill('solid', start_color=HDR_WEEK)
    wk_cell.alignment = Alignment(horizontal='center')

    ws.merge_cells(start_row=3, start_column=7, end_row=3, end_column=9)
    dy_cell = ws.cell(row=3, column=7, value="← Last 3 Days →")
    dy_cell.font      = Font(name='Arial', bold=True, size=9, color=WHITE)
    dy_cell.fill      = PatternFill('solid', start_color=HDR_DAY)
    dy_cell.alignment = Alignment(horizontal='center')

    # Column label row (row 4)
    headers  = ['City', 'Branch', 'Format'] + metric_cols
    col_meta = {}
    for h in ['City', 'Branch', 'Format']:
        col_meta[h] = (ROYAL_BLUE, WHITE)
    for c in week_cols:
        col_meta[c] = (HDR_WEEK, WHITE)
    for c in day_cols:
        col_meta[c] = (HDR_DAY, WHITE)

    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=ci, value=h)
        bg, fg = col_meta.get(h, (ROYAL_BLUE, WHITE))
        cell.fill      = PatternFill('solid', start_color=bg)
        cell.font      = Font(name='Arial', bold=True, size=10, color=fg)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = bdr
    ws.row_dimensions[4].height = 28

    # Data rows
    for ri, row in branches.iterrows():
        excel_row = ri + 5
        fill_bg   = LIGHT_ROW if ri % 2 == 0 else ALT_ROW
        vals = [row['City'], row['branchName'], row['Format']] + [row[c] for c in metric_cols]
        for ci, val in enumerate(vals, start=1):
            cell        = ws.cell(row=excel_row, column=ci)
            cell.border = bdr
            if ci <= 3:
                cell.value     = val
                cell.font      = Font(name='Arial', size=9)
                cell.fill      = PatternFill('solid', start_color=fill_bg)
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                if val is None:
                    cell.value     = "—"
                    cell.font      = Font(name='Arial', size=9, color='BBBBBB')
                    cell.fill      = PatternFill('solid', start_color=fill_bg)
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.value         = val
                    cell.number_format = '0.0%'
                    cell.font          = Font(name='Arial', size=9)
                    cell.fill          = PatternFill('solid', start_color=fill_bg)
                    cell.alignment     = Alignment(horizontal='center')

    total_data_rows = len(branches)

    # ── TOTAL row ─────────────────────────────────────────────────────────────
    total_row_num = 5 + total_data_rows
    TOTAL_BG      = "1A3A6B"
    TOTAL_FG      = "FFFFFF"

    # Merge first 3 cols for "TOTAL" label
    ws.merge_cells(start_row=total_row_num, start_column=1, end_row=total_row_num, end_column=3)
    lbl_cell            = ws.cell(row=total_row_num, column=1, value="TOTAL")
    lbl_cell.font       = Font(name='Arial', bold=True, size=10, color=TOTAL_FG)
    lbl_cell.fill       = PatternFill('solid', start_color=TOTAL_BG)
    lbl_cell.alignment  = Alignment(horizontal='center', vertical='center')
    lbl_cell.border     = bdr
    # border on merged cells 2 & 3
    for c_idx in [2, 3]:
        ws.cell(row=total_row_num, column=c_idx).border = bdr

    for ci, col in enumerate(metric_cols, start=4):
        val  = totals.get(col)
        cell = ws.cell(row=total_row_num, column=ci)
        cell.fill   = PatternFill('solid', start_color=TOTAL_BG)
        cell.border = bdr
        if val is None:
            cell.value     = "—"
            cell.font      = Font(name='Arial', bold=True, size=10, color="AAAAAA")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            cell.value         = val
            cell.number_format = '0.0%'
            cell.font          = Font(name='Arial', bold=True, size=10, color=TOTAL_FG)
            cell.alignment     = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[total_row_num].height = 22

    last_row     = total_row_num
    cf_last_row  = total_row_num - 1  # ── exclude TOTAL row from conditional formatting

    # Conditional formatting: ONLY on day cols, ONLY on data rows (not total)
    day_col_start = 3 + len(week_cols) + 1
    for ci_offset in range(len(day_cols)):
        ci         = day_col_start + ci_offset
        col_letter = get_column_letter(ci)
        ws.conditional_formatting.add(
            f"{col_letter}5:{col_letter}{cf_last_row}",
            ColorScaleRule(
                start_type='min', start_color='F8696B',
                end_type='max',   end_color='FFFFFF'
            )
        )

    col_widths = [14, 28, 14, 9, 9, 9, 9, 9, 9]
    for ci, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = 'D5'

    # 8. Raw Dump sheet ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Raw Dump")

    extra_cols_wanted = ['invoiceDate', 'customer_name', 'customer_email',
                         'customer_id', 'orderNumber', 'netAmount', 'grossAmount']
    available_extra = [c for c in extra_cols_wanted if c in df.columns]

    df['_qty'] = pd.to_numeric(df.get('item_quantity', 0), errors='coerce').fillna(0)
    sale_qty   = df.groupby('invoiceNumber')['_qty'].sum().reset_index()
    sale_qty.columns = ['invoiceNumber', 'sale_qty']

    raw = df.drop_duplicates(subset='invoiceNumber')[
        ['invoiceNumber', 'branchName', 'channel', 'date', 'clean_phone'] + available_extra
    ].copy()
    raw = raw.merge(mapping_df[['branchName', 'City', 'Format']], on='branchName', how='left')
    raw = raw.merge(sale_qty, on='invoiceNumber', how='left')
    raw['City']   = raw['City'].fillna('Unknown')
    raw['Format'] = raw['Format'].fillna('Unknown')
    raw['date']   = pd.to_datetime(raw['date'], errors='coerce').dt.date

    # ── EXCLUDE Cloud Kitchen and specific outlets from Raw Dump ────────────
    raw = raw[raw['Format'] != 'Cloud Kitchen']
    raw = raw[raw['branchName'] != 'RBI Officers Canteen']  # ── excluded outlet

    for nc in ['netAmount', 'grossAmount']:
        if nc in raw.columns:
            raw[nc] = pd.to_numeric(raw[nc], errors='coerce')

    raw_dump = (
        raw[raw['clean_phone'].notna()]
        .merge(order_counts, on='branchName', how='left')
        .sort_values(['total_orders', 'date', 'branchName'], ascending=[False, True, True])
        .drop_duplicates(subset='clean_phone', keep='first')  # ── one row per unique phone
        .drop(columns=['total_orders'])
        .reset_index(drop=True)
    )

    ordered_cols = ['invoiceNumber', 'date']
    for c in ['invoiceDate', 'branchName', 'City', 'Format', 'channel',
              'customer_name', 'clean_phone', 'customer_email', 'customer_id',
              'orderNumber', 'sale_qty', 'netAmount', 'grossAmount']:
        if c in raw_dump.columns:
            ordered_cols.append(c)
    raw_dump = raw_dump[ordered_cols]

    header_labels = {
        'invoiceNumber': 'Invoice No', 'date': 'Date', 'invoiceDate': 'Invoice Date',
        'branchName': 'Branch', 'City': 'City', 'Format': 'Format', 'channel': 'Channel',
        'customer_name': 'Customer Name', 'clean_phone': 'Phone Number',
        'customer_email': 'Email', 'customer_id': 'Customer ID', 'orderNumber': 'Order No',
        'sale_qty': 'Sale Qty', 'netAmount': 'Net Amount', 'grossAmount': 'Gross Amount',
    }

    numeric_cols_raw = {'netAmount', 'grossAmount', 'sale_qty'}
    left_cols        = {'branchName', 'customer_name', 'customer_email'}
    hdr_fill  = PatternFill('solid', start_color="1A3A6B")
    hdr_font  = Font(name='Arial', bold=True, size=10, color="FFFFFF")
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for ci, col in enumerate(ordered_cols, start=1):
        cell = ws2.cell(row=1, column=ci, value=header_labels.get(col, col))
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = hdr_align; cell.border = bdr
    ws2.row_dimensions[1].height = 28

    for ri, row_data in raw_dump.iterrows():
        excel_row = ri + 2
        fill_bg   = "F9FAFB" if ri % 2 == 0 else "FFFFFF"
        for ci, col in enumerate(ordered_cols, start=1):
            val  = row_data[col]
            cell = ws2.cell(row=excel_row, column=ci)
            cell.fill   = PatternFill('solid', start_color=fill_bg)
            cell.font   = Font(name='Arial', size=9)
            cell.border = bdr
            if col in numeric_cols_raw:
                try:
                    cell.value = float(val) if val is not None and str(val) != 'nan' else None
                except (ValueError, TypeError):
                    cell.value = None
                cell.number_format = '#,##0.00' if col in {'netAmount', 'grossAmount'} else '#,##0'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif col == 'clean_phone':
                cell.value         = str(val) if val is not None else None
                cell.number_format = '@'
                cell.alignment     = Alignment(horizontal='center', vertical='center')
            else:
                cell.value     = val
                cell.alignment = Alignment(
                    horizontal='left' if col in left_cols else 'center',
                    vertical='center'
                )

    col_widths2 = {
        'invoiceNumber': 16, 'date': 12, 'invoiceDate': 14, 'branchName': 26,
        'City': 12, 'Format': 14, 'channel': 12, 'customer_name': 20,
        'clean_phone': 15, 'customer_email': 26, 'customer_id': 14,
        'orderNumber': 14, 'sale_qty': 10, 'netAmount': 14, 'grossAmount': 14,
    }
    for ci, col in enumerate(ordered_cols, start=1):
        ws2.column_dimensions[get_column_letter(ci)].width = col_widths2.get(col, 14)

    ws2.freeze_panes = 'A2'

    wb.save(output_xlsx)
    phone_fill_pct = (inv['clean_phone'].nunique() / inv['invoiceNumber'].nunique() * 100) if len(inv) else 0
    print(f"✅ Report saved      → {output_xlsx}")
    print(f"📋 Log saved         → {log_path}")
    print(f"   Branches analysed             : {len(branches)}")
    print(f"   Total invoices (Dine In / TA) : {inv['invoiceNumber'].nunique():,}")
    print(f"   Raw Dump rows (with phone)    : {len(raw_dump):,}")
    print(f"   Overall phone capture         : {phone_fill_pct:.1f}%")


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import sys

    ROLLING_CSV = r"C:/Users/jeryy/OneDrive - BLISS CHOCOLATES INDIA PRIVATE LIMITED/Desktop/Documents/Phone No/Data/sales_rolling.csv"
    OUTPUT_DIR  = r"C:/Users/jeryy/OneDrive - BLISS CHOCOLATES INDIA PRIVATE LIMITED/Desktop/Documents/Phone No/output"

    input_file = sys.argv[1] if len(sys.argv) > 1 else ROLLING_CSV
    if len(sys.argv) > 2:
        output_file = sys.argv[2]
    else:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        from datetime import timedelta
        date_str    = (datetime.today() - timedelta(days=1)).strftime("%Y-%m-%d")
        output_file = os.path.join(OUTPUT_DIR, f"Phone No Record - {date_str}.xlsx")

    build_report(input_file, output_file)