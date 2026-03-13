
# """

# python send_report.py

# Auto-picks yesterday's Excel → reads it → sends responsive HTML email

# Desktop: full scrollable table
# Mobile:  fixed-width no-scroll table, all columns, abbreviated labels

# """

# import os, re, smtplib, ssl

# from datetime import datetime, timedelta

# from email.mime.multipart import MIMEMultipart

# from email.mime.base import MIMEBase

# from email.mime.text import MIMEText

# from email import encoders

# from openpyxl import load_workbook

# from openpyxl.utils import column_index_from_string



# # ── CONFIG ────────────────────────────────────────────────────────────────────

# OUTPUT_DIR      = r"C:/Users/jeryy/OneDrive - BLISS CHOCOLATES INDIA PRIVATE LIMITED/Desktop/Documents/Phone No/output"

# SMTP_HOST       = "smtp.office365.com"

# SMTP_PORT       = 587

# SENDER_EMAIL    = "finance.reports@smoorchocolates.com"

# SENDER_PASSWORD = "vnnyxwnxypqhxdqj"

# TO_RECIPIENTS   = [   
    
#     #  "muskan.sajwan@smoorchocolates.com",
#     # "trisha.krishna@smoorchocolates.com",
#     # "kanchan.achpal@smoorchocolates.com",		

#     # "vedant.gote@smoorchocolates.com",
#     # "adeshshetty@smoorchocolates.com",
#     # "mehtaj.khan@smoorchocolates.com",
#     # "ranbir.singh@smoorchocolates.com",
#     # "jagdish.duggal@smoorchocolates.com",
#     # "hussain@smoorchocolates.com",

#     # "suresh.patro@smoorchocolates.com",
#     # "ashwin@smoorchocolates.com",

#     # "deependra.rao@smoorchocolates.com",
#     "john.jerry@smoorchocolates.com",
#     # "siddhartha.siva@smoorchocolates.com",
#     # "rahul.kumar@smoorchocolates.com" 
#     ]

# CC_RECIPIENTS   = []

# BCC_RECIPIENTS  = []



# # ── COLORS ────────────────────────────────────────────────────────────────────

# ROYAL  = "#1A3A6B"

# WBLUE  = "#2C5282"

# WHITE  = "#FFFFFF"

# RLIGHT = "#F7F9FC"

# RALT   = "#EEF2F9"

# BORD   = "#C5CFE3"



# # ── ABBREVIATION MAPS ─────────────────────────────────────────────────────────

# CITY_SHORT = {
#     "Bangalore": "BLR", "Bengaluru": "BLR",
#     "Mumbai":    "MUM", "Bombay":    "MUM",
#     "Chennai":   "CHN", "Madras":    "CHN",
#     "Delhi":     "NCR", "New Delhi": "NCR", "NCR": "NCR",
#     "Gurgaon":   "NCR", "Noida":     "NCR",
#     "Pune":      "PUN",
#     "Hyderabad": "HYD",
#     "Kolkata":   "KOL",
#     "Ahmedabad": "AMD",
# }

# FORMAT_SHORT = {
#     "Signature":     "Sig",
#     "Cafe":          "Caf", "Café": "Caf",
#     "Kiosk":         "Kio",
#     "Choco Kitchen": "CK",  "ChocKitchen": "CK", "Choco-Kitchen": "CK",
#     "Cloud Kitchen": "CK",  "CloudKitchen": "CK", "Cloud-Kitchen": "CK",
#     "Airport":       "Air",
# }

# def short_city(val):
#     if not val: return ""
#     s = str(val).strip()
#     return CITY_SHORT.get(s, s[:3].upper())

# def short_format(val):
#     if not val: return ""
#     s = str(val).strip()
#     if s in FORMAT_SHORT: return FORMAT_SHORT[s]
#     for k, v in FORMAT_SHORT.items():
#         if k.lower() in s.lower():
#             return v
#     return s[:3]



# # ── COLOR SCALE: white=0% → red=100% ─────────────────────────────────────────

# def pct_to_bg(val):

#     t = max(0.0, min(1.0, float(val)))

#     r = int(255 + (248 - 255) * t)

#     g = int(255 + (105 - 255) * t)

#     b = int(255 + (107 - 255) * t)

#     return f"#{r:02X}{g:02X}{b:02X}"



# # ── FIND YESTERDAY'S FILE ─────────────────────────────────────────────────────

# def get_excel_file():

#     yesterday = (datetime.today() - timedelta(days=1)).strftime("%Y-%m-%d")

#     path = os.path.join(OUTPUT_DIR, f"Phone No Record - {yesterday}.xlsx")

#     if os.path.exists(path):

#         return path

#     files = sorted(

#         [f for f in os.listdir(OUTPUT_DIR)

#          if f.startswith("Phone No Record") and f.endswith(".xlsx")],

#         reverse=True

#     )

#     if files:

#         print(f"  Using latest file: {files[0]}")

#         return os.path.join(OUTPUT_DIR, files[0])

#     raise FileNotFoundError(f"No Excel found in {OUTPUT_DIR}")



# # ── READ EXCEL → HTML TABLES ──────────────────────────────────────────────────

# def excel_to_html(xlsx_path):

#     wb = load_workbook(xlsx_path, data_only=True)

#     ws = wb["Phone No Record Tracker"]



#     # Row 4: column headers + colors

#     headers = []

#     for c in range(1, ws.max_column + 1):

#         v = ws.cell(4, c).value

#         if v:

#             rgb = ws.cell(4, c).fill.fgColor.rgb[2:]

#             headers.append((c, v, "#" + rgb))



#     # CF cols = day cols

#     cf_cols = set()

#     for cf in ws.conditional_formatting:

#         m = re.match(r"([A-Z]+)", str(cf).strip("<>ConditionalFormatting "))

#         if m:

#             cf_cols.add(column_index_from_string(m.group(1)))



#     # Row 3 group spans

#     group_spans = {}

#     for mr in ws.merged_cells.ranges:

#         m = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", str(mr))

#         if m and int(m.group(2)) == 3:

#             sc = column_index_from_string(m.group(1))

#             ec = column_index_from_string(m.group(3))

#             lbl = ws.cell(3, sc).value

#             if lbl:

#                 group_spans[sc] = (ec, lbl)



#     city_idx   = next((i for i,(c,v,_) in enumerate(headers) if v=="City"), 0)

#     branch_idx = next((i for i,(c,v,_) in enumerate(headers) if v=="Branch"), 1)

#     fmt_idx    = next((i for i,(c,v,_) in enumerate(headers) if v=="Format"), None)



#     # Data rows

#     data = []

#     for r in range(5, ws.max_row + 1):

#         row = [ws.cell(r, c).value for c in range(1, len(headers) + 1)]

#         if not any(v is not None for v in row): continue

#         if row[city_idx] in (None, "", "Unknown"): continue

#         data.append(row)



#     title = ws["A1"].value or "Phone No Record Tracker"



#     # ── Desktop helpers ───────────────────────────────────────────────────────

#     BASE = ("padding:6px 8px; font-size:11px; font-weight:400; color:#1A1A1A;"

#             f" border-bottom:1px solid {BORD}; border-right:1px solid {BORD};"

#             " vertical-align:middle; white-space:nowrap;")

#     NUM  = BASE + " text-align:center;"

#     TXT  = BASE + " text-align:left;"



#     def render_val_td(c, name, val, bg, extra_style=""):

#         w_map = {"City": "90px", "Branch": "150px", "Format": "72px",

#                  "Hist High": "62px", "__week__": "58px", "__day__": "58px"}

#         w = w_map.get(name, "58px")

#         ws_ = f"width:{w}; min-width:{w};"

#         if name == "City":

#             return f'<td style="{TXT}{ws_} background:{bg}; {extra_style}">{val or ""}</td>'

#         elif name == "Branch":

#             return f'<td style="{TXT}{ws_} background:{bg}; {extra_style} max-width:150px; overflow:hidden; text-overflow:ellipsis;">{val or ""}</td>'

#         elif name == "Format":

#             return f'<td style="{BASE}{ws_} background:{bg}; text-align:center; font-size:10px; color:#666; {extra_style}">{val or ""}</td>'

#         elif name == "Hist High":

#             txt = "—" if val is None else f"{float(val)*100:.1f}%"

#             clr = "#BBB" if val is None else "#1A1A1A"

#             return f'<td style="{NUM}{ws_} background:{bg}; color:{clr}; border-left:2px solid {BORD}; {extra_style}">{txt}</td>'

#         elif c in cf_cols:

#             if val is None:

#                 return f'<td style="{NUM}{ws_} background:{RLIGHT}; color:#BBB; {extra_style}">—</td>'

#             cbg = pct_to_bg(float(val))

#             return f'<td style="{NUM}{ws_} background:{cbg}; {extra_style}">{float(val)*100:.1f}%</td>'

#         else:

#             txt = "—" if val is None else f"{float(val)*100:.1f}%"

#             clr = "#BBB" if val is None else "#1A1A1A"

#             return f'<td style="{NUM}{ws_} background:{bg}; color:{clr}; {extra_style}">{txt}</td>'



#     def th(label, bg, width, extra=""):

#         return (f'<th style="width:{width}; min-width:{width};'

#                 f' padding:7px 8px; font-size:10px; font-weight:700; background:{bg};'

#                 f' color:{WHITE}; text-align:center; white-space:nowrap;'

#                 f' border-right:1px solid rgba(255,255,255,0.18); {extra}">{label}</th>')



#     # ════════════════════════════════════════════════════════
#     #  DESKTOP TABLE
#     # ════════════════════════════════════════════════════════

#     def build_desktop_table():

#         group_row = ""

#         covered = set()

#         for c in range(1, len(headers) + 1):

#             if c in covered: continue

#             if c in group_spans:

#                 ec, lbl = group_spans[c]

#                 span = ec - c + 1

#                 bg   = WBLUE if "Week" in lbl else ROYAL

#                 group_row += (f'<th colspan="{span}" style="background:{bg}; color:{WHITE};'

#                               f' font-size:9px; font-weight:700; text-align:center;'

#                               f' padding:3px 8px; border-right:1px solid rgba(255,255,255,0.18);">{lbl}</th>')

#                 for x in range(c, ec+1): covered.add(x)

#             else:

#                 covered.add(c)

#                 group_row += f'<th style="background:{ROYAL}; padding:3px;"></th>'



#         col_row = ""

#         for c, v, hbg in headers:

#             extra = "border-left:2px solid rgba(255,255,255,0.35);" if v == "Hist High" else ""

#             label = "Hist&nbsp;High" if v == "Hist High" else v

#             w_map = {"City":"90px","Branch":"150px","Format":"72px","Hist High":"62px"}

#             w = w_map.get(v, "58px")

#             col_row += th(label, hbg, w, extra)



#         rows = ""

#         for ri, row in enumerate(data):

#             bg = RLIGHT if ri % 2 == 0 else RALT

#             rows += "<tr>"

#             for ci, (c, name, _) in enumerate(headers):

#                 rows += render_val_td(c, name, row[ci], bg)

#             rows += "</tr>"



#         return (f'<div style="overflow-x:auto; -webkit-overflow-scrolling:touch;">'

#                 f'<table style="border-collapse:collapse; font-family:Arial,sans-serif;'

#                 f' table-layout:auto; border:1px solid {BORD}; min-width:600px;">'

#                 f'<thead><tr>{group_row}</tr><tr>{col_row}</tr></thead>'

#                 f'<tbody>{rows}</tbody></table></div>')



#     # ════════════════════════════════════════════════════════
#     #  MOBILE TABLE — fixed width, NO scroll, ALL columns
#     #
#     #  City+Fmt merged into one stacked cell (38px) to free
#     #  space for full Branch/outlet name (118px).
#     #
#     #  Col widths (px):
#     #    CityFmt=38 | Branch=118 | each Week=28 | each Day=28 | HH=26
#     #  Font: 7.5px headers, 8px data
#     # ════════════════════════════════════════════════════════

#     def build_mobile_table():

#         week_hdrs = [(ci, c, v, hbg) for ci,(c,v,hbg) in enumerate(headers)
#                      if v not in ("City","Branch","Format","Hist High") and c not in cf_cols]

#         day_hdrs  = [(ci, c, v, hbg) for ci,(c,v,hbg) in enumerate(headers) if c in cf_cols]

#         hist_hdrs = [(ci, c, v, hbg) for ci,(c,v,hbg) in enumerate(headers) if v == "Hist High"]

#         n_week = len(week_hdrs)
#         n_day  = len(day_hdrs)
#         n_hist = len(hist_hdrs)

#         # Column pixel widths — City+Fmt merged, Branch gets full space
#         W = {"cityfmt": 38, "branch": 118, "week": 28, "day": 28, "hh": 26}

#         total_w = W["cityfmt"] + W["branch"] + n_week*W["week"] + n_day*W["day"] + n_hist*W["hh"]

#         # Micro cell base styles
#         CB = (f"border-bottom:1px solid {BORD}; border-right:1px solid {BORD};"
#               f" vertical-align:middle; padding:3px 1px; overflow:hidden;"
#               f" white-space:nowrap;")

#         def mth(label, bg, w_px, extra=""):
#             return (f'<th style="width:{w_px}px; min-width:{w_px}px; max-width:{w_px}px;'
#                     f' padding:3px 1px; font-size:7.5px; font-weight:700; background:{bg};'
#                     f' color:{WHITE}; text-align:center; overflow:hidden;'
#                     f' border-right:1px solid rgba(255,255,255,0.2); {extra}">{label}</th>')

#         def mgth(label, bg, span, extra=""):
#             return (f'<th colspan="{span}" style="padding:2px 1px; font-size:7px; font-weight:700;'
#                     f' background:{bg}; color:{WHITE}; text-align:center;'
#                     f' border-right:1px solid rgba(255,255,255,0.25); {extra}">{label}</th>')

#         # Group header row — City+Fmt+Branch = 2 cols now
#         group_row = (
#             mgth("", ROYAL, 2) +
#             (mgth("&#8592; Wks &#8594;", WBLUE, n_week) if n_week else "") +
#             (mgth("&#8592; Days &#8594;", ROYAL, n_day)  if n_day  else "") +
#             (mgth("HH", ROYAL, n_hist)                   if n_hist else "")
#         )

#         # Column header row — stacked City/Fmt label in one th
#         col_row = (
#             mth("City/Fmt", ROYAL, W["cityfmt"]) +
#             mth("Outlet",   ROYAL, W["branch"], "text-align:left;")
#         )
#         for _, _, v, hbg in week_hdrs:
#             col_row += mth(v, hbg, W["week"])
#         for _, _, v, hbg in day_hdrs:
#             col_row += mth(v, hbg, W["day"])
#         for _, _, v, hbg in hist_hdrs:
#             col_row += mth("HH", hbg, W["hh"])

#         # Data rows
#         rows = ""

#         for ri, row in enumerate(data):

#             bg = RLIGHT if ri % 2 == 0 else RALT

#             city_val   = short_city(row[city_idx])
#             branch_val = str(row[branch_idx] or "")
#             fmt_val    = short_format(row[fmt_idx]) if fmt_idx is not None else ""

#             rows += "<tr>"

#             # City + Format stacked in one cell: "BLR" on top, "Sig" below in grey
#             rows += (f'<td style="{CB} width:{W["cityfmt"]}px; max-width:{W["cityfmt"]}px;'
#                      f' background:{bg}; text-align:center; padding:2px 1px; line-height:1.2;">'
#                      f'<span style="display:block; font-size:8px; font-weight:700; color:{ROYAL};">{city_val}</span>'
#                      f'<span style="display:block; font-size:7px; color:#888;">{fmt_val}</span>'
#                      f'</td>')

#             # Branch — full name, wraps to 2 lines if needed (white-space:normal)
#             rows += (f'<td style="{CB} width:{W["branch"]}px; max-width:{W["branch"]}px;'
#                      f' background:{bg}; font-size:8px; font-weight:500; text-align:left;'
#                      f' padding-left:4px; white-space:normal; line-height:1.25;">{branch_val}</td>')

#             # Week cols
#             for ci, c, name, _ in week_hdrs:
#                 val = row[ci]
#                 txt = "—" if val is None else f"{float(val)*100:.0f}%"
#                 clr = "#BBB" if val is None else "#1A1A1A"
#                 rows += (f'<td style="{CB} width:{W["week"]}px; max-width:{W["week"]}px;'
#                          f' background:{bg}; font-size:8px; text-align:center; color:{clr};">{txt}</td>')

#             # Day cols (color scale)
#             for ci, c, name, _ in day_hdrs:
#                 val = row[ci]
#                 if val is None:
#                     rows += (f'<td style="{CB} width:{W["day"]}px; max-width:{W["day"]}px;'
#                              f' background:{RLIGHT}; font-size:8px; text-align:center; color:#BBB;">—</td>')
#                 else:
#                     cbg = pct_to_bg(float(val))
#                     rows += (f'<td style="{CB} width:{W["day"]}px; max-width:{W["day"]}px;'
#                              f' background:{cbg}; font-size:8px; text-align:center; color:#1A1A1A;">'
#                              f'{float(val)*100:.0f}%</td>')

#             # Hist High
#             for ci, c, name, _ in hist_hdrs:
#                 val = row[ci]
#                 txt = "—" if val is None else f"{float(val)*100:.0f}%"
#                 clr = "#BBB" if val is None else "#1A1A1A"
#                 rows += (f'<td style="{CB} width:{W["hh"]}px; max-width:{W["hh"]}px;'
#                          f' background:{bg}; font-size:8px; text-align:center; color:{clr};'
#                          f' border-left:1px solid {BORD};">{txt}</td>')

#             rows += "</tr>"

#         return (
#             f'<table style="border-collapse:collapse; font-family:Arial,sans-serif;'
#             f' table-layout:fixed; width:{total_w}px; border:1px solid {BORD};">'
#             f'<thead><tr>{group_row}</tr><tr>{col_row}</tr></thead>'
#             f'<tbody>{rows}</tbody>'
#             f'</table>'
#         )



#     desktop = build_desktop_table()

#     mobile  = build_mobile_table()

#     return desktop, mobile, title



# # ── ATTACH FILE ───────────────────────────────────────────────────────────────

# def attach_file(msg, path):

#     with open(path, "rb") as f:

#         part = MIMEBase("application", "octet-stream")

#         part.set_payload(f.read())

#     encoders.encode_base64(part)

#     part.add_header("Content-Disposition", f'attachment; filename="{os.path.basename(path)}"')

#     msg.attach(part)



# # ── MAIN ──────────────────────────────────────────────────────────────────────

# if __name__ == "__main__":

#     excel_file = get_excel_file()

#     print(f"Using: {excel_file}")



#     desktop_table, mobile_table, title = excel_to_html(excel_file)

#     subject = f"Phone No Record Tracker - {datetime.today().strftime('%d %b %Y')}"



#     html = f"""<!DOCTYPE html>
# <html xmlns:v="urn:schemas-microsoft-com:vml" xmlns:o="urn:schemas-microsoft-com:office:office">
# <head>
# <meta name="viewport" content="width=device-width, initial-scale=1.0">
# <!--[if mso]>
# <style type="text/css">
#   .mobile-view {{ display:none !important; mso-hide:all !important; }}
# </style>
# <![endif]-->
# <style>
#   /* Default (desktop email clients that read CSS) */
#   .desktop-view {{ display:block; }}
#   .mobile-view  {{ display:none; mso-hide:all; }}
#   /* Small screens: flip visibility */
#   @media only screen and (max-width: 600px) {{
#     .desktop-view {{ display:none !important; mso-hide:all !important; }}
#     .mobile-view  {{ display:block !important; mso-hide:none !important; }}
#     body {{ padding:4px !important; margin:0 !important; }}
#   }}
# </style>
# </head>
# <body style="margin:0; padding:16px; background:#EEF2F9; font-family:Arial,sans-serif;">

#   <!-- ░░ HEADER ░░ -->
#   <div style="background:{ROYAL}; padding:14px 16px; border-radius:4px 4px 0 0;">
#     <div style="font-size:14px; font-weight:700; color:{WHITE};">{title}</div>
#     <div style="font-size:10px; color:rgba(255,255,255,0.6); margin-top:2px;">
#       {datetime.today().strftime('%d %b %Y')}
#     </div>
#   </div>

#   <!-- ░░ DESKTOP (hidden on mobile via media query) ░░ -->
#   <div class="desktop-view" style="background:{WHITE}; border:1px solid {BORD};
#        border-top:none; padding:10px 16px 14px;">
#     <div style="font-size:10px; color:#2C5282; background:#EEF2F9;
#          border-left:3px solid {ROYAL}; padding:7px 12px; margin-bottom:10px;
#          border-radius:0 3px 3px 0; font-weight:500;">
#       Note: Outlets are sorted basis number of offline orders (high to low)
#     </div>
#     {desktop_table}
#   </div>

#   <!--[if !mso]><!-->
#   <!-- ░░ MOBILE: hidden on Outlook desktop via mso-hide:all ░░ -->
#   <div class="mobile-view" style="display:none; mso-hide:all; background:{WHITE};
#        border:1px solid {BORD}; border-top:none; padding:6px 4px 8px;">
#     <div style="font-size:9px; color:#2C5282; background:#EEF2F9;
#          border-left:3px solid {ROYAL}; padding:5px 8px; margin-bottom:6px;
#          border-radius:0 3px 3px 0; font-weight:500;">
#       Note: Outlets are sorted basis number of offline orders (high to low)
#     </div>
#     {mobile_table}
#   </div>
#   <!--<![endif]-->

#   <!-- ░░ FOOTER ░░ -->
#   <div style="background:{WHITE}; border:1px solid {BORD}; border-top:none;
#        padding:8px 16px 10px; border-radius:0 0 4px 4px;">
    
#     <div style="font-size:10px; color:#AAAAAA;">
#       Capture Rate = unique valid phones &divide; unique invoices &nbsp;|&nbsp;
#       Generated {datetime.now().strftime('%d %b %Y, %I:%M %p')}
#     </div>
#   </div>

# </body>
# </html>"""



#     alt   = MIMEMultipart("alternative")

#     alt.attach(MIMEText(html, "html"))

#     mixed = MIMEMultipart("mixed")

#     mixed["From"]    = SENDER_EMAIL

#     mixed["To"]      = ", ".join(TO_RECIPIENTS)

#     if CC_RECIPIENTS: mixed["CC"] = ", ".join(CC_RECIPIENTS)

#     mixed["Subject"] = subject

#     mixed.attach(alt)

#     attach_file(mixed, excel_file)



#     print(f"Connecting to {SMTP_HOST}:{SMTP_PORT} ...")

#     ctx = ssl.create_default_context()

#     try:

#         with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as s:

#             s.ehlo(); s.starttls(context=ctx); s.ehlo()

#             s.login(SENDER_EMAIL, SENDER_PASSWORD)

#             s.sendmail(SENDER_EMAIL,

#                        TO_RECIPIENTS + CC_RECIPIENTS + BCC_RECIPIENTS,

#                        mixed.as_string())

#         print(f"Sent to {', '.join(TO_RECIPIENTS)}")

#     except smtplib.SMTPAuthenticationError:

#         print("Auth failed — check App Password.")

#     except Exception as e:

#         print(f"Error: {e}")

# """

# python send_report.py

# Auto-picks yesterday's Excel → reads it → sends responsive HTML email

# Desktop: full scrollable table
# Mobile:  fixed-width no-scroll table, all columns, abbreviated labels

# """

# import os, re, smtplib, ssl

# from datetime import datetime, timedelta

# from email.mime.multipart import MIMEMultipart

# from email.mime.base import MIMEBase

# from email.mime.text import MIMEText

# from email import encoders

# from openpyxl import load_workbook

# from openpyxl.utils import column_index_from_string



# # ── CONFIG ────────────────────────────────────────────────────────────────────

# OUTPUT_DIR      = r"C:/Users/jeryy/OneDrive - BLISS CHOCOLATES INDIA PRIVATE LIMITED/Desktop/Documents/Phone No/output"

# SMTP_HOST       = "smtp.office365.com"

# SMTP_PORT       = 587

# SENDER_EMAIL    = "finance.reports@smoorchocolates.com"

# SENDER_PASSWORD = "vnnyxwnxypqhxdqj"

# TO_RECIPIENTS   = [   
    
#     #  "muskan.sajwan@smoorchocolates.com",
#     # "trisha.krishna@smoorchocolates.com",
#     # "kanchan.achpal@smoorchocolates.com",		
#     # "nitesh.kumar@smoorchocolates.com",
#     # "vedant.gote@smoorchocolates.com",
#     # "adeshshetty@smoorchocolates.com",
#     # "mehtaj.khan@smoorchocolates.com",
#     # "ranbir.singh@smoorchocolates.com",
#     # "jagdish.duggal@smoorchocolates.com",
#     # "hussain@smoorchocolates.com",

#     # "suresh.patro@smoorchocolates.com",
#     # "ashwin@smoorchocolates.com",

#     # "deependra.rao@smoorchocolates.com",
#     "john.jerry@smoorchocolates.com",
#     # "siddhartha.siva@smoorchocolates.com",
#     # "rahul.kumar@smoorchocolates.com" 
#     ]

# CC_RECIPIENTS   = []

# BCC_RECIPIENTS  = []



# # ── COLORS ────────────────────────────────────────────────────────────────────

# ROYAL  = "#1A3A6B"

# WBLUE  = "#2C5282"

# WHITE  = "#FFFFFF"

# RLIGHT = "#F7F9FC"

# RALT   = "#EEF2F9"

# BORD   = "#C5CFE3"



# # ── ABBREVIATION MAPS ─────────────────────────────────────────────────────────

# CITY_SHORT = {
#     "Bangalore": "BLR", "Bengaluru": "BLR",
#     "Mumbai":    "MUM", "Bombay":    "MUM",
#     "Chennai":   "CHN", "Madras":    "CHN",
#     "Delhi":     "NCR", "New Delhi": "NCR", "NCR": "NCR",
#     "Gurgaon":   "NCR", "Noida":     "NCR", "Delhi NCR": "NCR",
#     "Pune":      "PUN",
#     "Hyderabad": "HYD",
#     "Kolkata":   "KOL",
#     "Ahmedabad": "AMD",
# }

# FORMAT_SHORT = {
#     "Signature":     "Sig",
#     "Cafe":          "Caf", "Café": "Caf",
#     "Kiosk":         "Kio",
#     "Choco Kitchen": "CK",  "ChocKitchen": "CK", "Choco-Kitchen": "CK",
#     "Cloud Kitchen": "CK",  "CloudKitchen": "CK", "Cloud-Kitchen": "CK",
#     "Airport":       "Air",
# }

# def short_city(val):
#     if not val: return ""
#     s = str(val).strip()
#     return CITY_SHORT.get(s, s[:3].upper())

# def short_format(val):
#     if not val: return ""
#     s = str(val).strip()
#     if s in FORMAT_SHORT: return FORMAT_SHORT[s]
#     for k, v in FORMAT_SHORT.items():
#         if k.lower() in s.lower():
#             return v
#     return s[:3]



# # ── COLOR SCALE: red=0% (lowest) → white=100% (highest) ──────────────────────
# # Low capture rate = bad = RED
# # High capture rate = good = WHITE

# def pct_to_bg(val):
#     # t=0 → light red (#F8696B), t=1 → white (#FFFFFF)
#     t = max(0.0, min(1.0, float(val)))
#     r = int(248 + (255 - 248) * t)   # 248 → 255
#     g = int(105 + (255 - 105) * t)   # 105 → 255
#     b = int(107 + (255 - 107) * t)   # 107 → 255
#     return f"#{r:02X}{g:02X}{b:02X}"



# # ── FIND YESTERDAY'S FILE ─────────────────────────────────────────────────────

# def get_excel_file():
#     # Always try yesterday first
#     yesterday     = (datetime.today() - timedelta(days=1)).strftime("%Y-%m-%d")
#     yesterday_path = os.path.join(OUTPUT_DIR, f"Phone No Record - {yesterday}.xlsx")

#     if os.path.exists(yesterday_path):
#         print(f"  ✅ Found yesterday's file: Phone No Record - {yesterday}.xlsx")
#         return yesterday_path

#     # Fallback: latest available file
#     print(f"  ⚠️  Yesterday's file not found: Phone No Record - {yesterday}.xlsx")
#     print(f"  🔍 Searching for latest available file in: {OUTPUT_DIR}")

#     files = sorted(
#         [f for f in os.listdir(OUTPUT_DIR)
#          if f.startswith("Phone No Record") and f.endswith(".xlsx")],
#         reverse=True
#     )

#     if files:
#         print(f"  📂 Using latest file instead: {files[0]}")
#         return os.path.join(OUTPUT_DIR, files[0])

#     raise FileNotFoundError(
#         f"❌ No 'Phone No Record' Excel file found in {OUTPUT_DIR}\n"
#         f"   Expected: Phone No Record - {yesterday}.xlsx"
#     )



# # ── READ EXCEL → HTML TABLES ──────────────────────────────────────────────────

# def excel_to_html(xlsx_path):

#     wb = load_workbook(xlsx_path, data_only=True)

#     ws = wb["Phone No Record Tracker"]



#     # Row 4: column headers + colors

#     headers = []

#     for c in range(1, ws.max_column + 1):

#         v = ws.cell(4, c).value

#         if v:

#             rgb = ws.cell(4, c).fill.fgColor.rgb[2:]

#             headers.append((c, v, "#" + rgb))



#     # CF cols = day cols

#     cf_cols = set()

#     for cf in ws.conditional_formatting:

#         m = re.match(r"([A-Z]+)", str(cf).strip("<>ConditionalFormatting "))

#         if m:

#             cf_cols.add(column_index_from_string(m.group(1)))



#     # Row 3 group spans

#     group_spans = {}

#     for mr in ws.merged_cells.ranges:

#         m = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", str(mr))

#         if m and int(m.group(2)) == 3:

#             sc = column_index_from_string(m.group(1))

#             ec = column_index_from_string(m.group(3))

#             lbl = ws.cell(3, sc).value

#             if lbl:

#                 group_spans[sc] = (ec, lbl)



#     city_idx   = next((i for i,(c,v,_) in enumerate(headers) if v=="City"), 0)

#     branch_idx = next((i for i,(c,v,_) in enumerate(headers) if v=="Branch"), 1)

#     fmt_idx    = next((i for i,(c,v,_) in enumerate(headers) if v=="Format"), None)



#     # Data rows

#     data = []

#     for r in range(5, ws.max_row + 1):

#         row = [ws.cell(r, c).value for c in range(1, len(headers) + 1)]

#         if not any(v is not None for v in row): continue

#         if row[city_idx] in (None, "", "Unknown"): continue

#         data.append(row)



#     title = ws["A1"].value or "Phone No Record Tracker"



#     # ── Desktop helpers ───────────────────────────────────────────────────────

#     BASE = ("padding:6px 8px; font-size:11px; font-weight:400; color:#1A1A1A;"

#             f" border-bottom:1px solid {BORD}; border-right:1px solid {BORD};"

#             " vertical-align:middle; white-space:nowrap;")

#     NUM  = BASE + " text-align:center;"

#     TXT  = BASE + " text-align:left;"



#     def render_val_td(c, name, val, bg, extra_style=""):

#         w_map = {"City": "90px", "Branch": "150px", "Format": "72px",

#                  "Hist High": "62px", "__week__": "58px", "__day__": "58px"}

#         w = w_map.get(name, "58px")

#         ws_ = f"width:{w}; min-width:{w};"

#         if name == "City":

#             return f'<td style="{TXT}{ws_} background:{bg}; {extra_style}">{val or ""}</td>'

#         elif name == "Branch":

#             return f'<td style="{TXT}{ws_} background:{bg}; {extra_style} max-width:150px; overflow:hidden; text-overflow:ellipsis;">{val or ""}</td>'

#         elif name == "Format":

#             return f'<td style="{BASE}{ws_} background:{bg}; text-align:center; font-size:10px; color:#666; {extra_style}">{val or ""}</td>'

#         elif name == "Hist High":

#             txt = "—" if val is None else f"{float(val)*100:.1f}%"

#             clr = "#BBB" if val is None else "#1A1A1A"

#             return f'<td style="{NUM}{ws_} background:{bg}; color:{clr}; border-left:2px solid {BORD}; {extra_style}">{txt}</td>'

#         elif c in cf_cols:

#             if val is None:

#                 return f'<td style="{NUM}{ws_} background:{RLIGHT}; color:#BBB; {extra_style}">—</td>'

#             # ── RED = low, WHITE = high ───────────────────────────────────────
#             cbg = pct_to_bg(float(val))

#             return f'<td style="{NUM}{ws_} background:{cbg}; {extra_style}">{float(val)*100:.1f}%</td>'

#         else:

#             txt = "—" if val is None else f"{float(val)*100:.1f}%"

#             clr = "#BBB" if val is None else "#1A1A1A"

#             return f'<td style="{NUM}{ws_} background:{bg}; color:{clr}; {extra_style}">{txt}</td>'



#     def th(label, bg, width, extra=""):

#         return (f'<th style="width:{width}; min-width:{width};'

#                 f' padding:7px 8px; font-size:10px; font-weight:700; background:{bg};'

#                 f' color:{WHITE}; text-align:center; white-space:nowrap;'

#                 f' border-right:1px solid rgba(255,255,255,0.18); {extra}">{label}</th>')



#     # ════════════════════════════════════════════════════════
#     #  DESKTOP TABLE
#     # ════════════════════════════════════════════════════════

#     def build_desktop_table():

#         group_row = ""

#         covered = set()

#         for c in range(1, len(headers) + 1):

#             if c in covered: continue

#             if c in group_spans:

#                 ec, lbl = group_spans[c]

#                 span = ec - c + 1

#                 bg   = WBLUE if "Week" in lbl else ROYAL

#                 group_row += (f'<th colspan="{span}" style="background:{bg}; color:{WHITE};'

#                               f' font-size:9px; font-weight:700; text-align:center;'

#                               f' padding:3px 8px; border-right:1px solid rgba(255,255,255,0.18);">{lbl}</th>')

#                 for x in range(c, ec+1): covered.add(x)

#             else:

#                 covered.add(c)

#                 group_row += f'<th style="background:{ROYAL}; padding:3px;"></th>'



#         col_row = ""

#         for c, v, hbg in headers:

#             extra = "border-left:2px solid rgba(255,255,255,0.35);" if v == "Hist High" else ""

#             label = "Hist&nbsp;High" if v == "Hist High" else v

#             w_map = {"City":"90px","Branch":"150px","Format":"72px","Hist High":"62px"}

#             w = w_map.get(v, "58px")

#             col_row += th(label, hbg, w, extra)



#         rows = ""

#         for ri, row in enumerate(data):

#             bg = RLIGHT if ri % 2 == 0 else RALT

#             rows += "<tr>"

#             for ci, (c, name, _) in enumerate(headers):

#                 rows += render_val_td(c, name, row[ci], bg)

#             rows += "</tr>"



#         return (f'<div style="overflow-x:auto; -webkit-overflow-scrolling:touch;">'

#                 f'<table style="border-collapse:collapse; font-family:Arial,sans-serif;'

#                 f' table-layout:auto; border:1px solid {BORD}; min-width:600px;">'

#                 f'<thead><tr>{group_row}</tr><tr>{col_row}</tr></thead>'

#                 f'<tbody>{rows}</tbody></table></div>')



#     # ════════════════════════════════════════════════════════
#     #  MOBILE TABLE — fixed width, NO scroll, ALL columns
#     # ════════════════════════════════════════════════════════

#     def build_mobile_table():

#         week_hdrs = [(ci, c, v, hbg) for ci,(c,v,hbg) in enumerate(headers)
#                      if v not in ("City","Branch","Format","Hist High") and c not in cf_cols]

#         day_hdrs  = [(ci, c, v, hbg) for ci,(c,v,hbg) in enumerate(headers) if c in cf_cols]

#         hist_hdrs = [(ci, c, v, hbg) for ci,(c,v,hbg) in enumerate(headers) if v == "Hist High"]

#         n_week = len(week_hdrs)
#         n_day  = len(day_hdrs)
#         n_hist = len(hist_hdrs)

#         W = {"cityfmt": 38, "branch": 118, "week": 28, "day": 28, "hh": 26}

#         total_w = W["cityfmt"] + W["branch"] + n_week*W["week"] + n_day*W["day"] + n_hist*W["hh"]

#         CB = (f"border-bottom:1px solid {BORD}; border-right:1px solid {BORD};"
#               f" vertical-align:middle; padding:3px 1px; overflow:hidden;"
#               f" white-space:nowrap;")

#         def mth(label, bg, w_px, extra=""):
#             return (f'<th style="width:{w_px}px; min-width:{w_px}px; max-width:{w_px}px;'
#                     f' padding:3px 1px; font-size:7.5px; font-weight:700; background:{bg};'
#                     f' color:{WHITE}; text-align:center; overflow:hidden;'
#                     f' border-right:1px solid rgba(255,255,255,0.2); {extra}">{label}</th>')

#         def mgth(label, bg, span, extra=""):
#             return (f'<th colspan="{span}" style="padding:2px 1px; font-size:7px; font-weight:700;'
#                     f' background:{bg}; color:{WHITE}; text-align:center;'
#                     f' border-right:1px solid rgba(255,255,255,0.25); {extra}">{label}</th>')

#         group_row = (
#             mgth("", ROYAL, 2) +
#             (mgth("&#8592; Wks &#8594;", WBLUE, n_week) if n_week else "") +
#             (mgth("&#8592; Days &#8594;", ROYAL, n_day)  if n_day  else "") +
#             (mgth("HH", ROYAL, n_hist)                   if n_hist else "")
#         )

#         col_row = (
#             mth("City/Fmt", ROYAL, W["cityfmt"]) +
#             mth("Outlet",   ROYAL, W["branch"], "text-align:left;")
#         )
#         for _, _, v, hbg in week_hdrs:
#             col_row += mth(v, hbg, W["week"])
#         for _, _, v, hbg in day_hdrs:
#             col_row += mth(v, hbg, W["day"])
#         for _, _, v, hbg in hist_hdrs:
#             col_row += mth("HH", hbg, W["hh"])

#         rows = ""

#         for ri, row in enumerate(data):

#             bg = RLIGHT if ri % 2 == 0 else RALT

#             city_val   = short_city(row[city_idx])
#             branch_val = str(row[branch_idx] or "")
#             fmt_val    = short_format(row[fmt_idx]) if fmt_idx is not None else ""

#             rows += "<tr>"

#             rows += (f'<td style="{CB} width:{W["cityfmt"]}px; max-width:{W["cityfmt"]}px;'
#                      f' background:{bg}; text-align:center; padding:2px 1px; line-height:1.2;">'
#                      f'<span style="display:block; font-size:8px; font-weight:700; color:{ROYAL};">{city_val}</span>'
#                      f'<span style="display:block; font-size:7px; color:#888;">{fmt_val}</span>'
#                      f'</td>')

#             rows += (f'<td style="{CB} width:{W["branch"]}px; max-width:{W["branch"]}px;'
#                      f' background:{bg}; font-size:8px; font-weight:500; text-align:left;'
#                      f' padding-left:4px; white-space:normal; line-height:1.25;">{branch_val}</td>')

#             for ci, c, name, _ in week_hdrs:
#                 val = row[ci]
#                 txt = "—" if val is None else f"{float(val)*100:.0f}%"
#                 clr = "#BBB" if val is None else "#1A1A1A"
#                 rows += (f'<td style="{CB} width:{W["week"]}px; max-width:{W["week"]}px;'
#                          f' background:{bg}; font-size:8px; text-align:center; color:{clr};">{txt}</td>')

#             # ── Day cols: RED = low, WHITE = high ─────────────────────────────
#             for ci, c, name, _ in day_hdrs:
#                 val = row[ci]
#                 if val is None:
#                     rows += (f'<td style="{CB} width:{W["day"]}px; max-width:{W["day"]}px;'
#                              f' background:{RLIGHT}; font-size:8px; text-align:center; color:#BBB;">—</td>')
#                 else:
#                     cbg = pct_to_bg(float(val))
#                     rows += (f'<td style="{CB} width:{W["day"]}px; max-width:{W["day"]}px;'
#                              f' background:{cbg}; font-size:8px; text-align:center; color:#1A1A1A;">'
#                              f'{float(val)*100:.0f}%</td>')

#             for ci, c, name, _ in hist_hdrs:
#                 val = row[ci]
#                 txt = "—" if val is None else f"{float(val)*100:.0f}%"
#                 clr = "#BBB" if val is None else "#1A1A1A"
#                 rows += (f'<td style="{CB} width:{W["hh"]}px; max-width:{W["hh"]}px;'
#                          f' background:{bg}; font-size:8px; text-align:center; color:{clr};'
#                          f' border-left:1px solid {BORD};">{txt}</td>')

#             rows += "</tr>"

#         return (
#             f'<table style="border-collapse:collapse; font-family:Arial,sans-serif;'
#             f' table-layout:fixed; width:{total_w}px; border:1px solid {BORD};">'
#             f'<thead><tr>{group_row}</tr><tr>{col_row}</tr></thead>'
#             f'<tbody>{rows}</tbody>'
#             f'</table>'
#         )



#     desktop = build_desktop_table()

#     mobile  = build_mobile_table()

#     return desktop, mobile, title



# # ── ATTACH FILE ───────────────────────────────────────────────────────────────

# def attach_file(msg, path):

#     with open(path, "rb") as f:

#         part = MIMEBase("application", "octet-stream")

#         part.set_payload(f.read())

#     encoders.encode_base64(part)

#     part.add_header("Content-Disposition", f'attachment; filename="{os.path.basename(path)}"')

#     msg.attach(part)



# # ── MAIN ──────────────────────────────────────────────────────────────────────

# if __name__ == "__main__":

#     excel_file = get_excel_file()

#     print(f"Using: {excel_file}")



#     desktop_table, mobile_table, title = excel_to_html(excel_file)

#     # Use the file date in the subject, not today
#     file_date_str = os.path.basename(excel_file).replace("Phone No Record - ", "").replace(".xlsx", "")
#     try:
#         file_date = datetime.strptime(file_date_str, "%Y-%m-%d").strftime("%d %b %Y")
#     except Exception:
#         file_date = datetime.today().strftime("%d %b %Y")

#     subject = f"Phone No Record Tracker - {file_date}"



#     html = f"""<!DOCTYPE html>
# <html xmlns:v="urn:schemas-microsoft-com:vml" xmlns:o="urn:schemas-microsoft-com:office:office">
# <head>
# <meta name="viewport" content="width=device-width, initial-scale=1.0">
# <!--[if mso]>
# <style type="text/css">
#   .mobile-view {{ display:none !important; mso-hide:all !important; }}
# </style>
# <![endif]-->
# <style>
#   /* Default (desktop email clients that read CSS) */
#   .desktop-view {{ display:block; }}
#   .mobile-view  {{ display:none; mso-hide:all; }}
#   /* Small screens: flip visibility */
#   @media only screen and (max-width: 600px) {{
#     .desktop-view {{ display:none !important; mso-hide:all !important; }}
#     .mobile-view  {{ display:block !important; mso-hide:none !important; }}
#     body {{ padding:4px !important; margin:0 !important; }}
#   }}
# </style>
# </head>
# <body style="margin:0; padding:16px; background:#EEF2F9; font-family:Arial,sans-serif;">

#   <!-- ░░ HEADER ░░ -->
#   <div style="background:{ROYAL}; padding:14px 16px; border-radius:4px 4px 0 0;">
#     <div style="font-size:14px; font-weight:700; color:{WHITE};">{title}</div>
#     <div style="font-size:10px; color:rgba(255,255,255,0.6); margin-top:2px;">
#       Data for: {file_date}
#     </div>
#   </div>

#   <!-- ░░ DESKTOP (hidden on mobile via media query) ░░ -->
#   <div class="desktop-view" style="background:{WHITE}; border:1px solid {BORD};
#        border-top:none; padding:10px 16px 14px;">
#     <div style="font-size:10px; color:#2C5282; background:#EEF2F9;
#          border-left:3px solid {ROYAL}; padding:7px 12px; margin-bottom:10px;
#          border-radius:0 3px 3px 0; font-weight:500;">
#       Note: Outlets are sorted basis number of offline orders (high to low) &nbsp;|&nbsp;
#     </div>
#     {desktop_table}
#   </div>

#   <!--[if !mso]><!-->
#   <!-- ░░ MOBILE: hidden on Outlook desktop via mso-hide:all ░░ -->
#   <div class="mobile-view" style="display:none; mso-hide:all; background:{WHITE};
#        border:1px solid {BORD}; border-top:none; padding:6px 4px 8px;">
#     <div style="font-size:9px; color:#2C5282; background:#EEF2F9;
#          border-left:3px solid {ROYAL}; padding:5px 8px; margin-bottom:6px;
#          border-radius:0 3px 3px 0; font-weight:500;">
#       Note: Outlets sorted by offline orders (high to low) &nbsp;|&nbsp;
#       🔴 Red = low &nbsp; ⬜ White = high
#     </div>
#     {mobile_table}
#   </div>
#   <!--<![endif]-->

#   <!-- ░░ FOOTER ░░ -->
#   <div style="background:{WHITE}; border:1px solid {BORD}; border-top:none;
#        padding:8px 16px 10px; border-radius:0 0 4px 4px;">
#     <div style="font-size:10px; color:#AAAAAA;">
#       Capture Rate = unique valid phones &divide; unique invoices &nbsp;|&nbsp;
#       Generated {datetime.now().strftime('%d %b %Y, %I:%M %p')}
#     </div>
#   </div>

# </body>
# </html>"""



#     alt   = MIMEMultipart("alternative")

#     alt.attach(MIMEText(html, "html"))

#     mixed = MIMEMultipart("mixed")

#     mixed["From"]    = SENDER_EMAIL

#     mixed["To"]      = ", ".join(TO_RECIPIENTS)

#     if CC_RECIPIENTS: mixed["CC"] = ", ".join(CC_RECIPIENTS)

#     mixed["Subject"] = subject

#     mixed.attach(alt)

#     attach_file(mixed, excel_file)



#     print(f"Connecting to {SMTP_HOST}:{SMTP_PORT} ...")

#     ctx = ssl.create_default_context()

#     try:

#         with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as s:

#             s.ehlo(); s.starttls(context=ctx); s.ehlo()

#             s.login(SENDER_EMAIL, SENDER_PASSWORD)

#             s.sendmail(SENDER_EMAIL,

#                        TO_RECIPIENTS + CC_RECIPIENTS + BCC_RECIPIENTS,

#                        mixed.as_string())

#         print(f"✅ Sent to {', '.join(TO_RECIPIENTS)}")

#     except smtplib.SMTPAuthenticationError:

#         print("❌ Auth failed — check App Password.")

#     except Exception as e:

#         print(f"❌ Error: {e}")



"""

python send_report.py

Auto-picks yesterday's Excel → reads it → sends responsive HTML email

Desktop: full scrollable table
Mobile:  fixed-width no-scroll table, all columns, abbreviated labels

"""

import os, re, smtplib, ssl

from datetime import datetime, timedelta

from email.mime.multipart import MIMEMultipart

from email.mime.base import MIMEBase

from email.mime.text import MIMEText

from email import encoders

from openpyxl import load_workbook

from openpyxl.utils import column_index_from_string



# ── CONFIG ────────────────────────────────────────────────────────────────────

OUTPUT_DIR      = r"C:/Users/jeryy/OneDrive - BLISS CHOCOLATES INDIA PRIVATE LIMITED/Desktop/Documents/Phone No/output"

SMTP_HOST       = "smtp.office365.com"

SMTP_PORT       = 587

SENDER_EMAIL    = "finance.reports@smoorchocolates.com"

SENDER_PASSWORD = "vnnyxwnxypqhxdqj"

TO_RECIPIENTS   = [   
    
    #  "muskan.sajwan@smoorchocolates.com",
    # "trisha.krishna@smoorchocolates.com",
    # "kanchan.achpal@smoorchocolates.com",		

    # "vedant.gote@smoorchocolates.com",
    # "adeshshetty@smoorchocolates.com",
    # "mehtaj.khan@smoorchocolates.com",
    # "ranbir.singh@smoorchocolates.com",
    # "jagdish.duggal@smoorchocolates.com",
    # "hussain@smoorchocolates.com",

    # "suresh.patro@smoorchocolates.com",
    # "ashwin@smoorchocolates.com",

    # "deependra.rao@smoorchocolates.com",
    "john.jerry@smoorchocolates.com",
    # "siddhartha.siva@smoorchocolates.com",
    # "rahul.kumar@smoorchocolates.com" 
    ]

CC_RECIPIENTS   = []

BCC_RECIPIENTS  = []



# ── COLORS ────────────────────────────────────────────────────────────────────

ROYAL  = "#1A3A6B"

WBLUE  = "#2C5282"

WHITE  = "#FFFFFF"

RLIGHT = "#F7F9FC"

RALT   = "#EEF2F9"

BORD   = "#C5CFE3"



# ── ABBREVIATION MAPS ─────────────────────────────────────────────────────────

CITY_SHORT = {
    "Bangalore": "BLR", "Bengaluru": "BLR",
    "Mumbai":    "MUM", "Bombay":    "MUM",
    "Chennai":   "CHN", "Madras":    "CHN",
    "Delhi":     "NCR", "New Delhi": "NCR", "NCR": "NCR",
    "Gurgaon":   "NCR", "Noida":     "NCR", "Delhi NCR": "NCR",
    "Pune":      "PUN",
    "Hyderabad": "HYD",
    "Kolkata":   "KOL",
    "Ahmedabad": "AMD",
}

FORMAT_SHORT = {
    "Signature":     "Sig",
    "Cafe":          "Caf", "Café": "Caf",
    "Kiosk":         "Kio",
    "Choco Kitchen": "CK",  "ChocKitchen": "CK", "Choco-Kitchen": "CK",
    "Cloud Kitchen": "CK",  "CloudKitchen": "CK", "Cloud-Kitchen": "CK",
    "Airport":       "Air",
}

def short_city(val):
    if not val: return ""
    s = str(val).strip()
    return CITY_SHORT.get(s, s[:3].upper())

def short_format(val):
    if not val: return ""
    s = str(val).strip()
    if s in FORMAT_SHORT: return FORMAT_SHORT[s]
    for k, v in FORMAT_SHORT.items():
        if k.lower() in s.lower():
            return v
    return s[:3]



# ── COLOR SCALE: red=0% (lowest) → white=100% (highest) ──────────────────────
# Low capture rate = bad = RED
# High capture rate = good = WHITE

def pct_to_bg(val):
    # t=0 → light red (#F8696B), t=1 → white (#FFFFFF)
    t = max(0.0, min(1.0, float(val)))
    r = int(248 + (255 - 248) * t)   # 248 → 255
    g = int(105 + (255 - 105) * t)   # 105 → 255
    b = int(107 + (255 - 107) * t)   # 107 → 255
    return f"#{r:02X}{g:02X}{b:02X}"



# ── FIND YESTERDAY'S FILE ─────────────────────────────────────────────────────

def get_excel_file():
    # Always try yesterday first
    yesterday     = (datetime.today() - timedelta(days=1)).strftime("%Y-%m-%d")
    yesterday_path = os.path.join(OUTPUT_DIR, f"Phone No Record - {yesterday}.xlsx")

    if os.path.exists(yesterday_path):
        print(f"  ✅ Found yesterday's file: Phone No Record - {yesterday}.xlsx")
        return yesterday_path

    # Fallback: latest available file
    print(f"  ⚠️  Yesterday's file not found: Phone No Record - {yesterday}.xlsx")
    print(f"  🔍 Searching for latest available file in: {OUTPUT_DIR}")

    files = sorted(
        [f for f in os.listdir(OUTPUT_DIR)
         if f.startswith("Phone No Record") and f.endswith(".xlsx")],
        reverse=True
    )

    if files:
        print(f"  📂 Using latest file instead: {files[0]}")
        return os.path.join(OUTPUT_DIR, files[0])

    raise FileNotFoundError(
        f"❌ No 'Phone No Record' Excel file found in {OUTPUT_DIR}\n"
        f"   Expected: Phone No Record - {yesterday}.xlsx"
    )



# ── READ EXCEL → HTML TABLES ──────────────────────────────────────────────────

def excel_to_html(xlsx_path):

    wb = load_workbook(xlsx_path, data_only=True)

    ws = wb["Phone No Record Tracker"]



    # Row 4: column headers + colors

    headers = []

    for c in range(1, ws.max_column + 1):

        v = ws.cell(4, c).value

        if v:

            rgb = ws.cell(4, c).fill.fgColor.rgb[2:]

            headers.append((c, v, "#" + rgb))



    # CF cols = day cols

    cf_cols = set()

    for cf in ws.conditional_formatting:

        m = re.match(r"([A-Z]+)", str(cf).strip("<>ConditionalFormatting "))

        if m:

            cf_cols.add(column_index_from_string(m.group(1)))



    # Row 3 group spans

    group_spans = {}

    for mr in ws.merged_cells.ranges:

        m = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", str(mr))

        if m and int(m.group(2)) == 3:

            sc = column_index_from_string(m.group(1))

            ec = column_index_from_string(m.group(3))

            lbl = ws.cell(3, sc).value

            if lbl:

                group_spans[sc] = (ec, lbl)



    city_idx   = next((i for i,(c,v,_) in enumerate(headers) if v=="City"), 0)

    branch_idx = next((i for i,(c,v,_) in enumerate(headers) if v=="Branch"), 1)

    fmt_idx    = next((i for i,(c,v,_) in enumerate(headers) if v=="Format"), None)



    # Data rows

    data       = []
    total_row  = None

    for r in range(5, ws.max_row + 1):

        row = [ws.cell(r, c).value for c in range(1, len(headers) + 1)]

        if not any(v is not None for v in row): continue

        # Detect TOTAL row — Excel merges cols 1-3 so value lands on col 0 (City)
        if str(row[0] or "").strip().upper() == "TOTAL" or str(row[branch_idx] or "").strip().upper() == "TOTAL":
            total_row = row
            continue

        if row[city_idx] in (None, "", "Unknown"): continue

        data.append(row)



    title = ws["A1"].value or "Phone No Record Tracker"



    # ── Desktop helpers ───────────────────────────────────────────────────────

    BASE = ("padding:6px 8px; font-size:11px; font-weight:400; color:#1A1A1A;"

            f" border-bottom:1px solid {BORD}; border-right:1px solid {BORD};"

            " vertical-align:middle; white-space:nowrap;")

    NUM  = BASE + " text-align:center;"

    TXT  = BASE + " text-align:left;"



    def render_val_td(c, name, val, bg, extra_style=""):

        w_map = {"City": "90px", "Branch": "150px", "Format": "72px",

                 "Hist High": "62px", "__week__": "58px", "__day__": "58px"}

        w = w_map.get(name, "58px")

        ws_ = f"width:{w}; min-width:{w};"

        if name == "City":

            return f'<td style="{TXT}{ws_} background:{bg}; {extra_style}">{val or ""}</td>'

        elif name == "Branch":

            return f'<td style="{TXT}{ws_} background:{bg}; {extra_style} max-width:150px; overflow:hidden; text-overflow:ellipsis;">{val or ""}</td>'

        elif name == "Format":

            return f'<td style="{BASE}{ws_} background:{bg}; text-align:center; font-size:10px; color:#666; {extra_style}">{val or ""}</td>'

        elif name == "Hist High":

            txt = "—" if val is None else f"{float(val)*100:.1f}%"

            clr = "#BBB" if val is None else "#1A1A1A"

            return f'<td style="{NUM}{ws_} background:{bg}; color:{clr}; border-left:2px solid {BORD}; {extra_style}">{txt}</td>'

        elif c in cf_cols:

            if val is None:

                return f'<td style="{NUM}{ws_} background:{RLIGHT}; color:#BBB; {extra_style}">—</td>'

            # ── RED = low, WHITE = high ───────────────────────────────────────
            cbg = pct_to_bg(float(val))

            return f'<td style="{NUM}{ws_} background:{cbg}; {extra_style}">{float(val)*100:.1f}%</td>'

        else:

            txt = "—" if val is None else f"{float(val)*100:.1f}%"

            clr = "#BBB" if val is None else "#1A1A1A"

            return f'<td style="{NUM}{ws_} background:{bg}; color:{clr}; {extra_style}">{txt}</td>'



    def th(label, bg, width, extra=""):

        return (f'<th style="width:{width}; min-width:{width};'

                f' padding:7px 8px; font-size:10px; font-weight:700; background:{bg};'

                f' color:{WHITE}; text-align:center; white-space:nowrap;'

                f' border-right:1px solid rgba(255,255,255,0.18); {extra}">{label}</th>')



    # ════════════════════════════════════════════════════════
    #  DESKTOP TABLE
    # ════════════════════════════════════════════════════════

    def build_desktop_table():

        group_row = ""

        covered = set()

        for c in range(1, len(headers) + 1):

            if c in covered: continue

            if c in group_spans:

                ec, lbl = group_spans[c]

                span = ec - c + 1

                bg   = WBLUE if "Week" in lbl else ROYAL

                group_row += (f'<th colspan="{span}" style="background:{bg}; color:{WHITE};'

                              f' font-size:9px; font-weight:700; text-align:center;'

                              f' padding:3px 8px; border-right:1px solid rgba(255,255,255,0.18);">{lbl}</th>')

                for x in range(c, ec+1): covered.add(x)

            else:

                covered.add(c)

                group_row += f'<th style="background:{ROYAL}; padding:3px;"></th>'



        col_row = ""

        for c, v, hbg in headers:

            extra = "border-left:2px solid rgba(255,255,255,0.35);" if v == "Hist High" else ""

            label = "Hist&nbsp;High" if v == "Hist High" else v

            w_map = {"City":"90px","Branch":"150px","Format":"72px","Hist High":"62px"}

            w = w_map.get(v, "58px")

            col_row += th(label, hbg, w, extra)



        rows = ""

        for ri, row in enumerate(data):

            bg = RLIGHT if ri % 2 == 0 else RALT

            rows += "<tr>"

            for ci, (c, name, _) in enumerate(headers):

                rows += render_val_td(c, name, row[ci], bg)

            rows += "</tr>"

        # ── TOTAL row: Royal Blue bg, white bold text ─────────────────────────
        if total_row is not None:
            TS = (f'padding:6px 8px; font-size:11px; font-weight:700; color:{WHITE};'
                  f' background:{ROYAL}; text-align:center;'
                  f' border-right:1px solid rgba(255,255,255,0.2);'
                  f' border-top:2px solid rgba(255,255,255,0.3);')
            rows += '<tr>'
            rows += f'<td colspan="3" style="{TS}">TOTAL</td>'
            for ci, (c, name, _) in enumerate(headers):
                if ci < 3:
                    continue
                val = total_row[ci]
                txt = "—" if val is None else f"{float(val)*100:.1f}%"
                clr = "rgba(255,255,255,0.4)" if val is None else WHITE
                rows += f'<td style="{TS} color:{clr};">{txt}</td>'
            rows += "</tr>"

        return (f'<div style="overflow-x:auto; -webkit-overflow-scrolling:touch;">'

                f'<table style="border-collapse:collapse; font-family:Arial,sans-serif;'

                f' table-layout:auto; border:1px solid {BORD}; min-width:600px;">'

                f'<thead><tr>{group_row}</tr><tr>{col_row}</tr></thead>'

                f'<tbody>{rows}</tbody></table></div>')



    # ════════════════════════════════════════════════════════
    #  MOBILE TABLE — fixed width, NO scroll, ALL columns
    # ════════════════════════════════════════════════════════

    def build_mobile_table():

        week_hdrs = [(ci, c, v, hbg) for ci,(c,v,hbg) in enumerate(headers)
                     if v not in ("City","Branch","Format","Hist High") and c not in cf_cols]

        day_hdrs  = [(ci, c, v, hbg) for ci,(c,v,hbg) in enumerate(headers) if c in cf_cols]

        hist_hdrs = [(ci, c, v, hbg) for ci,(c,v,hbg) in enumerate(headers) if v == "Hist High"]

        n_week = len(week_hdrs)
        n_day  = len(day_hdrs)
        n_hist = len(hist_hdrs)

        W = {"cityfmt": 38, "branch": 118, "week": 28, "day": 28, "hh": 26}

        total_w = W["cityfmt"] + W["branch"] + n_week*W["week"] + n_day*W["day"] + n_hist*W["hh"]

        CB = (f"border-bottom:1px solid {BORD}; border-right:1px solid {BORD};"
              f" vertical-align:middle; padding:3px 1px; overflow:hidden;"
              f" white-space:nowrap;")

        def mth(label, bg, w_px, extra=""):
            return (f'<th style="width:{w_px}px; min-width:{w_px}px; max-width:{w_px}px;'
                    f' padding:3px 1px; font-size:7.5px; font-weight:700; background:{bg};'
                    f' color:{WHITE}; text-align:center; overflow:hidden;'
                    f' border-right:1px solid rgba(255,255,255,0.2); {extra}">{label}</th>')

        def mgth(label, bg, span, extra=""):
            return (f'<th colspan="{span}" style="padding:2px 1px; font-size:7px; font-weight:700;'
                    f' background:{bg}; color:{WHITE}; text-align:center;'
                    f' border-right:1px solid rgba(255,255,255,0.25); {extra}">{label}</th>')

        group_row = (
            mgth("", ROYAL, 2) +
            (mgth("&#8592; Wks &#8594;", WBLUE, n_week) if n_week else "") +
            (mgth("&#8592; Days &#8594;", ROYAL, n_day)  if n_day  else "") +
            (mgth("HH", ROYAL, n_hist)                   if n_hist else "")
        )

        col_row = (
            mth("City/Fmt", ROYAL, W["cityfmt"]) +
            mth("Outlet",   ROYAL, W["branch"], "text-align:left;")
        )
        for _, _, v, hbg in week_hdrs:
            col_row += mth(v, hbg, W["week"])
        for _, _, v, hbg in day_hdrs:
            col_row += mth(v, hbg, W["day"])
        for _, _, v, hbg in hist_hdrs:
            col_row += mth("HH", hbg, W["hh"])

        rows = ""

        for ri, row in enumerate(data):

            bg = RLIGHT if ri % 2 == 0 else RALT

            city_val   = short_city(row[city_idx])
            branch_val = str(row[branch_idx] or "")
            fmt_val    = short_format(row[fmt_idx]) if fmt_idx is not None else ""

            rows += "<tr>"

            rows += (f'<td style="{CB} width:{W["cityfmt"]}px; max-width:{W["cityfmt"]}px;'
                     f' background:{bg}; text-align:center; padding:2px 1px; line-height:1.2;">'
                     f'<span style="display:block; font-size:8px; font-weight:700; color:{ROYAL};">{city_val}</span>'
                     f'<span style="display:block; font-size:7px; color:#888;">{fmt_val}</span>'
                     f'</td>')

            rows += (f'<td style="{CB} width:{W["branch"]}px; max-width:{W["branch"]}px;'
                     f' background:{bg}; font-size:8px; font-weight:500; text-align:left;'
                     f' padding-left:4px; white-space:normal; line-height:1.25;">{branch_val}</td>')

            for ci, c, name, _ in week_hdrs:
                val = row[ci]
                txt = "—" if val is None else f"{float(val)*100:.0f}%"
                clr = "#BBB" if val is None else "#1A1A1A"
                rows += (f'<td style="{CB} width:{W["week"]}px; max-width:{W["week"]}px;'
                         f' background:{bg}; font-size:8px; text-align:center; color:{clr};">{txt}</td>')

            # ── Day cols: RED = low, WHITE = high ─────────────────────────────
            for ci, c, name, _ in day_hdrs:
                val = row[ci]
                if val is None:
                    rows += (f'<td style="{CB} width:{W["day"]}px; max-width:{W["day"]}px;'
                             f' background:{RLIGHT}; font-size:8px; text-align:center; color:#BBB;">—</td>')
                else:
                    cbg = pct_to_bg(float(val))
                    rows += (f'<td style="{CB} width:{W["day"]}px; max-width:{W["day"]}px;'
                             f' background:{cbg}; font-size:8px; text-align:center; color:#1A1A1A;">'
                             f'{float(val)*100:.0f}%</td>')

            for ci, c, name, _ in hist_hdrs:
                val = row[ci]
                txt = "—" if val is None else f"{float(val)*100:.0f}%"
                clr = "#BBB" if val is None else "#1A1A1A"
                rows += (f'<td style="{CB} width:{W["hh"]}px; max-width:{W["hh"]}px;'
                         f' background:{bg}; font-size:8px; text-align:center; color:{clr};'
                         f' border-left:1px solid {BORD};">{txt}</td>')

            rows += "</tr>"

        # ── TOTAL row: Royal Blue bg, white bold text ─────────────────────────
        if total_row is not None:
            TOT_CELL = (f"border-top:2px solid rgba(255,255,255,0.3);"
                        f" border-right:1px solid rgba(255,255,255,0.2);"
                        f" vertical-align:middle; padding:3px 1px;")
            rows += f'<tr style="background:{ROYAL};">'
            # City+Format cell → blank
            rows += (f'<td style="{TOT_CELL} width:{W["cityfmt"]}px; max-width:{W["cityfmt"]}px;'
                     f' background:{ROYAL};"></td>')
            # Branch cell → "TOTAL" label
            rows += (f'<td style="{TOT_CELL} width:{W["branch"]}px; max-width:{W["branch"]}px;'
                     f' background:{ROYAL}; font-size:9px; font-weight:700; color:{WHITE};'
                     f' text-align:left; padding-left:4px;">TOTAL</td>')
            for ci, c, name, _ in week_hdrs:
                val = total_row[ci]
                txt = "—" if val is None else f"{float(val)*100:.0f}%"
                clr = "rgba(255,255,255,0.4)" if val is None else WHITE
                rows += (f'<td style="{TOT_CELL} width:{W["week"]}px; max-width:{W["week"]}px;'
                         f' background:{ROYAL}; font-size:8px; font-weight:700;'
                         f' text-align:center; color:{clr};">{txt}</td>')
            for ci, c, name, _ in day_hdrs:
                val = total_row[ci]
                txt = "—" if val is None else f"{float(val)*100:.0f}%"
                clr = "rgba(255,255,255,0.4)" if val is None else WHITE
                rows += (f'<td style="{TOT_CELL} width:{W["day"]}px; max-width:{W["day"]}px;'
                         f' background:{ROYAL}; font-size:8px; font-weight:700;'
                         f' text-align:center; color:{clr};">{txt}</td>')
            for ci, c, name, _ in hist_hdrs:
                val = total_row[ci]
                txt = "—" if val is None else f"{float(val)*100:.0f}%"
                clr = "rgba(255,255,255,0.4)" if val is None else WHITE
                rows += (f'<td style="{TOT_CELL} width:{W["hh"]}px; max-width:{W["hh"]}px;'
                         f' background:{ROYAL}; font-size:8px; font-weight:700;'
                         f' text-align:center; color:{clr};">{txt}</td>')
            rows += "</tr>"

        return (
            f'<table style="border-collapse:collapse; font-family:Arial,sans-serif;'
            f' table-layout:fixed; width:{total_w}px; border:1px solid {BORD};">'
            f'<thead><tr>{group_row}</tr><tr>{col_row}</tr></thead>'
            f'<tbody>{rows}</tbody>'
            f'</table>'
        )



    desktop = build_desktop_table()

    mobile  = build_mobile_table()

    return desktop, mobile, title



# ── ATTACH FILE ───────────────────────────────────────────────────────────────

def attach_file(msg, path):

    with open(path, "rb") as f:

        part = MIMEBase("application", "octet-stream")

        part.set_payload(f.read())

    encoders.encode_base64(part)

    part.add_header("Content-Disposition", f'attachment; filename="{os.path.basename(path)}"')

    msg.attach(part)



# ── MAIN ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":

    excel_file = get_excel_file()

    print(f"Using: {excel_file}")



    desktop_table, mobile_table, title = excel_to_html(excel_file)

    # Use the file date in the subject, not today
    file_date_str = os.path.basename(excel_file).replace("Phone No Record - ", "").replace(".xlsx", "")
    try:
        file_date = datetime.strptime(file_date_str, "%Y-%m-%d").strftime("%d %b %Y")
    except Exception:
        file_date = datetime.today().strftime("%d %b %Y")

    subject = f"Phone No Record Tracker - {file_date}"



    html = f"""<!DOCTYPE html>
<html xmlns:v="urn:schemas-microsoft-com:vml" xmlns:o="urn:schemas-microsoft-com:office:office">
<head>
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<!--[if mso]>
<style type="text/css">
  .mobile-view {{ display:none !important; mso-hide:all !important; }}
</style>
<![endif]-->
<style>
  /* Default (desktop email clients that read CSS) */
  .desktop-view {{ display:block; }}
  .mobile-view  {{ display:none; mso-hide:all; }}
  /* Small screens: flip visibility */
  @media only screen and (max-width: 600px) {{
    .desktop-view {{ display:none !important; mso-hide:all !important; }}
    .mobile-view  {{ display:block !important; mso-hide:none !important; }}
    body {{ padding:4px !important; margin:0 !important; }}
  }}
</style>
</head>
<body style="margin:0; padding:16px; background:#EEF2F9; font-family:Arial,sans-serif;">

  <!-- ░░ HEADER ░░ -->
  <div style="background:{ROYAL}; padding:14px 16px; border-radius:4px 4px 0 0;">
    <div style="font-size:14px; font-weight:700; color:{WHITE};">{title}</div>
    <div style="font-size:10px; color:rgba(255,255,255,0.6); margin-top:2px;">
      Data for: {file_date}
    </div>
  </div>

  <!-- ░░ DESKTOP (hidden on mobile via media query) ░░ -->
  <div class="desktop-view" style="background:{WHITE}; border:1px solid {BORD};
       border-top:none; padding:10px 16px 14px;">
    <div style="font-size:10px; color:#2C5282; background:#EEF2F9;
         border-left:3px solid {ROYAL}; padding:7px 12px; margin-bottom:10px;
         border-radius:0 3px 3px 0; font-weight:500;">
      Note: Outlets are sorted basis number of offline orders (high to low) &nbsp;|&nbsp;
    </div>
    {desktop_table}
  </div>

  <!--[if !mso]><!-->
  <!-- ░░ MOBILE: hidden on Outlook desktop via mso-hide:all ░░ -->
  <div class="mobile-view" style="display:none; mso-hide:all; background:{WHITE};
       border:1px solid {BORD}; border-top:none; padding:6px 4px 8px;">
    <div style="font-size:9px; color:#2C5282; background:#EEF2F9;
         border-left:3px solid {ROYAL}; padding:5px 8px; margin-bottom:6px;
         border-radius:0 3px 3px 0; font-weight:500;">
      Note: Outlets sorted by offline orders (high to low) &nbsp;|&nbsp;
      🔴 Red = low &nbsp; ⬜ White = high
    </div>
    {mobile_table}
  </div>
  <!--<![endif]-->

  <!-- ░░ FOOTER ░░ -->
  <div style="background:{WHITE}; border:1px solid {BORD}; border-top:none;
       padding:8px 16px 10px; border-radius:0 0 4px 4px;">
    <div style="font-size:10px; color:#AAAAAA;">
      Capture Rate = unique valid phones &divide; unique invoices &nbsp;|&nbsp;
      Generated {datetime.now().strftime('%d %b %Y, %I:%M %p')}
    </div>
  </div>

</body>
</html>"""



    alt   = MIMEMultipart("alternative")

    alt.attach(MIMEText(html, "html"))

    mixed = MIMEMultipart("mixed")

    mixed["From"]    = SENDER_EMAIL

    mixed["To"]      = ", ".join(TO_RECIPIENTS)

    if CC_RECIPIENTS: mixed["CC"] = ", ".join(CC_RECIPIENTS)

    mixed["Subject"] = subject

    mixed.attach(alt)

    attach_file(mixed, excel_file)



    print(f"Connecting to {SMTP_HOST}:{SMTP_PORT} ...")

    ctx = ssl.create_default_context()

    try:

        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as s:

            s.ehlo(); s.starttls(context=ctx); s.ehlo()

            s.login(SENDER_EMAIL, SENDER_PASSWORD)

            s.sendmail(SENDER_EMAIL,

                       TO_RECIPIENTS + CC_RECIPIENTS + BCC_RECIPIENTS,

                       mixed.as_string())

        print(f"✅ Sent to {', '.join(TO_RECIPIENTS)}")

    except smtplib.SMTPAuthenticationError:

        print("❌ Auth failed — check App Password.")

    except Exception as e:

        print(f"❌ Error: {e}")